# 词典管理路由。
import json
import os
from contextlib import suppress
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from time import perf_counter
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, StreamingResponse
from loguru import logger
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from starlette.background import BackgroundTask

from ..config import get_config
from ..db import db_cursor, db_stream_cursor
from ..response import success_response
from ..services import dictionary_correction
from .deps import require_auth

router = APIRouter(prefix="/dictionary", tags=["dictionary"])

DICTIONARY_TEMPLATE_HEADERS: Tuple[str, ...] = ("原文 key", "标准译文", "译文变体JSON", "分类", "备注")


class DictionaryCreateRequest(BaseModel):
    termKey: str
    termValue: str
    variantValues: Optional[List[str]] = None
    category: Optional[str] = None
    remark: Optional[str] = None


class DictionaryUpdateRequest(BaseModel):
    termValue: str
    variantValues: Optional[List[str]] = None
    category: Optional[str] = None
    remark: Optional[str] = None


CORRECTION_STATUS_LABELS = {
    dictionary_correction.CORRECTION_STATUS_IDLE: "无需纠错",
    dictionary_correction.CORRECTION_STATUS_PENDING: "待纠错",
    dictionary_correction.CORRECTION_STATUS_RUNNING: "纠错中",
    dictionary_correction.CORRECTION_STATUS_DONE: "已完成",
    dictionary_correction.CORRECTION_STATUS_FAILED: "失败",
}


def _apply_pagination(page: int, page_size: int) -> int:
    if page < 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="page 必须 >= 1")
    if page_size < 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 必须 >= 1")
    return (page - 1) * page_size


def _require_non_empty_text(value: str, field_name: str) -> str:
    cleaned = value.strip()
    if not cleaned:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} 不能为空")
    return cleaned


def _normalize_optional_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned if cleaned else None


def _normalize_variant_values(values: Optional[List[Any]], term_value: str, field_name: str = "译文变体") -> Optional[List[str]]:
    if values is None:
        return None
    if not isinstance(values, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} 必须为字符串数组")

    normalized: List[str] = []
    seen = set()
    for index, item in enumerate(values, start=1):
        if not isinstance(item, str):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"{field_name} 第 {index} 项必须为字符串",
            )
        cleaned = item.strip()
        if not cleaned or cleaned == term_value or cleaned in seen:
            continue
        seen.add(cleaned)
        normalized.append(cleaned)
    return normalized


def _serialize_variant_values(values: Optional[List[str]]) -> Optional[str]:
    if not values:
        return None
    return json.dumps(values, ensure_ascii=False)


def _deserialize_variant_values(value: Any) -> List[str]:
    if value in (None, ""):
        return []
    if isinstance(value, (bytes, bytearray)):
        value = value.decode("utf-8")
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            logger.warning("dictionary variantValues JSON parse failed, fallback to empty list: value={}", value)
            return []
    else:
        parsed = value
    if not isinstance(parsed, list):
        return []
    result: List[str] = []
    seen = set()
    for item in parsed:
        if not isinstance(item, str):
            continue
        cleaned = item.strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def _parse_variant_values_cell(value: Any, row_number: int) -> Optional[List[str]]:
    raw = _parse_optional_str(value)
    if raw is None:
        return None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行字段 译文变体JSON 不是合法 JSON: {error.msg}",
        ) from error
    if not isinstance(parsed, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行字段 译文变体JSON 必须为 JSON 数组",
        )
    for index, item in enumerate(parsed, start=1):
        if not isinstance(item, str):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_number} 行字段 译文变体JSON 第 {index} 项必须为字符串",
            )
    return parsed


def _resolve_correction_updates(
    *,
    term_key: str,
    term_value: str,
    variant_values: Optional[List[str]],
    is_active: bool,
    existing: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    normalized_variants = dictionary_correction.normalize_variant_values(variant_values)
    current_version = int(existing["correctionVersion"]) if existing else 0
    applied_version = int(existing["appliedCorrectionVersion"]) if existing else 0
    current_status = int(existing["correctionStatus"]) if existing else dictionary_correction.CORRECTION_STATUS_IDLE
    previous_term_key = existing["termKey"] if existing else None
    previous_term_value = existing["termValue"] if existing else None
    previous_variants = dictionary_correction.normalize_variant_values(existing["variantValues"]) if existing else []
    previous_is_active = bool(existing["isActive"]) if existing else False

    changed = existing is None or any(
        [
            previous_term_key != term_key,
            previous_term_value != term_value,
            previous_variants != normalized_variants,
            previous_is_active != is_active,
        ]
    )
    resolved = dictionary_correction.should_requeue_correction(
        is_active,
        normalized_variants,
        changed,
        current_version,
        applied_version,
        current_status,
    )
    return {
        "variantValues": normalized_variants,
        **resolved,
    }


def _is_empty_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _parse_required_str(value: Any, field_name: str, row_number: int) -> str:
    if _is_empty_cell(value):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行字段 {field_name} 不能为空")
    return str(value).strip()


def _parse_optional_str(value: Any) -> Optional[str]:
    if _is_empty_cell(value):
        return None
    return str(value).strip()


def _validate_template_header(header_values: List[Any]) -> None:
    if len(header_values) < len(DICTIONARY_TEMPLATE_HEADERS):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传模板缺少必要列")

    expected = list(DICTIONARY_TEMPLATE_HEADERS)
    actual = ["" if value is None else str(value).strip() for value in header_values[: len(DICTIONARY_TEMPLATE_HEADERS)]]
    if actual != expected:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"上传模板表头不匹配，必须为: {'/'.join(DICTIONARY_TEMPLATE_HEADERS)}",
        )

    extra_values = header_values[len(DICTIONARY_TEMPLATE_HEADERS) :]
    for column_index, value in enumerate(extra_values, start=len(DICTIONARY_TEMPLATE_HEADERS) + 1):
        if not _is_empty_cell(value):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"上传模板存在多余表头列: 第 {column_index} 列",
            )


def _load_upload_sheet(file_bytes: bytes):
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空")
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"上传文件解析失败: {error}") from error
    if not workbook.worksheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件缺少工作表")
    return workbook.worksheets[0]


def _cleanup_temp_file(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        return


def _build_dictionary_conditions(
    keyword: Optional[str],
    termKey: Optional[str],
    termValue: Optional[str],
    category: Optional[str],
    isActive: Optional[bool],
) -> Tuple[List[str], List[Any]]:
    conditions: List[str] = []
    params: List[Any] = []

    if keyword is not None:
        keyword_value = f"%{keyword}%"
        conditions.append('(de."termKey" LIKE %s OR de."termValue" LIKE %s)')
        params.append(keyword_value)
        params.append(keyword_value)
    if termKey is not None:
        conditions.append('de."termKey" LIKE %s')
        params.append(f"%{termKey}%")
    if termValue is not None:
        conditions.append('de."termValue" LIKE %s')
        params.append(f"%{termValue}%")
    if category is not None:
        conditions.append("de.category = %s")
        params.append(category)
    if isActive is not None:
        conditions.append('de."isActive" = %s')
        params.append(isActive)

    return conditions, params


@router.get("")
def list_dictionary(
    keyword: Optional[str] = None,
    termKey: Optional[str] = None,
    termValue: Optional[str] = None,
    category: Optional[str] = None,
    isActive: Optional[bool] = None,
    page: int = 1,
    pageSize: Optional[int] = Query(default=None, alias="pageSize"),
    user: Dict[str, Any] = Depends(require_auth),
):
    """查询词典条目，支持筛选与分页。"""
    logger.info(
        "Dict list: keyword={} termKey={} termValue={} category={} isActive={} page={} pageSize={} userId={}",
        keyword,
        termKey,
        termValue,
        category,
        isActive,
        page,
        pageSize,
        user["userId"],
    )
    config = get_config()
    pagination = config["pagination"]
    default_page_size = pagination["default_page_size"]
    max_page_size = pagination["max_page_size"]

    effective_page_size = pageSize if pageSize is not None else default_page_size
    if effective_page_size > max_page_size:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 超出最大限制")

    offset = _apply_pagination(page, effective_page_size)
    conditions, params = _build_dictionary_conditions(keyword, termKey, termValue, category, isActive)
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM dictionary_entries de
            {where_clause}
            """,
            tuple(params),
        )
        total = cursor.fetchone()["total"]

        cursor.execute(
            f"""
            SELECT
              de.id,
              de."termKey" AS "termKey",
              de."termValue" AS "termValue",
              de."variantValues" AS "variantValues",
              de."correctionVersion" AS "correctionVersion",
              de."appliedCorrectionVersion" AS "appliedCorrectionVersion",
              de."correctionStatus" AS "correctionStatus",
              de."correctionLastStartedAt" AS "correctionLastStartedAt",
              de."correctionLastFinishedAt" AS "correctionLastFinishedAt",
              de."correctionLastError" AS "correctionLastError",
              de."correctionUpdatedTextCount" AS "correctionUpdatedTextCount",
              de.category,
              de.remark,
              de."isActive" AS "isActive",
              de."lastModifiedBy" AS "lastModifiedBy",
              u.username AS "lastModifiedByName",
              de."crtTime" AS "crtTime",
              de."uptTime" AS "uptTime"
            FROM dictionary_entries de
            LEFT JOIN users u ON u.id = de."lastModifiedBy"
            {where_clause}
            ORDER BY de."uptTime" DESC, de.id DESC
            LIMIT %s OFFSET %s
            """,
            tuple(params + [effective_page_size, offset]),
        )
        items = cursor.fetchall()
        for item in items:
            item["variantValues"] = _deserialize_variant_values(item.get("variantValues"))
            item["correctionStatusLabel"] = CORRECTION_STATUS_LABELS.get(item["correctionStatus"], "未知")

    logger.info("Dict list complete: total={} page={} pageSize={} userId={}", total, page, effective_page_size, user["userId"])
    return success_response(
        {
            "items": items,
            "total": total,
            "page": page,
            "pageSize": effective_page_size,
        }
    )


@router.post("")
def create_dictionary(request: DictionaryCreateRequest, user: Dict[str, Any] = Depends(require_auth)):
    """新增词典条目。"""
    term_key = _require_non_empty_text(request.termKey, "原文 key")
    term_value = _require_non_empty_text(request.termValue, "译文 value")
    variant_values = _normalize_variant_values(request.variantValues, term_value)
    category = _normalize_optional_text(request.category)
    remark = _normalize_optional_text(request.remark)
    correction_updates = _resolve_correction_updates(
        term_key=term_key,
        term_value=term_value,
        variant_values=variant_values,
        is_active=True,
        existing=None,
    )

    logger.info("Dict create: termKey={} category={} userId={}", term_key, category, user["userId"])
    with db_cursor() as cursor:
        cursor.execute('SELECT id FROM dictionary_entries WHERE "termKey" = %s', (term_key,))
        if cursor.fetchone() is not None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="原文 key 已存在")

        cursor.execute(
            """
            INSERT INTO dictionary_entries (
              "termKey",
              "termValue",
              "variantValues",
              "correctionVersion",
              "appliedCorrectionVersion",
              "correctionStatus",
              "correctionLastStartedAt",
              "correctionLastFinishedAt",
              "correctionLastError",
              "correctionUpdatedTextCount",
              category,
              remark,
              "isActive",
              "lastModifiedBy",
              "crtTime",
              "uptTime"
            )
            VALUES (%s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 0, %s, %s, TRUE, %s, NOW(), NOW())
            """,
            (
                term_key,
                term_value,
                _serialize_variant_values(correction_updates["variantValues"]),
                correction_updates["correctionVersion"],
                correction_updates["appliedCorrectionVersion"],
                correction_updates["correctionStatus"],
                category,
                remark,
                user["userId"],
            ),
        )
        entry_id = cursor.lastrowid

    logger.info("Dict created: entryId={} termKey={} userId={}", entry_id, term_key, user["userId"])
    return success_response({"id": entry_id})


@router.put("/{entryId}")
def update_dictionary(entryId: int, request: DictionaryUpdateRequest, user: Dict[str, Any] = Depends(require_auth)):
    """更新词典条目。"""
    term_value = _require_non_empty_text(request.termValue, "译文 value")
    variant_values = _normalize_variant_values(request.variantValues, term_value)
    category = _normalize_optional_text(request.category)
    remark = _normalize_optional_text(request.remark)

    logger.info("Dict update: entryId={} category={} userId={}", entryId, category, user["userId"])
    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT
              id,
              "termKey" AS "termKey",
              "termValue" AS "termValue",
              "variantValues" AS "variantValues",
              "isActive" AS "isActive",
              "correctionVersion" AS "correctionVersion",
              "appliedCorrectionVersion" AS "appliedCorrectionVersion",
              "correctionStatus" AS "correctionStatus"
            FROM dictionary_entries
            WHERE id = %s
            """,
            (entryId,),
        )
        entry = cursor.fetchone()
        if entry is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="词条不存在")
        entry["variantValues"] = _deserialize_variant_values(entry.get("variantValues"))
        correction_updates = _resolve_correction_updates(
            term_key=entry["termKey"],
            term_value=term_value,
            variant_values=variant_values,
            is_active=entry["isActive"],
            existing=entry,
        )

        cursor.execute(
            (
                """
                UPDATE dictionary_entries
                SET
                  "termValue" = %s,
                  "variantValues" = %s,
                  "correctionVersion" = %s,
                  "appliedCorrectionVersion" = %s,
                  "correctionStatus" = %s,
                  "correctionLastStartedAt" = NULL,
                  "correctionLastFinishedAt" = NULL,
                  "correctionLastError" = NULL,
                  "correctionUpdatedTextCount" = 0,
                  category = %s,
                  remark = %s,
                  "lastModifiedBy" = %s,
                  "uptTime" = NOW()
                WHERE id = %s
                """
                if correction_updates["resetCorrectionMeta"]
                else """
                UPDATE dictionary_entries
                SET
                  "termValue" = %s,
                  "variantValues" = %s,
                  "correctionVersion" = %s,
                  "appliedCorrectionVersion" = %s,
                  "correctionStatus" = %s,
                  category = %s,
                  remark = %s,
                  "lastModifiedBy" = %s,
                  "uptTime" = NOW()
                WHERE id = %s
                """,
                (
                    term_value,
                    _serialize_variant_values(correction_updates["variantValues"]),
                    correction_updates["correctionVersion"],
                    correction_updates["appliedCorrectionVersion"],
                    correction_updates["correctionStatus"],
                    category,
                    remark,
                    user["userId"],
                    entryId,
                ),
            )
        )

    logger.info("Dict updated: entryId={} termKey={} userId={}", entryId, entry["termKey"], user["userId"])
    return success_response({"id": entryId})


@router.post("/{entryId}/correct")
def correct_dictionary_entry(entryId: int, user: Dict[str, Any] = Depends(require_auth)):
    logger.info("Dict correction requested: entryId={} userId={}", entryId, user["userId"])
    try:
        result = dictionary_correction.run_dictionary_correction(entryId)
    except Exception as error:
        dictionary_correction.mark_dictionary_correction_failed(entryId, str(error))
        logger.exception("Dict correction failed: entryId={} userId={} error={}", entryId, user["userId"], error)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"系统纠错失败: {error}") from error
    return success_response(
        {
            "dictionaryId": result.dictionary_id,
            "matchedTextCount": result.matched_text_count,
            "updatedTextCount": result.updated_text_count,
            "status": result.status,
            "statusLabel": CORRECTION_STATUS_LABELS.get(result.status, "未知"),
            "appliedCorrectionVersion": result.applied_version,
            "startedAt": result.started_at,
            "finishedAt": result.finished_at,
            "error": result.error,
        }
    )


@router.get("/template")
def download_dictionary_template(user: Dict[str, Any] = Depends(require_auth)):
    """下载词典导入模板。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dictionary"
    sheet.append(list(DICTIONARY_TEMPLATE_HEADERS))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    logger.info("download_dictionary_template: userId={}", user["userId"])
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="dictionary_template.xlsx"'},
    )


@router.get("/download")
def download_dictionary(
    keyword: Optional[str] = None,
    termKey: Optional[str] = None,
    termValue: Optional[str] = None,
    category: Optional[str] = None,
    isActive: Optional[bool] = None,
    _: Dict[str, Any] = Depends(require_auth),
):
    """根据筛选条件导出词典数据。"""
    request_started_at = perf_counter()
    logger.info(
        "download_dictionary start: keyword={} termKey={} termValue={} category={} isActive={}",
        keyword,
        termKey,
        termValue,
        category,
        isActive,
    )
    config = get_config()
    text_import_export = config["text_import_export"]
    max_download_rows = text_import_export["max_download_rows"]
    download_fetch_batch_size = text_import_export["download_fetch_batch_size"]
    download_progress_log_every_batches = text_import_export["download_progress_log_every_batches"]
    download_temp_dir = text_import_export["download_temp_dir"]

    if not os.path.isdir(download_temp_dir):
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="导出临时目录不存在")

    conditions, params = _build_dictionary_conditions(keyword, termKey, termValue, category, isActive)
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="dictionary")
    sheet.append(list(DICTIONARY_TEMPLATE_HEADERS))

    export_count = 0
    fetched_row_count = 0
    batch_count = 0
    tmp_path: Optional[str] = None
    try:
        with db_stream_cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                  de."termKey" AS "termKey",
                  de."termValue" AS "termValue",
                  de."variantValues" AS "variantValues",
                  de.category,
                  de.remark
                FROM dictionary_entries de
                {where_clause}
                ORDER BY de."uptTime" DESC, de.id DESC
                """,
                tuple(params),
            )
            while True:
                rows = cursor.fetchmany(download_fetch_batch_size)
                if not rows:
                    break
                batch_count += 1
                fetched_row_count += len(rows)
                for row in rows:
                    export_count += 1
                    if export_count > max_download_rows:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"导出数据量超过限制（{max_download_rows}），请缩小筛选范围后重试",
                        )
                    sheet.append(
                        [
                            row["termKey"],
                            row["termValue"],
                            _serialize_variant_values(_deserialize_variant_values(row.get("variantValues"))),
                            row["category"],
                            row["remark"],
                        ]
                    )
                if batch_count == 1 or batch_count % download_progress_log_every_batches == 0:
                    logger.info(
                        "download_dictionary progress: batch={} batchRows={} fetchedRows={} exportRows={} elapsedSec={:.3f}",
                        batch_count,
                        len(rows),
                        fetched_row_count,
                        export_count,
                        perf_counter() - request_started_at,
                    )

        if export_count == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前筛选条件无可导出数据")

        tmp_file = NamedTemporaryFile(
            prefix="tmp_dictionary_export_",
            suffix=".xlsx",
            dir=download_temp_dir,
            delete=False,
        )
        tmp_path = tmp_file.name
        tmp_file.close()
        workbook.save(tmp_path)
    except Exception:
        logger.exception(
            "download_dictionary failed: fetchedRows={} exportRows={} batches={} elapsedSec={:.3f}",
            fetched_row_count,
            export_count,
            batch_count,
            perf_counter() - request_started_at,
        )
        workbook.close()
        if tmp_path is not None:
            _cleanup_temp_file(tmp_path)
        raise
    workbook.close()

    export_name = f"dictionary_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    logger.info(
        "download_dictionary done: exportName={} exportRows={} fetchedRows={} batches={} elapsedSec={:.3f}",
        export_name,
        export_count,
        fetched_row_count,
        batch_count,
        perf_counter() - request_started_at,
    )
    return FileResponse(
        path=tmp_path,
        filename=export_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(_cleanup_temp_file, tmp_path),
    )


@router.post("/upload")
async def upload_dictionary(
    request: Request,
    fileName: str = Query(..., alias="fileName"),
    user: Dict[str, Any] = Depends(require_auth),
):
    """按模板上传词典，termKey 已存在则覆盖，否则新增。"""
    logger.info("Upload dictionary start: fileName={} userId={}", fileName, user["userId"])
    if not fileName:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件缺少文件名")
    if not fileName.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件必须为 .xlsx 格式")

    config = get_config()
    text_import_export = config["text_import_export"]
    max_upload_rows = text_import_export["max_upload_rows"]

    file_bytes = await request.body()
    sheet = _load_upload_sheet(file_bytes)

    header_rows = list(sheet.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空")
    header_values = [cell.value for cell in header_rows[0]]
    _validate_template_header(header_values)

    parsed_rows: List[Dict[str, Any]] = []
    upload_term_keys: List[str] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row[: len(DICTIONARY_TEMPLATE_HEADERS)])
        if len(cells) < len(DICTIONARY_TEMPLATE_HEADERS):
            cells.extend([None] * (len(DICTIONARY_TEMPLATE_HEADERS) - len(cells)))
        if all(_is_empty_cell(item) for item in cells):
            continue

        extra_cells = list(row[len(DICTIONARY_TEMPLATE_HEADERS) :])
        if any(not _is_empty_cell(item) for item in extra_cells):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_index} 行存在模板外数据，请删除多余列",
            )

        term_key = _parse_required_str(cells[0], "原文 key", row_index)
        term_value = _parse_required_str(cells[1], "标准译文", row_index)
        variant_values = _parse_variant_values_cell(cells[2], row_index)
        upload_term_keys.append(term_key)
        parsed_rows.append(
            {
                "rowNumber": row_index,
                "termKey": term_key,
                "termValue": term_value,
                "variantValues": _normalize_variant_values(variant_values, term_value, "译文变体JSON"),
                "category": _parse_optional_str(cells[3]),
                "remark": _parse_optional_str(cells[4]),
            }
        )

    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件没有可处理的数据行")

    if len(parsed_rows) > max_upload_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"上传行数超限，最大允许 {max_upload_rows} 行",
        )

    if len(set(upload_term_keys)) != len(upload_term_keys):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件存在重复原文 key")

    placeholders = ",".join(["%s"] * len(upload_term_keys))
    created_count = 0
    updated_count = 0
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT id, "termKey" AS "termKey"
            FROM dictionary_entries
            WHERE "termKey" IN ({placeholders})
            """,
            tuple(upload_term_keys),
        )
        existing_rows = cursor.fetchall()
        existing_term_keys = [row["termKey"] for row in existing_rows]
        if len(existing_term_keys) != len(set(existing_term_keys)):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="数据库存在重复原文 key，请先清理后再导入")
        existing_map = {row["termKey"]: row for row in existing_rows}

        for item in parsed_rows:
            existing = existing_map.get(item["termKey"])
            if existing is None:
                correction_updates = _resolve_correction_updates(
                    term_key=item["termKey"],
                    term_value=item["termValue"],
                    variant_values=item["variantValues"],
                    is_active=True,
                    existing=None,
                )
                cursor.execute(
                    """
                    INSERT INTO dictionary_entries (
                      "termKey",
                      "termValue",
                      "variantValues",
                      "correctionVersion",
                      "appliedCorrectionVersion",
                      "correctionStatus",
                      "correctionLastStartedAt",
                      "correctionLastFinishedAt",
                      "correctionLastError",
                      "correctionUpdatedTextCount",
                      category,
                      remark,
                      "isActive",
                      "lastModifiedBy",
                      "crtTime",
                      "uptTime"
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, NULL, NULL, NULL, 0, %s, %s, TRUE, %s, NOW(), NOW())
                    """,
                    (
                        item["termKey"],
                        item["termValue"],
                        _serialize_variant_values(correction_updates["variantValues"]),
                        correction_updates["correctionVersion"],
                        correction_updates["appliedCorrectionVersion"],
                        correction_updates["correctionStatus"],
                        item["category"],
                        item["remark"],
                        user["userId"],
                    ),
                )
                created_count += 1
            else:
                cursor.execute(
                    """
                    SELECT
                      id,
                      "termKey" AS "termKey",
                      "termValue" AS "termValue",
                      "variantValues" AS "variantValues",
                      "isActive" AS "isActive",
                      "correctionVersion" AS "correctionVersion",
                      "appliedCorrectionVersion" AS "appliedCorrectionVersion",
                      "correctionStatus" AS "correctionStatus"
                    FROM dictionary_entries
                    WHERE id = %s
                    """,
                    (existing["id"],),
                )
                existing_entry = cursor.fetchone()
                existing_entry["variantValues"] = _deserialize_variant_values(existing_entry.get("variantValues"))
                correction_updates = _resolve_correction_updates(
                    term_key=existing_entry["termKey"],
                    term_value=item["termValue"],
                    variant_values=item["variantValues"],
                    is_active=existing_entry["isActive"],
                    existing=existing_entry,
                )
                update_sql = (
                    """
                    UPDATE dictionary_entries
                    SET
                      "termValue" = %s,
                      "variantValues" = %s,
                      "correctionVersion" = %s,
                      "appliedCorrectionVersion" = %s,
                      "correctionStatus" = %s,
                      "correctionLastStartedAt" = NULL,
                      "correctionLastFinishedAt" = NULL,
                      "correctionLastError" = NULL,
                      "correctionUpdatedTextCount" = 0,
                      category = %s,
                      remark = %s,
                      "lastModifiedBy" = %s,
                      "uptTime" = NOW()
                    WHERE id = %s
                    """
                    if correction_updates["resetCorrectionMeta"]
                    else """
                    UPDATE dictionary_entries
                    SET
                      "termValue" = %s,
                      "variantValues" = %s,
                      "correctionVersion" = %s,
                      "appliedCorrectionVersion" = %s,
                      "correctionStatus" = %s,
                      category = %s,
                      remark = %s,
                      "lastModifiedBy" = %s,
                      "uptTime" = NOW()
                    WHERE id = %s
                    """
                )
                update_params = (
                    item["termValue"],
                    _serialize_variant_values(correction_updates["variantValues"]),
                    correction_updates["correctionVersion"],
                    correction_updates["appliedCorrectionVersion"],
                    correction_updates["correctionStatus"],
                    item["category"],
                    item["remark"],
                    user["userId"],
                    existing["id"],
                )
                cursor.execute(update_sql, update_params)
                updated_count += 1

    logger.info(
        "Upload dictionary complete: fileName={} createdCount={} updatedCount={} userId={}",
        fileName,
        created_count,
        updated_count,
        user["userId"],
    )
    return success_response({"createdCount": created_count, "updatedCount": updated_count})
