# 主文本列表与详情路由。
import os
import threading
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from time import perf_counter
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from loguru import logger
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from starlette.background import BackgroundTask

from ..config import get_config
from ..db import db_cursor, db_stream_cursor
from ..response import success_response
from .deps import require_auth

router = APIRouter(prefix="/texts", tags=["texts"])


TEXT_TEMPLATE_HEADERS: Tuple[str, ...] = ("编号", "FID", "TextId", "Part", "原文", "译文", "状态")
PACKAGE_HEADERS: Tuple[str, ...] = ("fid", "translation")
_EXCEL_CELL_CHAR_LIMIT = 32767
_package_download_lock = threading.Semaphore(1)
STATUS_LABEL_TO_VALUE: Dict[str, int] = {"新增": 1, "修改": 2, "已完成": 3}
STATUS_VALUE_TO_LABEL: Dict[int, str] = {value: label for label, value in STATUS_LABEL_TO_VALUE.items()}
STATUS_VALUE_SET = {1, 2, 3}
TEXT_MATCH_MODE_SET = {"fuzzy", "exact"}


def _apply_pagination(page: int, page_size: int) -> int:
    if page < 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="page 必须 >= 1")
    if page_size < 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 必须 >= 1")
    return (page - 1) * page_size


def _parse_text_match_mode(value: Optional[str], field_name: str) -> str:
    if value is None or value == "":
        return "fuzzy"
    if value not in TEXT_MATCH_MODE_SET:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{field_name} 必须为 fuzzy/exact",
        )
    return value


def _build_text_match_clause(column_sql: str, keyword: str, match_mode: str) -> Tuple[str, str]:
    if match_mode == "exact":
        return f"{column_sql} = %s", keyword
    return f"{column_sql} LIKE %s", f"%{keyword}%"


def _is_empty_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _parse_required_int(value: Any, field_name: str, row_number: int) -> int:
    if _is_empty_cell(value):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行字段 {field_name} 不能为空")
    if isinstance(value, bool):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行字段 {field_name} 类型错误，必须为整数",
        )
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit():
            return int(stripped)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"第 {row_number} 行字段 {field_name} 类型错误，必须为整数",
    )


def _parse_status(value: Any, row_number: int) -> int:
    if _is_empty_cell(value):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行字段 状态 不能为空")
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in STATUS_LABEL_TO_VALUE:
            return STATUS_LABEL_TO_VALUE[stripped]
        if stripped.isdigit():
            status_value = int(stripped)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_number} 行字段 状态 非法，必须为 1/2/3 或 新增/修改/已完成",
            )
    else:
        status_value = _parse_required_int(value, "状态", row_number)

    if status_value not in STATUS_VALUE_SET:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_number} 行字段 状态 非法，必须为 1/2/3 或 新增/修改/已完成",
        )
    return status_value


def _format_status_label(status_value: Any) -> str:
    if isinstance(status_value, bool):
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="状态字段类型异常")
    if not isinstance(status_value, int):
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="状态字段类型异常")
    if status_value not in STATUS_VALUE_TO_LABEL:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="状态枚举缺失映射")
    return STATUS_VALUE_TO_LABEL[status_value]


def _parse_required_str(value: Any, field_name: str, row_number: int) -> str:
    if _is_empty_cell(value):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_number} 行字段 {field_name} 不能为空")
    return str(value)


def _normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _validate_template_header(header_values: List[Any]) -> None:
    if len(header_values) < len(TEXT_TEMPLATE_HEADERS):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传模板缺少必要列")

    expected = list(TEXT_TEMPLATE_HEADERS)
    actual = ["" if value is None else str(value).strip() for value in header_values[: len(TEXT_TEMPLATE_HEADERS)]]
    if actual != expected:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"上传模板表头不匹配，必须为: {'/'.join(TEXT_TEMPLATE_HEADERS)}",
        )

    extra_values = header_values[len(TEXT_TEMPLATE_HEADERS) :]
    for column_index, value in enumerate(extra_values, start=len(TEXT_TEMPLATE_HEADERS) + 1):
        if not _is_empty_cell(value):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"上传模板存在多余表头列: 第 {column_index} 列",
            )


def _load_upload_sheet(file_bytes: bytes):
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空")
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"上传文件解析失败: {error}") from error
    if not workbook.worksheets:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件缺少工作表")
    return workbook.worksheets[0]


def _cleanup_temp_file(path: str) -> None:
    try:
        os.remove(path)
    except FileNotFoundError:
        return


def _format_part_range(parts: List[int]) -> str:
    """将 part 列表压缩为范围字符串，如 [1,2,3,5,6] -> '1-3,5-6'，单个 part 输出 'n-n'。"""
    if not parts:
        return ""
    sorted_parts = sorted(parts)
    ranges = []
    start = sorted_parts[0]
    end = sorted_parts[0]
    for p in sorted_parts[1:]:
        if p == end + 1:
            end = p
        else:
            ranges.append(f"{start}-{end}")
            start = p
            end = p
    ranges.append(f"{start}-{end}")
    return ",".join(ranges)


def _split_translation_into_rows(fid: str, translation: str) -> List[List[str]]:
    """将超长 translation 按 ||| segment 边界拆分为多行写入 xlsx。
    segment（textId::::::[text]）为最小原子单位，绝不在内部截断。
    非末行末尾追加 '|||' 表示续行；单个 segment 超过 limit 时独占一行（完整保留）。
    """
    segments = translation.split("|||")
    rows: List[List[str]] = []
    current_parts: List[str] = []
    current_len = 0
    sep_len = 3  # len("|||")

    for seg in segments:
        seg_len = len(seg)
        if not current_parts:
            current_parts.append(seg)
            current_len = seg_len
        else:
            if current_len + sep_len + seg_len > _EXCEL_CELL_CHAR_LIMIT - sep_len:
                rows.append([fid, "|||".join(current_parts) + "|||"])
                current_parts = [seg]
                current_len = seg_len
            else:
                current_parts.append(seg)
                current_len += sep_len + seg_len

    if current_parts:
        rows.append([fid, "|||".join(current_parts)])

    return rows


def _merge_fid_rows(fid_rows: List[Dict[str, Any]]) -> Tuple[str, str]:
    """合并同一 fid 的多个 part 为一行，还原分段协议格式。
    - textId 不含 ':::' → textId::::::[text]   (格式1)
    - textId 含 ':::'   → textId:::[text]       (格式2/3)
    空译文取原文填充。返回 (fid, translation)。
    """
    fid = fid_rows[0]["fid"]
    translated_segments = []
    for row in fid_rows:
        text_id = str(row["textId"])
        source = row["sourceText"] or ""
        translated = row["translatedText"] or source
        if ":::" in text_id:
            translated_segments.append(f"{text_id}:::[{translated}]")
        else:
            translated_segments.append(f"{text_id}::::::[{translated}]")

    return fid, "|||".join(translated_segments)


def _build_download_conditions(
    fid: Optional[str],
    textId: Optional[str],
    status_filter: Optional[int],
    sourceKeyword: Optional[str],
    sourceMatchMode: str,
    translatedKeyword: Optional[str],
    translatedMatchMode: str,
    updatedFrom: Optional[str],
    updatedTo: Optional[str],
    claimer: Optional[str],
    claimed: Optional[bool],
) -> Tuple[List[str], List[Any]]:
    conditions: List[str] = []
    params: List[Any] = []

    if fid is not None:
        conditions.append("tm.fid = %s")
        params.append(fid)
    if textId is not None:
        if textId == "":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="textId 不能为空")
        conditions.append('tm."textId" LIKE %s')
        params.append(f"{textId}%")
    if status_filter is not None:
        if status_filter not in (1, 2, 3):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="status 必须为 1/2/3")
        conditions.append("tm.status = %s")
        params.append(status_filter)
    if sourceKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause('tm."sourceText"', sourceKeyword, sourceMatchMode)
        conditions.append(condition_sql)
        params.append(condition_param)
    if translatedKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause(
            'tm."translatedText"', translatedKeyword, translatedMatchMode
        )
        conditions.append(condition_sql)
        params.append(condition_param)
    if updatedFrom is not None:
        conditions.append('tm."uptTime" >= %s')
        params.append(updatedFrom)
    if updatedTo is not None:
        conditions.append('tm."uptTime" <= %s')
        params.append(updatedTo)
    if claimer is not None:
        conditions.append(
            """
            (
              SELECT u.username
              FROM text_claims c
              JOIN users u ON u.id = c."userId"
              WHERE c."textId" = tm.id
              ORDER BY c."claimedAt" DESC, c.id DESC
              LIMIT 1
            ) LIKE %s
            """
        )
        params.append(f"%{claimer}%")
    if claimed is True:
        conditions.append('EXISTS (SELECT 1 FROM text_claims c WHERE c."textId" = tm.id)')
    if claimed is False:
        conditions.append('NOT EXISTS (SELECT 1 FROM text_claims c WHERE c."textId" = tm.id)')

    return conditions, params


@router.get("")
def list_texts(
    fid: Optional[str] = None,
    textId: Optional[str] = Query(default=None, alias="textId"),
    status_filter: Optional[int] = Query(default=None, alias="status"),
    sourceKeyword: Optional[str] = None,
    sourceMatchModeRaw: Optional[str] = Query(default=None, alias="sourceMatchMode"),
    translatedKeyword: Optional[str] = None,
    translatedMatchModeRaw: Optional[str] = Query(default=None, alias="translatedMatchMode"),
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    claimer: Optional[str] = None,
    claimed: Optional[bool] = None,
    page: int = 1,
    pageSize: Optional[int] = Query(default=None, alias="pageSize"),
    _: Dict[str, Any] = Depends(require_auth),
):
    """获取平铺主文本列表（全量 part），支持筛选与分页。"""
    config = get_config()
    pagination = config["pagination"]
    default_page_size = pagination["default_page_size"]
    max_page_size = pagination["max_page_size"]
    source_match_mode = _parse_text_match_mode(sourceMatchModeRaw, "sourceMatchMode")
    translated_match_mode = _parse_text_match_mode(translatedMatchModeRaw, "translatedMatchMode")

    effective_page_size = pageSize if pageSize is not None else default_page_size
    if effective_page_size > max_page_size:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 超出最大限制")

    offset = _apply_pagination(page, effective_page_size)
    conditions, params = _build_download_conditions(
        fid=fid,
        textId=textId,
        status_filter=status_filter,
        sourceKeyword=sourceKeyword,
        sourceMatchMode=source_match_mode,
        translatedKeyword=translatedKeyword,
        translatedMatchMode=translated_match_mode,
        updatedFrom=updatedFrom,
        updatedTo=updatedTo,
        claimer=claimer,
        claimed=claimed,
    )
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    max_text_length = config["text_list"]["max_text_length"]

    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM text_main tm
            {where_clause}
            """,
            tuple(params),
        )
        total = cursor.fetchone()["total"]

        cursor.execute(
            f"""
            SELECT
              tm.id,
              tm.fid,
              tm."textId" AS "textId",
              tm.part,
              CASE
                WHEN length(tm."sourceText") > %s THEN CONCAT(SUBSTRING(tm."sourceText", 1, %s), '...')
                ELSE tm."sourceText"
              END AS "sourceText",
              CASE
                WHEN tm."translatedText" IS NOT NULL AND length(tm."translatedText") > %s
                  THEN CONCAT(SUBSTRING(tm."translatedText", 1, %s), '...')
                ELSE tm."translatedText"
              END AS "translatedText",
              tm.status,
              tm."editCount" AS "editCount",
              tm."uptTime" AS "uptTime",
              tm."crtTime" AS "crtTime",
              (
                SELECT c.id
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimId",
              (
                SELECT u.username
                FROM text_claims c
                JOIN users u ON u.id = c."userId"
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedBy",
              (
                SELECT c."claimedAt"
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedAt",
              EXISTS (
                SELECT 1
                FROM text_claims c
                WHERE c."textId" = tm.id
              ) AS "isClaimed"
            FROM text_main tm
            {where_clause}
            ORDER BY tm."uptTime" DESC, tm.id DESC
            LIMIT %s OFFSET %s
            """,
            tuple([max_text_length, max_text_length, max_text_length, max_text_length] + params + [effective_page_size, offset]),
        )
        items = cursor.fetchall()
        for item in items:
            item["isClaimed"] = bool(item["isClaimed"])

    return success_response(
        {
            "items": items,
            "total": total,
            "page": page,
            "pageSize": effective_page_size,
        }
    )


@router.get("/parents")
def list_parent_texts(
    fid: Optional[str] = None,
    status_filter: Optional[int] = Query(default=None, alias="status"),
    sourceKeyword: Optional[str] = None,
    sourceMatchModeRaw: Optional[str] = Query(default=None, alias="sourceMatchMode"),
    translatedKeyword: Optional[str] = None,
    translatedMatchModeRaw: Optional[str] = Query(default=None, alias="translatedMatchMode"),
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    claimer: Optional[str] = None,
    claimed: Optional[bool] = None,
    page: int = 1,
    pageSize: Optional[int] = Query(default=None, alias="pageSize"),
    _: Dict[str, Any] = Depends(require_auth),
):
    """获取父级主文本列表（仅 part=1），支持筛选与分页。"""
    config = get_config()
    pagination = config["pagination"]
    default_page_size = pagination["default_page_size"]
    max_page_size = pagination["max_page_size"]
    source_match_mode = _parse_text_match_mode(sourceMatchModeRaw, "sourceMatchMode")
    translated_match_mode = _parse_text_match_mode(translatedMatchModeRaw, "translatedMatchMode")

    effective_page_size = pageSize if pageSize is not None else default_page_size
    if effective_page_size > max_page_size:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 超出最大限制")

    offset = _apply_pagination(page, effective_page_size)

    conditions: List[str] = []
    params: List[Any] = []

    if fid is not None:
        conditions.append("tm.fid = %s")
        params.append(fid)
    conditions.append("tm.part = %s")
    params.append(1)
    if status_filter is not None:
        if status_filter not in (1, 2, 3):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="status 必须为 1/2/3")
        conditions.append("tm.status = %s")
        params.append(status_filter)
    if sourceKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause('tmx."sourceText"', sourceKeyword, source_match_mode)
        conditions.append(
            f"""
            EXISTS (
                SELECT 1
                FROM text_main tmx
                WHERE tmx.fid = tm.fid
                  AND {condition_sql}
            )
            """
        )
        params.append(condition_param)
    if translatedKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause(
            'tmx."translatedText"', translatedKeyword, translated_match_mode
        )
        conditions.append(
            f"""
            EXISTS (
                SELECT 1
                FROM text_main tmx
                WHERE tmx.fid = tm.fid
                  AND {condition_sql}
            )
            """
        )
        params.append(condition_param)
    if updatedFrom is not None:
        conditions.append('tm."uptTime" >= %s')
        params.append(updatedFrom)
    if updatedTo is not None:
        conditions.append('tm."uptTime" <= %s')
        params.append(updatedTo)
    if claimer is not None:
        conditions.append(
            """
            (
              SELECT u.username
              FROM text_claims c
              JOIN users u ON u.id = c."userId"
              WHERE c."textId" = tm.id
              ORDER BY c."claimedAt" DESC, c.id DESC
              LIMIT 1
            ) LIKE %s
            """
        )
        params.append(f"%{claimer}%")
    if claimed is True:
        conditions.append('EXISTS (SELECT 1 FROM text_claims c WHERE c."textId" = tm.id)')
    if claimed is False:
        conditions.append('NOT EXISTS (SELECT 1 FROM text_claims c WHERE c."textId" = tm.id)')

    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    max_text_length = config["text_list"]["max_text_length"]

    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM text_main tm
            {where_clause}
            """,
            tuple(params),
        )
        total = cursor.fetchone()["total"]

        cursor.execute(
            f"""
            SELECT
              tm.id,
              tm.fid,
              tm."textId" AS "textId",
              tm.part,
              CASE
                WHEN length(tm."sourceText") > %s THEN CONCAT(SUBSTRING(tm."sourceText", 1, %s), '...')
                ELSE tm."sourceText"
              END AS "sourceText",
              CASE
                WHEN tm."translatedText" IS NOT NULL AND length(tm."translatedText") > %s
                  THEN CONCAT(SUBSTRING(tm."translatedText", 1, %s), '...')
                ELSE tm."translatedText"
              END AS "translatedText",
              tm.status,
              tm."editCount" AS "editCount",
              tm."uptTime" AS "uptTime",
              tm."crtTime" AS "crtTime",
              (
                SELECT c.id
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimId",
              (
                SELECT u.username
                FROM text_claims c
                JOIN users u ON u.id = c."userId"
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedBy",
              (
                SELECT c."claimedAt"
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedAt",
              EXISTS (
                SELECT 1
                FROM text_claims c
                WHERE c."textId" = tm.id
              ) AS "isClaimed"
            FROM text_main tm
            {where_clause}
            ORDER BY tm."uptTime" DESC
            LIMIT %s OFFSET %s
            """,
            tuple([max_text_length, max_text_length, max_text_length, max_text_length] + params + [effective_page_size, offset]),
        )
        items = cursor.fetchall()
        for item in items:
            item["isClaimed"] = bool(item["isClaimed"])

    return success_response(
        {
            "items": items,
            "total": total,
            "page": page,
            "pageSize": effective_page_size,
        }
    )


@router.get("/children")
def list_child_texts(
    fid: str,
    textId: Optional[str] = Query(default=None, alias="textId"),
    sourceKeyword: Optional[str] = None,
    sourceMatchModeRaw: Optional[str] = Query(default=None, alias="sourceMatchMode"),
    translatedKeyword: Optional[str] = None,
    translatedMatchModeRaw: Optional[str] = Query(default=None, alias="translatedMatchMode"),
    page: int = 1,
    pageSize: Optional[int] = Query(default=None, alias="pageSize"),
    _: Dict[str, Any] = Depends(require_auth),
):
    """获取指定 fid 的子列表（默认排除 part=1），支持筛选与分页。"""
    config = get_config()
    pagination = config["pagination"]
    default_page_size = pagination["default_page_size"]
    max_page_size = pagination["max_page_size"]
    source_match_mode = _parse_text_match_mode(sourceMatchModeRaw, "sourceMatchMode")
    translated_match_mode = _parse_text_match_mode(translatedMatchModeRaw, "translatedMatchMode")

    effective_page_size = pageSize if pageSize is not None else default_page_size
    if effective_page_size > max_page_size:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="pageSize 超出最大限制")

    offset = _apply_pagination(page, effective_page_size)

    params: List[Any] = [fid]
    where_clause = "WHERE tm.fid = %s AND tm.part <> 1"
    if textId is not None:
        where_clause += ' AND tm."textId" LIKE %s'
        params.append(f"{textId}%")
    if sourceKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause('tm."sourceText"', sourceKeyword, source_match_mode)
        where_clause += f" AND {condition_sql}"
        params.append(condition_param)
    if translatedKeyword is not None:
        condition_sql, condition_param = _build_text_match_clause(
            'tm."translatedText"', translatedKeyword, translated_match_mode
        )
        where_clause += f" AND {condition_sql}"
        params.append(condition_param)

    max_text_length = config["text_list"]["max_text_length"]

    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM text_main tm
            {where_clause}
            """,
            tuple(params),
        )
        total = cursor.fetchone()["total"]

        cursor.execute(
            f"""
            SELECT
              tm.id,
              tm.fid,
              tm."textId" AS "textId",
              tm.part,
              CASE
                WHEN length(tm."sourceText") > %s THEN CONCAT(SUBSTRING(tm."sourceText", 1, %s), '...')
                ELSE tm."sourceText"
              END AS "sourceText",
              CASE
                WHEN tm."translatedText" IS NOT NULL AND length(tm."translatedText") > %s
                  THEN CONCAT(SUBSTRING(tm."translatedText", 1, %s), '...')
                ELSE tm."translatedText"
              END AS "translatedText",
              tm.status,
              tm."editCount" AS "editCount",
              tm."uptTime" AS "uptTime",
              tm."crtTime" AS "crtTime",
              (
                SELECT c.id
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimId",
              (
                SELECT u.username
                FROM text_claims c
                JOIN users u ON u.id = c."userId"
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedBy",
              (
                SELECT c."claimedAt"
                FROM text_claims c
                WHERE c."textId" = tm.id
                ORDER BY c."claimedAt" DESC, c.id DESC
                LIMIT 1
              ) AS "claimedAt",
              EXISTS (
                SELECT 1
                FROM text_claims c
                WHERE c."textId" = tm.id
              ) AS "isClaimed"
            FROM text_main tm
            {where_clause}
            ORDER BY tm.part ASC
            LIMIT %s OFFSET %s
            """,
            tuple([max_text_length, max_text_length, max_text_length, max_text_length] + params + [effective_page_size, offset]),
        )
        items = cursor.fetchall()
        for item in items:
            item["isClaimed"] = bool(item["isClaimed"])

    return success_response(
        {
            "items": items,
            "total": total,
            "page": page,
            "pageSize": effective_page_size,
        }
    )


@router.get("/template")
def download_text_template(_: Dict[str, Any] = Depends(require_auth)):
    """下载上传模板（仅表头）。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "texts"
    sheet.append(list(TEXT_TEMPLATE_HEADERS))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="text_template.xlsx"'},
    )


@router.get("/download")
def download_texts(
    fid: Optional[str] = None,
    textId: Optional[str] = Query(default=None, alias="textId"),
    status_filter: Optional[int] = Query(default=None, alias="status"),
    sourceKeyword: Optional[str] = None,
    sourceMatchModeRaw: Optional[str] = Query(default=None, alias="sourceMatchMode"),
    translatedKeyword: Optional[str] = None,
    translatedMatchModeRaw: Optional[str] = Query(default=None, alias="translatedMatchMode"),
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    claimer: Optional[str] = None,
    claimed: Optional[bool] = None,
    _: Dict[str, Any] = Depends(require_auth),
):
    """根据筛选条件导出文本数据（流式 + 低内存）。"""
    request_started_at = perf_counter()
    logger.info(
        "download_texts start: fid={} status={} sourceKeyword={} translatedKeyword={} updatedFrom={} updatedTo={} claimer={} claimed={}",
        fid,
        status_filter,
        sourceKeyword,
        translatedKeyword,
        updatedFrom,
        updatedTo,
        claimer,
        claimed,
    )
    source_match_mode = _parse_text_match_mode(sourceMatchModeRaw, "sourceMatchMode")
    translated_match_mode = _parse_text_match_mode(translatedMatchModeRaw, "translatedMatchMode")
    config = get_config()
    text_import_export = config["text_import_export"]
    max_download_rows = text_import_export["max_download_rows"]
    download_fetch_batch_size = text_import_export["download_fetch_batch_size"]
    download_progress_log_every_batches = text_import_export["download_progress_log_every_batches"]
    download_temp_dir = text_import_export["download_temp_dir"]

    if not os.path.isdir(download_temp_dir):
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="导出临时目录不存在")

    conditions, params = _build_download_conditions(
        fid=fid,
        textId=textId,
        status_filter=status_filter,
        sourceKeyword=sourceKeyword,
        sourceMatchMode=source_match_mode,
        translatedKeyword=translatedKeyword,
        translatedMatchMode=translated_match_mode,
        updatedFrom=updatedFrom,
        updatedTo=updatedTo,
        claimer=claimer,
        claimed=claimed,
    )
    where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    logger.info(
        "download_texts stage=build_conditions done: conditionCount={} elapsedSec={:.3f}",
        len(conditions),
        perf_counter() - request_started_at,
    )

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="texts")
    sheet.append(list(TEXT_TEMPLATE_HEADERS))
    logger.info(
        "download_texts stage=init_workbook done: elapsedSec={:.3f}",
        perf_counter() - request_started_at,
    )

    export_count = 0
    fetched_row_count = 0
    batch_count = 0
    tmp_path: Optional[str] = None
    try:
        db_read_started_at = perf_counter()
        logger.info("download_texts stage=db_stream start")
        with db_stream_cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                  tm.id,
                  tm.fid,
                  tm."textId" AS "textId",
                  tm.part,
                  tm."sourceText" AS "sourceText",
                  tm."translatedText" AS "translatedText",
                  tm.status
                FROM text_main tm
                {where_clause}
                ORDER BY tm."uptTime" DESC, tm.id DESC
                """,
                tuple(params),
            )
            while True:
                rows = cursor.fetchmany(download_fetch_batch_size)
                if not rows:
                    break
                batch_count += 1
                fetched_row_count += len(rows)
                for row in rows:
                    export_count += 1
                    if export_count > max_download_rows:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"导出数据量超过限制（{max_download_rows}），请缩小筛选范围后重试",
                        )
                    sheet.append(
                        [
                            row["id"],
                            row["fid"],
                            row["textId"],
                            row["part"],
                            row["sourceText"],
                            row["translatedText"],
                            _format_status_label(row["status"]),
                        ]
                    )
                if batch_count == 1 or batch_count % download_progress_log_every_batches == 0:
                    logger.info(
                        "download_texts stage=db_stream progress: batch={} batchRows={} fetchedRows={} exportRows={} elapsedSec={:.3f}",
                        batch_count,
                        len(rows),
                        fetched_row_count,
                        export_count,
                        perf_counter() - request_started_at,
                    )
        logger.info(
            "download_texts stage=db_stream done: batches={} fetchedRows={} exportRows={} stageElapsedSec={:.3f} totalElapsedSec={:.3f}",
            batch_count,
            fetched_row_count,
            export_count,
            perf_counter() - db_read_started_at,
            perf_counter() - request_started_at,
        )

        if export_count == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前筛选条件无可导出数据")

        tmp_file = NamedTemporaryFile(
            prefix="tmp_text_export_",
            suffix=".xlsx",
            dir=download_temp_dir,
            delete=False,
        )
        tmp_path = tmp_file.name
        tmp_file.close()
        save_started_at = perf_counter()
        logger.info("download_texts stage=save_xlsx start: tmpPath={}", tmp_path)
        workbook.save(tmp_path)
        logger.info(
            "download_texts stage=save_xlsx done: fileSizeBytes={} stageElapsedSec={:.3f} totalElapsedSec={:.3f}",
            os.path.getsize(tmp_path),
            perf_counter() - save_started_at,
            perf_counter() - request_started_at,
        )
    except Exception:
        logger.exception(
            "download_texts failed: fetchedRows={} exportRows={} batches={} elapsedSec={:.3f}",
            fetched_row_count,
            export_count,
            batch_count,
            perf_counter() - request_started_at,
        )
        workbook.close()
        if tmp_path is not None:
            _cleanup_temp_file(tmp_path)
        raise
    workbook.close()

    export_name = f"text_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    logger.info(
        "download_texts done: exportName={} tmpPath={} exportRows={} fetchedRows={} batches={} totalElapsedSec={:.3f}",
        export_name,
        tmp_path,
        export_count,
        fetched_row_count,
        batch_count,
        perf_counter() - request_started_at,
    )
    return FileResponse(
        path=tmp_path,
        filename=export_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(_cleanup_temp_file, tmp_path),
    )


@router.get("/download-package")
def download_package(
    fid: Optional[str] = None,
    status_filter: Optional[int] = Query(default=None, alias="status"),
    sourceKeyword: Optional[str] = None,
    sourceMatchModeRaw: Optional[str] = Query(default=None, alias="sourceMatchMode"),
    translatedKeyword: Optional[str] = None,
    translatedMatchModeRaw: Optional[str] = Query(default=None, alias="translatedMatchMode"),
    updatedFrom: Optional[str] = None,
    updatedTo: Optional[str] = None,
    claimer: Optional[str] = None,
    claimed: Optional[bool] = None,
    _: Dict[str, Any] = Depends(require_auth),
):
    """下载汉化包：按 fid + part 顺序流式读取，Python 端按 fid 增量合并。
    translation 超过单元格字符限制时按 segment 边界自动分行，不截断任何 segment。
    使用进程内信号量避免并发导出拖垮系统。
    """
    request_started_at = perf_counter()
    logger.info(
        "download_package start: fid={} status={} sourceKeyword={} translatedKeyword={} updatedFrom={} updatedTo={} claimer={} claimed={}",
        fid,
        status_filter,
        sourceKeyword,
        translatedKeyword,
        updatedFrom,
        updatedTo,
        claimer,
        claimed,
    )
    source_match_mode = _parse_text_match_mode(sourceMatchModeRaw, "sourceMatchMode")
    translated_match_mode = _parse_text_match_mode(translatedMatchModeRaw, "translatedMatchMode")
    if not _package_download_lock.acquire(blocking=False):
        logger.warning(
            "download_package rejected: semaphore busy elapsedSec={:.3f}",
            perf_counter() - request_started_at,
        )
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="汉化包正在生成中，请稍后再试",
        )
    try:
        config = get_config()
        text_import_export = config["text_import_export"]
        max_download_rows = text_import_export["max_download_rows"]
        download_fetch_batch_size = text_import_export["download_fetch_batch_size"]
        download_progress_log_every_batches = text_import_export["download_progress_log_every_batches"]
        download_temp_dir = text_import_export["download_temp_dir"]

        if not os.path.isdir(download_temp_dir):
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="导出临时目录不存在")

        conditions, params = _build_download_conditions(
            fid=fid,
            textId=None,
            status_filter=status_filter,
            sourceKeyword=sourceKeyword,
            sourceMatchMode=source_match_mode,
            translatedKeyword=translatedKeyword,
            translatedMatchMode=translated_match_mode,
            updatedFrom=updatedFrom,
            updatedTo=updatedTo,
            claimer=claimer,
            claimed=claimed,
        )
        where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        logger.info(
            "download_package stage=build_conditions done: conditionCount={} elapsedSec={:.3f}",
            len(conditions),
            perf_counter() - request_started_at,
        )

        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="texts")
        sheet.append(list(PACKAGE_HEADERS))
        logger.info(
            "download_package stage=init_workbook done: elapsedSec={:.3f}",
            perf_counter() - request_started_at,
        )

        output_fid_count = 0
        fetched_part_rows = 0
        batch_count = 0
        tmp_path: Optional[str] = None

        try:
            current_fid: Optional[str] = None
            current_segments: List[str] = []

            def flush_current_fid() -> None:
                nonlocal output_fid_count, current_fid, current_segments
                if current_fid is None:
                    return
                output_fid_count += 1
                if output_fid_count > max_download_rows:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"导出数据量超过限制（{max_download_rows}），请缩小筛选范围后重试",
                    )
                translation = "|||".join(current_segments)
                if len(translation) > _EXCEL_CELL_CHAR_LIMIT:
                    logger.warning(f"fid={current_fid} translation 超过 Excel 单元格字符上限，按 segment 边界分行")
                    for split_row in _split_translation_into_rows(current_fid, translation):
                        sheet.append(split_row)
                else:
                    sheet.append([current_fid, translation])
                current_fid = None
                current_segments = []

            with db_stream_cursor() as cursor:
                db_read_started_at = perf_counter()
                logger.info("download_package stage=db_stream start")
                cursor.execute(
                    f"""
                    SELECT
                      tm.fid,
                      tm."textId" AS "textId",
                      tm.part,
                      tm."sourceText" AS "sourceText",
                      tm."translatedText" AS "translatedText"
                    FROM text_main tm
                    {where_clause}
                    ORDER BY tm.fid ASC, tm.part ASC
                    """,
                    tuple(params),
                )
                while True:
                    rows = cursor.fetchmany(download_fetch_batch_size)
                    if not rows:
                        break
                    batch_count += 1
                    fetched_part_rows += len(rows)
                    for row in rows:
                        row_fid = row["fid"]
                        if current_fid is None:
                            current_fid = row_fid
                        elif row_fid != current_fid:
                            flush_current_fid()
                            current_fid = row_fid

                        source_text = row["sourceText"] or ""
                        translated_text = row["translatedText"] or source_text
                        text_id = str(row["textId"])
                        if ":::" in text_id:
                            current_segments.append(f"{text_id}:::[{translated_text}]")
                        else:
                            current_segments.append(f"{text_id}::::::[{translated_text}]")
                    if batch_count == 1 or batch_count % download_progress_log_every_batches == 0:
                        logger.info(
                            "download_package stage=db_stream progress: batch={} batchRows={} fetchedPartRows={} flushedFidRows={} elapsedSec={:.3f}",
                            batch_count,
                            len(rows),
                            fetched_part_rows,
                            output_fid_count,
                            perf_counter() - request_started_at,
                        )

                flush_current_fid()
                logger.info(
                    "download_package stage=db_stream done: batches={} fetchedPartRows={} fidRows={} stageElapsedSec={:.3f} totalElapsedSec={:.3f}",
                    batch_count,
                    fetched_part_rows,
                    output_fid_count,
                    perf_counter() - db_read_started_at,
                    perf_counter() - request_started_at,
                )

            if output_fid_count == 0:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前筛选条件无可导出数据")

            tmp_file = NamedTemporaryFile(
                prefix="tmp_package_export_",
                suffix=".xlsx",
                dir=download_temp_dir,
                delete=False,
            )
            tmp_path = tmp_file.name
            tmp_file.close()
            save_started_at = perf_counter()
            logger.info("download_package stage=save_xlsx start: tmpPath={}", tmp_path)
            workbook.save(tmp_path)
            logger.info(
                "download_package stage=save_xlsx done: fileSizeBytes={} stageElapsedSec={:.3f} totalElapsedSec={:.3f}",
                os.path.getsize(tmp_path),
                perf_counter() - save_started_at,
                perf_counter() - request_started_at,
            )
        except Exception:
            logger.exception(
                "download_package failed: fetchedPartRows={} fidRows={} batches={} elapsedSec={:.3f}",
                fetched_part_rows,
                output_fid_count,
                batch_count,
                perf_counter() - request_started_at,
            )
            workbook.close()
            if tmp_path is not None:
                _cleanup_temp_file(tmp_path)
            raise
        workbook.close()
    finally:
        _package_download_lock.release()
        logger.info(
            "download_package stage=release_lock done: elapsedSec={:.3f}",
            perf_counter() - request_started_at,
        )

    logger.info(
        "download_package done: tmpPath={} fidRows={} fetchedPartRows={} batches={} totalElapsedSec={:.3f}",
        tmp_path,
        output_fid_count,
        fetched_part_rows,
        batch_count,
        perf_counter() - request_started_at,
    )
    return FileResponse(
        path=tmp_path,
        filename="text_work.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(_cleanup_temp_file, tmp_path),
    )


@router.post("/upload")
async def upload_text_template(
    request: Request,
    fileName: str = Query(..., alias="fileName"),
    reason: Optional[str] = Query(default=None, alias="reason"),
    user: Dict[str, Any] = Depends(require_auth),
):
    """按模板上传翻译结果并覆盖译文与状态。"""
    logger.info(f"Upload start: fileName={fileName} userId={user['userId']} reason={reason}")
    if not fileName:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件缺少文件名")
    if not fileName.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件必须为 .xlsx 格式")

    config = get_config()
    text_import_export = config["text_import_export"]
    max_upload_rows = text_import_export["max_upload_rows"]

    file_bytes = await request.body()
    sheet = _load_upload_sheet(file_bytes)

    header_rows = list(sheet.iter_rows(min_row=1, max_row=1))
    if not header_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件为空")
    header_values = [cell.value for cell in header_rows[0]]
    _validate_template_header(header_values)

    parsed_rows: List[Dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row[: len(TEXT_TEMPLATE_HEADERS)])
        if len(cells) < len(TEXT_TEMPLATE_HEADERS):
            cells.extend([None] * (len(TEXT_TEMPLATE_HEADERS) - len(cells)))
        if all(_is_empty_cell(item) for item in cells):
            continue

        extra_cells = list(row[len(TEXT_TEMPLATE_HEADERS) :])
        if any(not _is_empty_cell(item) for item in extra_cells):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {row_index} 行存在模板外数据，请删除多余列",
            )

        row_id = _parse_required_int(cells[0], "编号", row_index)
        text_id = _parse_required_str(cells[2], "TextId", row_index)
        part = _parse_required_int(cells[3], "Part", row_index)
        if row_id <= 0 or part <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"第 {row_index} 行编号/Part 必须 > 0")

        parsed_rows.append(
            {
                "rowNumber": row_index,
                "id": row_id,
                "fid": _parse_required_str(cells[1], "FID", row_index),
                "textId": text_id,
                "part": part,
                "translatedText": _normalize_text(cells[5]),
                "status": _parse_status(cells[6], row_index),
            }
        )

    if not parsed_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件没有可处理的数据行")

    if len(parsed_rows) > max_upload_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"上传行数超限，最大允许 {max_upload_rows} 行",
        )

    ids = [item["id"] for item in parsed_rows]
    if len(set(ids)) != len(ids):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="上传文件存在重复编号")

    placeholders = ",".join(["%s"] * len(ids))
    with db_cursor() as cursor:
        cursor.execute(
            f"""
            SELECT id, fid, "textId" AS "textId", part, "translatedText" AS "translatedText"
            FROM text_main
            WHERE id IN ({placeholders})
            """,
            tuple(ids),
        )
        db_rows = cursor.fetchall()
        db_map = {item["id"]: item for item in db_rows}

        for item in parsed_rows:
            db_item = db_map.get(item["id"])
            if db_item is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"第 {item['rowNumber']} 行编号不存在: {item['id']}",
                )
            if db_item["fid"] != item["fid"] or db_item["textId"] != item["textId"] or db_item["part"] != item["part"]:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"第 {item['rowNumber']} 行校验失败: 编号/FID/TextId/Part 与数据库不匹配",
                )

        for item in parsed_rows:
            db_item = db_map[item["id"]]
            before_text = db_item["translatedText"] or ""
            cursor.execute(
                """
                UPDATE text_main
                SET "translatedText" = %s, status = %s, "editCount" = "editCount" + 1, "uptTime" = NOW()
                WHERE id = %s
                """,
                (item["translatedText"], item["status"], item["id"]),
            )
            cursor.execute(
                """
                INSERT INTO text_changes ("textId", "userId", "beforeText", "afterText", reason)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (item["id"], user["userId"], before_text, item["translatedText"] or "", reason),
            )

    logger.info(f"Upload complete: fileName={fileName} updatedCount={len(parsed_rows)} userId={user['userId']}")
    return success_response({"updatedCount": len(parsed_rows)})


@router.get("/by-textid")
def get_text_by_textid(
    fid: str,
    textId: str = Query(..., alias="textId"),
    _: Dict[str, Any] = Depends(require_auth),
):
    """根据 fid + textId 获取主文本详情。"""
    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT
              id,
              fid,
              "textId" AS "textId",
              part,
              "sourceText" AS "sourceText",
              "translatedText" AS "translatedText",
              status,
              "editCount" AS "editCount",
              "uptTime" AS "uptTime",
              "crtTime" AS "crtTime"
            FROM text_main
            WHERE fid = %s AND "textId" = %s
            """,
            (fid, textId),
        )
        rows = cursor.fetchall()
        if not rows:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="文本不存在")
        if len(rows) > 1:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="textId 在该 fid 下存在重复数据")
        text = rows[0]

        cursor.execute(
            """
            SELECT id, "userId" AS "userId", "claimedAt" AS "claimedAt"
            FROM text_claims
            WHERE "textId" = %s
            ORDER BY "claimedAt" DESC
            """,
            (text["id"],),
        )
        claims = cursor.fetchall()

        cursor.execute(
            """
            SELECT
              id,
              "userId" AS "userId",
              "lockedAt" AS "lockedAt",
              "expiresAt" AS "expiresAt",
              "releasedAt" AS "releasedAt"
            FROM text_locks
            WHERE "textId" = %s
            ORDER BY "lockedAt" DESC
            """,
            (text["id"],),
        )
        locks = cursor.fetchall()

    return success_response(
        {
            "text": text,
            "claims": claims,
            "locks": locks,
        }
    )


@router.get("/{textId}")
def get_text(textId: int, _: Dict[str, Any] = Depends(require_auth)):
    """获取主文本详情以及认领/锁定信息。"""
    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT
              id,
              fid,
              "textId" AS "textId",
              part,
              "sourceText" AS "sourceText",
              "translatedText" AS "translatedText",
              status,
              "editCount" AS "editCount",
              "uptTime" AS "uptTime",
              "crtTime" AS "crtTime"
            FROM text_main
            WHERE id = %s
            """,
            (textId,),
        )
        text = cursor.fetchone()
        if text is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="文本不存在")

        cursor.execute(
            """
            SELECT id, "userId" AS "userId", "claimedAt" AS "claimedAt"
            FROM text_claims
            WHERE "textId" = %s
            ORDER BY "claimedAt" DESC
            """,
            (textId,),
        )
        claims = cursor.fetchall()

        cursor.execute(
            """
            SELECT
              id,
              "userId" AS "userId",
              "lockedAt" AS "lockedAt",
              "expiresAt" AS "expiresAt",
              "releasedAt" AS "releasedAt"
            FROM text_locks
            WHERE "textId" = %s
            ORDER BY "lockedAt" DESC
            """,
            (textId,),
        )
        locks = cursor.fetchall()

    return success_response(
        {
            "text": text,
            "claims": claims,
            "locks": locks,
        }
    )


class TranslateRequest(BaseModel):
    translatedText: str
    reason: Optional[str] = None
    isCompleted: Optional[bool] = None


@router.put("/{textId}/translate")
def update_translation(
    textId: int,
    request: TranslateRequest,
    user: Dict[str, Any] = Depends(require_auth),
):
    """保存译文并写入变更记录。"""
    logger.info(f"Translate: textId={textId} userId={user['userId']} isCompleted={request.isCompleted} reason={request.reason}")
    with db_cursor() as cursor:
        cursor.execute(
            'SELECT "translatedText" FROM text_main WHERE id = %s',
            (textId,),
        )
        row = cursor.fetchone()
        if row is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="文本不存在")

        beforeText = row["translatedText"] or ""
        if request.isCompleted:
            cursor.execute(
                """
                UPDATE text_main
                SET "translatedText" = %s, status = %s, "editCount" = "editCount" + 1, "uptTime" = NOW()
                WHERE id = %s
                """,
                (request.translatedText, 3, textId),
            )
        else:
            cursor.execute(
                """
                UPDATE text_main
                SET "translatedText" = %s, status = %s, "editCount" = "editCount" + 1, "uptTime" = NOW()
                WHERE id = %s
                """,
                (request.translatedText, 2, textId),
            )
        cursor.execute(
            """
            INSERT INTO text_changes ("textId", "userId", "beforeText", "afterText", reason)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (textId, user["userId"], beforeText, request.translatedText, request.reason),
        )

    return success_response({"id": textId})
