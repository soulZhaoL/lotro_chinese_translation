# 词典接口测试。
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from server.app import app
from server.db import db_cursor
from server.services import dictionary_correction


def _login(client: TestClient, seed_user):
    response = client.post("/auth/login", json={"username": seed_user["username"], "password": seed_user["password"]})
    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == "0000"
    return payload["data"]["token"]


def _build_dictionary_xlsx(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dictionary"
    sheet.append(["原文 key", "标准译文", "译文变体JSON", "分类", "备注"])
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_dictionary_filter_and_create_with_remark(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    create = client.post(
        "/dictionary",
        json={
            "termKey": "orc",
            "termValue": "兽人",
            "variantValues": ["半兽人", "奥克", "兽人"],
            "category": "race",
            "remark": "怪物称呼",
        },
        headers=headers,
    )
    assert create.status_code == 200

    response = client.get("/dictionary?keyword=orc", headers=headers)
    assert response.status_code == 200
    data = response.json()["data"]
    assert data["total"] == 1
    assert data["items"][0]["termKey"] == "orc"
    assert data["items"][0]["variantValues"] == ["半兽人", "奥克"]
    assert data["items"][0]["remark"] == "怪物称呼"
    assert data["items"][0]["lastModifiedByName"] == seed_user["username"]


def test_dictionary_update(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    create = client.post(
        "/dictionary",
        json={"termKey": "orc", "termValue": "兽人", "category": "race"},
        headers=headers,
    )
    entry_id = create.json()["data"]["id"]

    update = client.put(
        f"/dictionary/{entry_id}",
        json={
            "termValue": "半兽人",
            "variantValues": ["兽人", "奥克", "半兽人", "奥克"],
            "category": "quest",
            "remark": "剧情相关",
        },
        headers=headers,
    )
    assert update.status_code == 200

    response = client.get("/dictionary?termKey=orc", headers=headers)
    item = response.json()["data"]["items"][0]
    assert item["termValue"] == "半兽人"
    assert item["variantValues"] == ["兽人", "奥克"]
    assert item["category"] == "quest"
    assert item["remark"] == "剧情相关"


def test_dictionary_template_download_and_upload_upsert(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    create = client.post(
        "/dictionary",
        json={"termKey": "orc", "termValue": "兽人", "category": "race", "remark": "旧备注"},
        headers=headers,
    )
    assert create.status_code == 200

    template_response = client.get("/dictionary/template", headers=headers)
    assert template_response.status_code == 200
    workbook = load_workbook(BytesIO(template_response.content))
    sheet = workbook.active
    headers_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers_row == ["原文 key", "标准译文", "译文变体JSON", "分类", "备注"]

    upload_bytes = _build_dictionary_xlsx(
        [
          ["orc", "半兽人", '["兽族"]', "race", "已覆盖"],
          ["elf", "精灵", '["尖耳朵"]', "race", "新增词条"],
        ]
    )
    upload_response = client.post(
        "/dictionary/upload?fileName=dictionary.xlsx",
        content=upload_bytes,
        headers={
            **headers,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    assert upload_response.status_code == 200
    payload = upload_response.json()["data"]
    assert payload["createdCount"] == 1
    assert payload["updatedCount"] == 1

    orc_response = client.get("/dictionary?termKey=orc", headers=headers)
    orc_item = orc_response.json()["data"]["items"][0]
    assert orc_item["termValue"] == "半兽人"
    assert orc_item["variantValues"] == ["兽族"]
    assert orc_item["remark"] == "已覆盖"

    elf_response = client.get("/dictionary?termKey=elf", headers=headers)
    elf_item = elf_response.json()["data"]["items"][0]
    assert elf_item["termValue"] == "精灵"
    assert elf_item["variantValues"] == ["尖耳朵"]
    assert elf_item["remark"] == "新增词条"


def test_dictionary_upload_rejects_duplicate_term_key(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    upload_bytes = _build_dictionary_xlsx(
        [
          ["orc", "兽人", None, "race", None],
          ["orc", "半兽人", None, "race", "重复"],
        ]
    )
    upload_response = client.post(
        "/dictionary/upload?fileName=dictionary.xlsx",
        content=upload_bytes,
        headers={
            **headers,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    assert upload_response.status_code == 400
    assert upload_response.json()["detail"] == "上传文件存在重复原文 key"


def test_dictionary_upload_rejects_invalid_variant_json(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    upload_bytes = _build_dictionary_xlsx(
        [
          ["skill_t", "skill_t", '["skill1"", "skill2"]', "skill", "非法 JSON"],
        ]
    )
    upload_response = client.post(
        "/dictionary/upload?fileName=dictionary.xlsx",
        content=upload_bytes,
        headers={
            **headers,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    assert upload_response.status_code == 400
    assert upload_response.json()["detail"].startswith("第 2 行字段 译文变体JSON 不是合法 JSON")


def test_dictionary_correct_all_requeues_non_running_entries(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    create_orc = client.post(
        "/dictionary",
        json={"termKey": "orc", "termValue": "兽人", "variantValues": ["奥克"], "category": "race"},
        headers=headers,
    )
    assert create_orc.status_code == 200
    orc_id = create_orc.json()["data"]["id"]

    create_elf = client.post(
        "/dictionary",
        json={"termKey": "elf", "termValue": "精灵", "category": "race"},
        headers=headers,
    )
    assert create_elf.status_code == 200
    elf_id = create_elf.json()["data"]["id"]

    create_man = client.post(
        "/dictionary",
        json={"termKey": "man", "termValue": "人类", "variantValues": ["迈安"], "category": "race"},
        headers=headers,
    )
    assert create_man.status_code == 200
    man_id = create_man.json()["data"]["id"]

    with db_cursor() as cursor:
        cursor.execute(
            """
            UPDATE dictionary_entries
            SET "correctionStatus" = %s
            WHERE id = %s
            """,
            (dictionary_correction.CORRECTION_STATUS_RUNNING, man_id),
        )

    response = client.post("/dictionary/correct-all", headers=headers)
    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["totalCount"] == 3
    assert payload["requeuedCount"] == 2
    assert payload["skippedRunningCount"] == 1

    with db_cursor() as cursor:
        cursor.execute(
            """
            SELECT
              id,
              "termKey" AS "termKey",
              "correctionVersion" AS "correctionVersion",
              "appliedCorrectionVersion" AS "appliedCorrectionVersion",
              "correctionStatus" AS "correctionStatus"
            FROM dictionary_entries
            ORDER BY id ASC
            """
        )
        rows = cursor.fetchall()

    row_map = {row["termKey"]: row for row in rows}
    assert row_map["orc"]["id"] == orc_id
    assert row_map["orc"]["correctionVersion"] == 2
    assert row_map["orc"]["appliedCorrectionVersion"] == 0
    assert row_map["orc"]["correctionStatus"] == dictionary_correction.CORRECTION_STATUS_PENDING

    assert row_map["elf"]["id"] == elf_id
    assert row_map["elf"]["correctionVersion"] == 1
    assert row_map["elf"]["appliedCorrectionVersion"] == 0
    assert row_map["elf"]["correctionStatus"] == dictionary_correction.CORRECTION_STATUS_PENDING

    assert row_map["man"]["id"] == man_id
    assert row_map["man"]["correctionVersion"] == 1
    assert row_map["man"]["appliedCorrectionVersion"] == 0
    assert row_map["man"]["correctionStatus"] == dictionary_correction.CORRECTION_STATUS_RUNNING


def test_dictionary_correction_records_lists_abnormal_items(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    create = client.post(
        "/dictionary",
        json={"termKey": "Bree", "termValue": "布雷", "variantValues": ["布里"], "category": "place"},
        headers=headers,
    )
    assert create.status_code == 200
    entry_id = create.json()["data"]["id"]

    with db_cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO dictionary_correction_logs (
              "dictionaryEntryId",
              "correctionVersion",
              "textMainId",
              fid,
              "textId",
              action,
              reason,
              "sourceMatchCount",
              "translatedMatchCount"
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                entry_id,
                1,
                1001,
                "620799665",
                "218649171",
                "skipped",
                "原文匹配 1 次，译文匹配 2 次，次数不一致",
                1,
                2,
            ),
        )

    response = client.get(
        f"/dictionary/{entry_id}/correction-records?onlyAbnormal=true&page=1&pageSize=20",
        headers=headers,
    )
    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["entryId"] == entry_id
    assert payload["termKey"] == "Bree"
    assert payload["correctionVersion"] == 1
    assert payload["total"] == 1
    assert payload["items"][0]["fid"] == "620799665"
    assert payload["items"][0]["textId"] == "218649171"
    assert payload["items"][0]["sourceMatchCount"] == 1
    assert payload["items"][0]["translatedMatchCount"] == 2
