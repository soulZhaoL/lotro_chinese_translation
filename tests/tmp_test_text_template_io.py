# 文本模板下载与上传测试。
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from server.app import app
from server.db import db_cursor


def _login(client: TestClient, seed_user):
    response = client.post("/auth/login", json={"username": seed_user["username"], "password": seed_user["password"]})
    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == "0000"
    return payload["data"]["token"]


def _build_xlsx(rows, headers=None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "texts"
    final_headers = headers or ["编号", "FID", "TextId", "Part", "原文", "译文", "状态"]
    sheet.append(final_headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_text_template_download(seed_user):
    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    response = client.get("/texts/template", headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("编号", "FID", "TextId", "Part", "原文", "译文", "状态")
    assert len(rows) == 1


def test_text_download_by_filter(seed_user):
    with db_cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_download", 12001, 1, "src_1", "dst_1", 2, 4),
        )
        row_a_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_download_other", 12002, 1, "src_2", "dst_2", 1, 0),
        )

    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    response = client.get("/texts/download?fid=file_download", headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("编号", "FID", "TextId", "Part", "原文", "译文", "状态")
    assert len(rows) == 2
    assert rows[1] == (row_a_id, "file_download", 12001, 1, "src_1", "dst_1", "修改")


def test_text_template_upload_success(seed_user):
    with db_cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_upload", 13001, 2, "src_upload", "old_translation", 2, 3),
        )
        text_id = cursor.lastrowid

    upload_bytes = _build_xlsx(
        [
            [text_id, "file_upload", 13001, 2, "src_upload", "new_translation", 3],
        ]
    )

    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    response = client.post(
        "/texts/upload?fileName=tmp_upload.xlsx&reason=%E7%BA%BF%E4%B8%8B%E5%9B%9E%E4%BC%A0",
        headers=headers,
        content=upload_bytes,
    )
    assert response.status_code == 200
    assert response.json()["data"]["updatedCount"] == 1

    with db_cursor() as cursor:
        cursor.execute('SELECT "translatedText", status, "editCount" FROM text_main WHERE id = %s', (text_id,))
        row = cursor.fetchone()
        assert row["translatedText"] == "new_translation"
        assert row["status"] == 3
        assert row["editCount"] == 4

        cursor.execute(
            """
            SELECT "beforeText", "afterText", reason
            FROM text_changes
            WHERE "textId" = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (text_id,),
        )
        change = cursor.fetchone()
        assert change["beforeText"] == "old_translation"
        assert change["afterText"] == "new_translation"
        assert change["reason"] == "线下回传"


def test_text_template_upload_mismatch_rollback(seed_user):
    with db_cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_mismatch", 14001, 1, "src_mismatch", "old_mismatch", 1, 0),
        )
        text_id = cursor.lastrowid

    upload_bytes = _build_xlsx(
        [
            [text_id, "file_mismatch", 14001, 9, "src_mismatch", "new_mismatch", 2],
        ]
    )

    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    response = client.post(
        "/texts/upload?fileName=tmp_mismatch.xlsx",
        headers=headers,
        content=upload_bytes,
    )
    assert response.status_code == 400
    assert "校验失败" in response.json()["message"]

    with db_cursor() as cursor:
        cursor.execute('SELECT "translatedText", status, "editCount" FROM text_main WHERE id = %s', (text_id,))
        row = cursor.fetchone()
        assert row["translatedText"] == "old_mismatch"
        assert row["status"] == 1
        assert row["editCount"] == 0

        cursor.execute('SELECT COUNT(*) AS total FROM text_changes WHERE "textId" = %s', (text_id,))
        assert cursor.fetchone()["total"] == 0


def test_text_list_and_download_order_consistent(seed_user):
    with db_cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_order", 10001, 1, "old_src", "old_dst", 1, 0),
        )
        older_id = cursor.lastrowid
        cursor.execute(
            """
            INSERT INTO text_main (fid, "textId", part, "sourceText", "translatedText", status, "editCount")
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            ("file_order", 20002, 1, "new_src", "new_dst", 2, 1),
        )
        newer_id = cursor.lastrowid
        cursor.execute('UPDATE text_main SET "uptTime" = %s WHERE id = %s', ("2026-01-01 00:00:00", older_id))
        cursor.execute('UPDATE text_main SET "uptTime" = %s WHERE id = %s', ("2026-01-02 00:00:00", newer_id))

    client = TestClient(app)
    token = _login(client, seed_user)
    headers = {"Authorization": f"Bearer {token}"}

    list_response = client.get("/texts?fid=file_order&page=1&pageSize=20", headers=headers)
    assert list_response.status_code == 200
    list_payload = list_response.json()["data"]["items"]
    assert len(list_payload) == 2
    assert list_payload[0]["id"] == newer_id

    download_response = client.get("/texts/download?fid=file_order", headers=headers)
    assert download_response.status_code == 200

    workbook = load_workbook(BytesIO(download_response.content), data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][0] == newer_id
    assert rows[1][6] == "修改"
