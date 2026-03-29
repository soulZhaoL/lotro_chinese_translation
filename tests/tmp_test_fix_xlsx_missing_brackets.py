from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from tools.valid_format.fix_xlsx_missing_brackets import (
    _build_missing_colon_patterns,
    _build_missing_close_patterns,
    _build_missing_open_patterns,
    _build_valid_patterns,
    _repair_cell_text,
    _repair_segment,
    run_from_config,
)


pytestmark = pytest.mark.no_db


def _patterns():
    id_pattern = r"\d{2,10}"
    return (
        _build_valid_patterns(id_pattern),
        _build_missing_colon_patterns(id_pattern),
        _build_missing_open_patterns(id_pattern),
        _build_missing_close_patterns(id_pattern),
    )


def test_repair_segment_adds_missing_opening_bracket():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    repaired, repair_kinds = _repair_segment(
        "91111505::::::'我当然来了，孩子！']",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert repair_kinds == ["missing_open"]
    assert repaired == "91111505::::::['我当然来了，孩子！']"


def test_repair_segment_adds_missing_closing_bracket():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    repaired, repair_kinds = _repair_segment(
        "91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！'",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert repair_kinds == ["missing_close"]
    assert repaired == "91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！']"


def test_repair_segment_adds_missing_colon_for_colon6_protocol():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    repaired, repair_kinds = _repair_segment(
        "228870261:::::[Maethad竞技场]",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert repair_kinds == ["missing_colon"]
    assert repaired == "228870261::::::[Maethad竞技场]"


def test_repair_segment_adds_missing_colon_and_closing_bracket_together():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    repaired, repair_kinds = _repair_segment(
        "228870261:::::[Maethad竞技场",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert repair_kinds == ["missing_colon", "missing_close"]
    assert repaired == "228870261::::::[Maethad竞技场]"


def test_repair_cell_text_only_repairs_clear_protocol_bracket_errors():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    result = _repair_cell_text(
        "263655938::::::['我会尽力帮助你。跟我来。']|||228870261:::::[Maethad竞技场]|||91111505::::::'我当然来了，孩子！']|||91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！'",
        "|||",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert result["text"] == (
        "263655938::::::['我会尽力帮助你。跟我来。']|||"
        "228870261::::::[Maethad竞技场]|||"
        "91111505::::::['我当然来了，孩子！']|||"
        "91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！']"
    )
    assert result["repairedMissingColon"] == 1
    assert result["repairedMissingOpen"] == 1
    assert result["repairedMissingClose"] == 1
    assert result["invalidSegments"] == []


def test_repair_cell_text_keeps_unfixable_invalid_segment():
    valid_patterns, missing_colon_patterns, missing_open_patterns, missing_close_patterns = _patterns()

    result = _repair_cell_text(
        "1加斯阿格温98::::::[你未能援助拉达加斯特]",
        "|||",
        valid_patterns,
        missing_colon_patterns,
        missing_open_patterns,
        missing_close_patterns,
        allow_missing_colon=True,
        allow_missing_opening=True,
        allow_missing_closing=True,
    )

    assert result["text"] == "1加斯阿格温98::::::[你未能援助拉达加斯特]"
    assert result["repairedMissingColon"] == 0
    assert result["repairedMissingOpen"] == 0
    assert result["repairedMissingClose"] == 0
    assert result["invalidSegments"] == [
        (1, "other_invalid", "1加斯阿格温98::::::[你未能援助拉达加斯特]")
    ]


def test_run_from_config_repairs_workbook_and_writes_output(tmp_path: Path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "sheet1"
    worksheet["A1"] = "fid"
    worksheet["B1"] = "split_part"
    worksheet["C1"] = "translation"
    worksheet["A2"] = "fid_1"
    worksheet["B2"] = "0"
    worksheet["C2"] = (
        "263655938::::::['我会尽力帮助你。跟我来。']|||"
        "228870261:::::[Maethad竞技场]|||"
        "91111505::::::'我当然来了，孩子！']|||"
        "91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！'"
    )
    input_path = tmp_path / "input.xlsx"
    workbook.save(input_path)

    output_path = tmp_path / "output.xlsx"
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        "\n".join(
            [
                f'base_dir: "{tmp_path}"',
                "",
                "input:",
                '  path: "input.xlsx"',
                '  sheet: "sheet1"',
                "  row_start: 2",
                '  row_end: "max"',
                '  column: "C"',
                "",
                "parsing:",
                '  splitDelimiter: "|||"',
                '  idPattern: "\\\\d{2,10}"',
                "",
                "repair:",
                "  allowMissingColonInColon6Format: true",
                "  allowMissingOpeningBracket: true",
                "  allowMissingClosingBracket: true",
                "",
                "output:",
                '  path: "output.xlsx"',
                "  overwrite: true",
                "",
                "behavior:",
                "  skipBlankCells: true",
                "  maxSamplesPerKind: 3",
                "",
            ]
        ),
        encoding="utf-8",
    )

    summary = run_from_config(config_path)

    assert summary["changedCells"] == 1
    assert summary["repairedMissingColon"] == 1
    assert summary["repairedMissingOpen"] == 1
    assert summary["repairedMissingClose"] == 1
    assert summary["overflowBlockedCells"] == 0
    assert summary["overflowBlockedSegments"] == 0

    fixed_workbook = load_workbook(output_path)
    fixed_value = fixed_workbook["sheet1"]["C2"].value
    assert fixed_value == (
        "263655938::::::['我会尽力帮助你。跟我来。']|||"
        "228870261::::::[Maethad竞技场]|||"
        "91111505::::::['我当然来了，孩子！']|||"
        "91111506::::::['你不该这么匆忙地来这里！我告诉过你这是愚蠢的！']"
    )


def test_run_from_config_reports_excel_length_overflow(tmp_path: Path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "sheet1"
    worksheet["A1"] = "fid"
    worksheet["B1"] = "split_part"
    worksheet["C1"] = "translation"
    worksheet["A2"] = "fid_1"
    worksheet["B2"] = "0"
    worksheet["C2"] = "12::::::[" + ("甲" * (32767 - len("12::::::[")))
    input_path = tmp_path / "input.xlsx"
    workbook.save(input_path)

    output_path = tmp_path / "output.xlsx"
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        "\n".join(
            [
                f'base_dir: "{tmp_path}"',
                "",
                "input:",
                '  path: "input.xlsx"',
                '  sheet: "sheet1"',
                "  row_start: 2",
                '  row_end: "max"',
                '  column: "C"',
                "",
                "parsing:",
                '  splitDelimiter: "|||"',
                '  idPattern: "\\\\d{2,10}"',
                "",
                "repair:",
                "  allowMissingColonInColon6Format: true",
                "  allowMissingOpeningBracket: true",
                "  allowMissingClosingBracket: true",
                "",
                "output:",
                '  path: "output.xlsx"',
                "  overwrite: true",
                "",
                "behavior:",
                "  skipBlankCells: true",
                "  maxSamplesPerKind: 3",
                "",
            ]
        ),
        encoding="utf-8",
    )

    summary = run_from_config(config_path)

    assert summary["changedCells"] == 0
    assert summary["repairedMissingColon"] == 0
    assert summary["repairedMissingOpen"] == 0
    assert summary["repairedMissingClose"] == 0
    assert summary["overflowBlockedCells"] == 1
    assert summary["overflowBlockedSegments"] == 1
    assert summary["remainingInvalidCounts"]["missing_close"] == 1

    fixed_workbook = load_workbook(output_path)
    fixed_value = fixed_workbook["sheet1"]["C2"].value
    assert fixed_value == worksheet["C2"].value
