import os
import re
import sqlite3
import threading
import queue
import tkinter as tk
from tkinter import filedialog, messagebox

from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# ============================================================
# SQLite helpers
# ============================================================
def get_single_table_name(db_path: str) -> str:
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type='table' AND name NOT LIKE 'sqlite_%' "
            "ORDER BY name"
        )
        rows = cur.fetchall()
        if not rows:
            raise RuntimeError(f"DB无可用表: {db_path}")
        if len(rows) > 1:
            raise RuntimeError(
                f"DB中发现多张表（请确保只有一张表）: {db_path}\n{[r[0] for r in rows]}"
            )
        return rows[0][0]
    finally:
        conn.close()


def load_fid_text_map(db_path: str) -> dict:
    table = get_single_table_name(db_path)
    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT fid, text_data FROM {table}")
        data = {}
        for fid, text in cur.fetchall():
            data[str(fid)] = "" if text is None else str(text)
        return data
    finally:
        conn.close()


def is_blank_text(s: str) -> bool:
    return s is None or str(s).strip() == ""


# ============================================================
# Splitting helpers
# ============================================================
def split_by_delimiter_smart(text: str, limit: int, delimiter="|||"):
    """
    先按 delimiter 分段，尽量不在段中间切断。
    若某段本身超过 limit，则对该段进行硬切。
    返回 list[str]
    """
    if text is None:
        text = ""

    if limit <= 0:
        return [text]

    parts = text.split(delimiter)
    chunks = []
    current = ""

    def flush():
        nonlocal current
        if current != "":
            chunks.append(current)
            current = ""

    for i, p in enumerate(parts):
        piece = p if i == 0 else (delimiter + p)

        if len(piece) > limit:
            flush()
            start = 0
            while start < len(piece):
                chunks.append(piece[start:start + limit])
                start += limit
            continue

        if current == "":
            current = piece
        else:
            if len(current) + len(piece) <= limit:
                current += piece
            else:
                flush()
                current = piece

    flush()
    return chunks if chunks else [""]


# ============================================================
# Diff helpers
# ============================================================
DELIM = "|||"
BRACKET_RE = re.compile(r"[[^[]]*]")  # 不支持嵌套


def diff_added_removed(old_text: str, new_text: str, delimiter=DELIM):
    old_parts = [p for p in (old_text or "").split(delimiter) if p != ""]
    new_parts = [p for p in (new_text or "").split(delimiter) if p != ""]
    old_set = set(old_parts)
    new_set = set(new_parts)

    added = [p for p in new_parts if p not in old_set]
    removed = [p for p in old_parts if p not in new_set]
    return added, removed


def build_changes_summary(old_text: str, new_text: str, change_type: str) -> str:
    """
    changes 列（按你的要求）：
    - 新增：输出“添加了什么”（added）
    - 修改：输出“最终变更后的内容”（added）；若同段内部微调导致 added 为空，给提示看标红
    - 删除：空
    """
    if change_type == "删除":
        return ""

    added, _removed = diff_added_removed(old_text, new_text, delimiter=DELIM)
    if added:
        return DELIM.join(added)

    if (old_text or "") != (new_text or ""):
        return "(有内容变更，详见 new_text_data 标红处)"
    return ""


# ============================================================
# Rich text highlighting: bracket-inner char diff
# ============================================================
def _append_text(rich: CellRichText, s: str, font: InlineFont):
    if s:
        rich.append(TextBlock(text=s, font=font))


def _highlight_changed_ranges_by_opcodes(base_text: str, other_text: str, which: str):
    """
    which:
      - 'old': 标红 old 中被删/被替换的部分
      - 'new': 标红 new 中新增/替换的部分
    """
    sm = SequenceMatcher(a=base_text, b=other_text, autojunk=False)
    ranges = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if which == "old":
            if tag in ("delete", "replace") and i2 > i1:
                ranges.append((i1, i2))
        else:
            if tag in ("insert", "replace") and i2 > i1:
                ranges.append((i1, i2))
    if not ranges:
        return []
    ranges.sort()
    merged = [ranges[0]]
    for s, e in ranges[1:]:
        ps, pe = merged[-1]
        if s <= pe:
            merged[-1] = (ps, max(pe, e))
        else:
            merged.append((s, e))
    return merged


def _render_with_red_ranges(text: str, red_ranges, black: InlineFont, red: InlineFont):
    if not text or not red_ranges:
        return text

    rich = CellRichText()
    pos = 0
    for s, e in red_ranges:
        if s > pos:
            _append_text(rich, text[pos:s], black)
        _append_text(rich, text[s:e], red)
        pos = e
    if pos < len(text):
        _append_text(rich, text[pos:], black)

    return rich


def build_rich_text_highlight_bracket_inner_changes(old_text: str, new_text: str, which: str):
    """
    仅对每个 '[...]' 内部做精细 diff；括号外不做全文diff（更快更稳）。
    """
    base = old_text if which == "old" else new_text
    other = new_text if which == "old" else old_text

    if not base:
        return ""

    base_matches = list(BRACKET_RE.finditer(base))
    if not base_matches:
        return base

    black = InlineFont(color="000000")
    red = InlineFont(color="FF0000")

    rich = CellRichText()
    pos = 0
    other_brackets = [m.group(0) for m in BRACKET_RE.finditer(other)]

    for idx, m in enumerate(base_matches):
        if m.start() > pos:
            _append_text(rich, base[pos:m.start()], black)

        seg = m.group(0)  # "[...]"
        inner = seg[1:-1]

        other_seg = other_brackets[idx] if idx < len(other_brackets) else ""
        other_inner = other_seg[1:-1] if (other_seg.startswith("[") and other_seg.endswith("]")) else ""

        red_ranges = _highlight_changed_ranges_by_opcodes(inner, other_inner, which=which)

        _append_text(rich, "[", black)
        rendered_inner = _render_with_red_ranges(inner, red_ranges, black=black, red=red)
        if isinstance(rendered_inner, CellRichText):
            for tb in rendered_inner:
                rich.append(tb)
        else:
            _append_text(rich, rendered_inner, black)
        _append_text(rich, "]", black)

        pos = m.end()

    if pos < len(base):
        _append_text(rich, base[pos:], black)

    # 若最终没有红色，返回普通字符串（减小文件 & 提速）
    for tb in rich:
        c = getattr(tb.font, "color", None)
        rgb = getattr(c, "rgb", None) if c is not None else None
        if rgb == "FF0000":
            return rich
    return base


# ============================================================
# Excel export
# ============================================================
HEADERS = [
    "fid",
    "split_part",
    "change_type",
    "old_text_data",
    "new_text_data",
    "changes",
]


def export_to_excel(rows, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "diff"

    ws.append(HEADERS)
    header_font = Font(bold=True)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for r in rows:
        ws.append([""] * len(HEADERS))
        row_idx = ws.max_row

        for col_idx, key in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.value = r.get(key, "")

    widths = {"A": 18, "B": 10, "C": 10, "D": 70, "E": 70, "F": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(out_path)


# ============================================================
# Core compare
# ============================================================
def compare_maps(old_map: dict, new_map: dict, split_limit: int):
    rows = []

    old_keys = set(old_map.keys())
    new_keys = set(new_map.keys())

    added_fids = sorted(new_keys - old_keys)
    deleted_fids = sorted(old_keys - new_keys)
    common_fids = sorted(old_keys & new_keys)

    # 新增
    for fid in added_fids:
        new_text = new_map.get(fid, "")
        if is_blank_text(new_text):
            continue

        parts = split_by_delimiter_smart(new_text, split_limit)
        split_parts = [0] if (len(parts) == 1) else list(range(1, len(parts) + 1))
        changes = build_changes_summary("", new_text, "新增")

        for sp in split_parts:
            rows.append({
                "fid": fid,
                "split_part": sp,
                "change_type": "新增",
                "old_text_data": "",
                "new_text_data": new_text,
                "changes": changes,
            })

    # 修改
    for fid in common_fids:
        old_text = old_map.get(fid, "")
        new_text = new_map.get(fid, "")

        if is_blank_text(old_text) or is_blank_text(new_text):
            continue
        if old_text == new_text:
            continue

        old_rich = build_rich_text_highlight_bracket_inner_changes(old_text, new_text, which="old")
        new_rich = build_rich_text_highlight_bracket_inner_changes(old_text, new_text, which="new")

        parts = split_by_delimiter_smart(new_text, split_limit)
        split_parts = [0] if (len(parts) == 1) else list(range(1, len(parts) + 1))
        changes = build_changes_summary(old_text, new_text, "修改")

        for sp in split_parts:
            rows.append({
                "fid": fid,
                "split_part": sp,
                "change_type": "修改",
                "old_text_data": old_rich,
                "new_text_data": new_rich,
                "changes": changes,
            })

    # 删除
    for fid in deleted_fids:
        old_text = old_map.get(fid, "")
        if is_blank_text(old_text):
            continue

        parts = split_by_delimiter_smart(old_text, split_limit)
        split_parts = [0] if (len(parts) == 1) else list(range(1, len(parts) + 1))

        for sp in split_parts:
            rows.append({
                "fid": fid,
                "split_part": sp,
                "change_type": "删除",
                "old_text_data": old_text,
                "new_text_data": "",
                "changes": "",
            })

    return rows


# ============================================================
# GUI
# ============================================================
def suggested_output_path(old_db_path: str, new_db_path: str) -> str:
    old_base = os.path.splitext(os.path.basename(old_db_path or ""))[0] or "old"
    new_base = os.path.splitext(os.path.basename(new_db_path or ""))[0] or "new"
    dirname = os.path.dirname(old_db_path) if old_db_path else os.getcwd()
    filename = f"diff__{old_base}__VS__{new_base}.xlsx"
    return os.path.join(dirname, filename)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DB text_data 对比导出 Excel")
        self.geometry("1020x520")
        self.minsize(900, 480)

        self.old_db = tk.StringVar()
        self.new_db = tk.StringVar()
        self.out_xlsx = tk.StringVar()
        self.split_limit = tk.StringVar(value="25000")

        self._out_auto = True
        self._running = False
        self._uiq = queue.Queue()

        pad = {"padx": 10, "pady": 6}

        tk.Label(self, text="旧DB路径：").grid(row=0, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.old_db, width=110).grid(row=0, column=1, sticky="we", **pad)
        tk.Button(self, text="选择...", command=self.pick_old).grid(row=0, column=2, **pad)

        tk.Label(self, text="新DB路径：").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.new_db, width=110).grid(row=1, column=1, sticky="we", **pad)
        tk.Button(self, text="选择...", command=self.pick_new).grid(row=1, column=2, **pad)

        tk.Label(self, text="输出Excel：").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.out_xlsx, width=110).grid(row=2, column=1, sticky="we", **pad)
        tk.Button(self, text="选择...", command=self.pick_out).grid(row=2, column=2, **pad)

        tk.Label(self, text="分割上限(字符数)：").grid(row=3, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.split_limit, width=20).grid(row=3, column=1, sticky="w", **pad)
        tk.Label(self, text="默认25000；建议≤32767").grid(row=3, column=1, sticky="w", padx=220, pady=6)

        self.status = tk.Label(self, text="就绪", anchor="w")
        self.status.grid(row=4, column=0, columnspan=2, sticky="we", padx=10, pady=6)

        self.run_btn = tk.Button(self, text="开始比较并导出", command=self.run, height=2)
        self.run_btn.grid(row=4, column=2, sticky="e", padx=10, pady=6)

        self.info = tk.Text(self, height=14, width=140)
        self.info.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=10, pady=10)
        self.info.insert(
            "end",
            "说明：\n"
            "- 比较 fid + text_data\n"
            "- 修改：仅对每个 [ ... ] 内部做精细diff，新旧分别把变化部分标红\n"
            "- split_part：不分割=0；分割后从1开始编号\n"
            "- text_data为空的fid会被跳过不输出\n"
            "- changes列：只输出最终新增/变更后的内容（不输出删除内容，不加 +/-）\n"
            "- 采用后台线程，避免界面卡死\n"
        )
        self.info.configure(state="disabled")

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(5, weight=1)

        self.old_db.trace_add("write", lambda *_: self._maybe_update_default_out())
        self.new_db.trace_add("write", lambda *_: self._maybe_update_default_out())

        # 轮询后台线程消息
        self.after(100, self._poll_queue)

    def _set_running(self, running: bool):
        self._running = running
        self.run_btn.config(state=("disabled" if running else "normal"))
        self.status.config(text=("运行中..." if running else "就绪"))

    def _poll_queue(self):
        try:
            while True:
                msg = self._uiq.get_nowait()
                kind = msg.get("kind")
                if kind == "log":
                    self._log(msg["text"])
                elif kind == "status":
                    self.status.config(text=msg["text"])
                elif kind == "done":
                    self._set_running(False)
                    out_p = msg.get("out_path")
                    if out_p:
                        messagebox.showinfo("完成", f"导出完成：\n{out_p}")
                elif kind == "error":
                    self._set_running(False)
                    messagebox.showerror("运行失败", msg.get("text", "未知错误"))
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _log(self, s: str):
        self.info.configure(state="normal")
        self.info.insert("end", s + "\n")
        self.info.see("end")
        self.info.configure(state="disabled")

    def _maybe_update_default_out(self):
        if not self._out_auto:
            return
        old_p = self.old_db.get().strip()
        new_p = self.new_db.get().strip()
        if old_p and new_p:
            self.out_xlsx.set(suggested_output_path(old_p, new_p))

    def pick_old(self):
        if self._running:
            return
        p = filedialog.askopenfilename(
            title="选择旧DB",
            filetypes=[("SQLite DB", "*.db *.sqlite *.sqlite3"), ("All", "*.*")]
        )
        if p:
            self.old_db.set(p)
            self._out_auto = True
            self._maybe_update_default_out()

    def pick_new(self):
        if self._running:
            return
        p = filedialog.askopenfilename(
            title="选择新DB",
            filetypes=[("SQLite DB", "*.db *.sqlite *.sqlite3"), ("All", "*.*")]
        )
        if p:
            self.new_db.set(p)
            self._out_auto = True
            self._maybe_update_default_out()

    def pick_out(self):
        if self._running:
            return
        old_p = self.old_db.get().strip()
        new_p = self.new_db.get().strip()
        initialfile = None
        initialdir = None
        if old_p and new_p:
            sug = suggested_output_path(old_p, new_p)
            initialdir = os.path.dirname(sug)
            initialfile = os.path.basename(sug)

        p = filedialog.asksaveasfilename(
            title="选择输出Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=initialdir,
            initialfile=initialfile
        )
        if p:
            self.out_xlsx.set(p)
            self._out_auto = False

    def run(self):
        if self._running:
            return

        old_p = self.old_db.get().strip()
        new_p = self.new_db.get().strip()
        out_p = self.out_xlsx.get().strip()

        if not old_p or not os.path.exists(old_p):
            messagebox.showerror("错误", "请选择有效的旧DB路径")
            return
        if not new_p or not os.path.exists(new_p):
            messagebox.showerror("错误", "请选择有效的新DB路径")
            return
        if not out_p:
            out_p = suggested_output_path(old_p, new_p)
            self.out_xlsx.set(out_p)

        try:
            limit = int(self.split_limit.get().strip())
        except Exception:
            messagebox.showerror("错误", "分割上限必须是整数")
            return

        self._set_running(True)
        self._uiq.put({"kind": "log", "text": "开始..."})

        def worker():
            try:
                self._uiq.put({"kind": "status", "text": "读取旧DB..."})
                self._uiq.put({"kind": "log", "text": "读取旧DB..."})
                old_map = load_fid_text_map(old_p)
                self._uiq.put({"kind": "log", "text": f"旧DB记录数: {len(old_map)}"})

                self._uiq.put({"kind": "status", "text": "读取新DB..."})
                self._uiq.put({"kind": "log", "text": "读取新DB..."})
                new_map = load_fid_text_map(new_p)
                self._uiq.put({"kind": "log", "text": f"新DB记录数: {len(new_map)}"})

                self._uiq.put({"kind": "status", "text": "比较差异..."})
                self._uiq.put({"kind": "log", "text": "比较差异..."})
                rows = compare_maps(old_map, new_map, limit)
                self._uiq.put({"kind": "log", "text": f"输出行数（含拆分）: {len(rows)}"})

                self._uiq.put({"kind": "status", "text": "写入Excel..."})
                self._uiq.put({"kind": "log", "text": "写入Excel..."})
                export_to_excel(rows, out_p)

                self._uiq.put({"kind": "log", "text": f"完成：{out_p}"})
                self._uiq.put({"kind": "done", "out_path": out_p})
            except Exception as e:
                self._uiq.put({"kind": "error", "text": str(e)})

        t = threading.Thread(target=worker, daemon=True)
        t.start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
