import os
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from copy import copy as pycopy

import pandas as pd
import openpyxl


# ----------------------------
# 读取与规范化
# ----------------------------
def read_excel_keep_str(path: str, sheet_name=None) -> pd.DataFrame:
    x = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    if isinstance(x, dict):
        if not x:
            raise ValueError("Excel里没有可读取的sheet")
        df = x[next(iter(x.keys()))]
    else:
        df = x
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    return df


def normalize_series(s: pd.Series) -> pd.Series:
    s = s.astype("string")
    s = s.str.strip()
    s = s.replace({"": pd.NA})
    return s


def normalize_split_part(s: pd.Series) -> pd.Series:
    s = normalize_series(s)

    def _fix(x):
        if x is pd.NA or x is None:
            return pd.NA
        t = str(x).strip()
        if t == "":
            return pd.NA
        # "1.0" -> "1"
        if t.endswith(".0") and t.replace(".", "", 1).isdigit():
            t2 = t[:-2]
            if t2.isdigit():
                return t2
        return t

    return s.map(_fix).astype("string")


def ensure_cols(df: pd.DataFrame, cols, label: str):
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"{label} 缺少列：{missing}；现有列：{list(df.columns)}")


# ----------------------------
# 样式复制：仅复制 sheet1 中 status=一致 的 text_data 单元格样式
# ----------------------------
def _ws_header_to_col(ws):
    """返回 header_name -> col_index 的映射（从第1行读表头）。"""
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        if isinstance(v, str):
            k = v.strip()
        else:
            k = str(v).strip()
        if k:
            m[k] = c
    return m


def _norm_key_cell(v):
    """把单元格值规范成用于匹配的 key 字符串。"""
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def apply_text_data_style_for_same_rows(
    old_path: str,
    out_path: str,
    *,
    old_sheet: str | None,
    out_sheet: str = "sheet1",
    fid_col: str = "fid",
    split_part_col: str = "split_part",
    text_col: str = "text_data",
    status_col: str = "status",
    same_status_value: str = "一致",
):
    """
    把旧Excel中 text_data 单元格的样式复制到输出Excel：
    - 仅对输出 sheet1 里 status==一致 的行
    - 按 (fid, split_part) 精确匹配旧表行
    """
    wb_old = openpyxl.load_workbook(old_path, data_only=True)
    ws_old = wb_old[old_sheet] if old_sheet else wb_old.active

    wb_out = openpyxl.load_workbook(out_path)
    if out_sheet not in wb_out.sheetnames:
        wb_old.close()
        wb_out.close()
        raise ValueError(f"输出文件中不存在工作表：{out_sheet}")
    ws_out = wb_out[out_sheet]

    old_h = _ws_header_to_col(ws_old)
    out_h = _ws_header_to_col(ws_out)

    # 检查必要列
    for col in (fid_col, split_part_col, text_col):
        if col not in old_h:
            wb_old.close()
            wb_out.close()
            raise ValueError(f"旧Excel中找不到列：{col}（请确认表头）")
    for col in (fid_col, split_part_col, text_col, status_col):
        if col not in out_h:
            wb_old.close()
            wb_out.close()
            raise ValueError(f"输出Excel的 {out_sheet} 中找不到列：{col}（请确认表头）")

    # 建旧表索引： (fid, split_part) -> old_row_idx
    old_index = {}
    dup_keys = set()

    for r in range(2, ws_old.max_row + 1):
        fid_v = _norm_key_cell(ws_old.cell(r, old_h[fid_col]).value)
        sp_v = _norm_key_cell(ws_old.cell(r, old_h[split_part_col]).value)
        if fid_v is None or sp_v is None:
            continue
        key = (fid_v, sp_v)
        if key in old_index:
            dup_keys.add(key)
        old_index[key] = r

    if dup_keys:
        # 旧表理论上你已在 pandas 里检查过唯一性，这里再兜底
        wb_old.close()
        wb_out.close()
        sample = list(dup_keys)[:10]
        raise ValueError(f"旧Excel存在重复 (fid, split_part) 键，示例：{sample}")

    out_fid_c = out_h[fid_col]
    out_sp_c = out_h[split_part_col]
    out_text_c = out_h[text_col]
    out_status_c = out_h[status_col]
    old_text_c = old_h[text_col]

    styled_count = 0
    miss_count = 0

    for r in range(2, ws_out.max_row + 1):
        status = ws_out.cell(r, out_status_c).value
        if status != same_status_value:
            continue

        fid_v = _norm_key_cell(ws_out.cell(r, out_fid_c).value)
        sp_v = _norm_key_cell(ws_out.cell(r, out_sp_c).value)
        if fid_v is None or sp_v is None:
            continue
        key = (fid_v, sp_v)

        old_r = old_index.get(key)
        if old_r is None:
            miss_count += 1
            continue

        src = ws_old.cell(old_r, old_text_c)
        dst = ws_out.cell(r, out_text_c)

        # 复制整套样式（底色/字体颜色等都会保留）
        dst._style = pycopy(src._style)
        dst.font = pycopy(src.font)
        dst.fill = pycopy(src.fill)
        dst.border = pycopy(src.border)
        dst.alignment = pycopy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = pycopy(src.protection)

        styled_count += 1

    wb_out.save(out_path)
    wb_old.close()
    wb_out.close()
    return styled_count, miss_count


# ----------------------------
# 核心逻辑
# ----------------------------
def build_two_sheet_output(
    old_path,
    new_path,
    out_path,
    old_sheet=None,
    new_sheet=None,
    progress_cb=None,
    log_cb=None,
):
    def log(msg):
        if log_cb:
            log_cb(msg)

    def progress(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    progress(1, "读取Excel中...")
    log(f"旧Excel: {old_path}")
    log(f"新Excel: {new_path}")
    log(f"输出: {out_path}")

    old_df = read_excel_keep_str(old_path, sheet_name=old_sheet if old_sheet else None)
    new_df = read_excel_keep_str(new_path, sheet_name=new_sheet if new_sheet else None)

    ensure_cols(old_df, ["fid", "split_part", "text_data", "translation"], "旧表")
    ensure_cols(new_df, ["fid", "split_part", "text_data"], "新表")

    progress(10, "规范化数据...")

    old = old_df[["fid", "split_part", "text_data", "translation"]].copy()
    new = new_df[["fid", "split_part", "text_data"]].copy()

    old["fid"] = normalize_series(old["fid"])
    new["fid"] = normalize_series(new["fid"])

    old["split_part"] = normalize_split_part(old["split_part"])
    new["split_part"] = normalize_split_part(new["split_part"])

    old["text_data"] = normalize_series(old["text_data"])
    new["text_data"] = normalize_series(new["text_data"])

    old["translation"] = normalize_series(old["translation"])

    old = old[old["fid"].notna() & old["split_part"].notna()].copy()
    new = new[new["fid"].notna() & new["split_part"].notna()].copy()

    progress(18, "检查 (fid, split_part) 唯一性...")

    if old.duplicated(subset=["fid", "split_part"], keep=False).any():
        sample = (
            old.loc[old.duplicated(["fid", "split_part"], keep=False), ["fid", "split_part"]]
            .drop_duplicates()
            .head(20)
        )
        raise ValueError("旧表存在重复键 (fid, split_part)：\n" + sample.to_string(index=False))

    if new.duplicated(subset=["fid", "split_part"], keep=False).any():
        sample = (
            new.loc[new.duplicated(["fid", "split_part"], keep=False), ["fid", "split_part"]]
            .drop_duplicates()
            .head(20)
        )
        raise ValueError("新表存在重复键 (fid, split_part)：\n" + sample.to_string(index=False))

    progress(30, "按 (fid, split_part) 对比新旧差异...")

    merged = old.merge(
        new,
        on=["fid", "split_part"],
        how="outer",
        suffixes=("_old", "_new"),
        indicator=True,
    )

    both = merged[merged["_merge"] == "both"].copy()
    added = merged[merged["_merge"] == "right_only"].copy()
    deleted = merged[merged["_merge"] == "left_only"].copy()

    # 忽略大小写比较：只有大小写差异 => 认为一致
    def norm_casefold(x):
        if x is None or x is pd.NA:
            return "<NA>"
        return str(x).casefold()

    old_cf = both["text_data_old"].map(norm_casefold)
    new_cf = both["text_data_new"].map(norm_casefold)

    same = both[old_cf == new_cf].copy()
    modified = both[old_cf != new_cf].copy()

    def sort_block(df):
        tmp = df.copy()
        tmp["_fid_num"] = pd.to_numeric(tmp["fid"], errors="coerce")
        tmp["_sp_num"] = pd.to_numeric(tmp["split_part"], errors="coerce")
        tmp = tmp.sort_values(
            by=["_fid_num", "fid", "_sp_num", "split_part"],
            ascending=[True, True, True, True],
            kind="stable",
            na_position="last",
        ).drop(columns=["_fid_num", "_sp_num"])
        return tmp

    # ----------------------------
    # Sheet1：一致(旧) -> 新增(新) -> 修改(新)，加 status
    # ----------------------------
    progress(55, "生成Sheet1...")

    sheet1_same = same[["fid", "split_part", "text_data_old", "translation"]].rename(
        columns={"text_data_old": "text_data"}
    )
    sheet1_same["status"] = "一致"

    sheet1_added = added[["fid", "split_part", "text_data_new"]].rename(
        columns={"text_data_new": "text_data"}
    )
    sheet1_added["translation"] = pd.NA
    sheet1_added["status"] = "新增"

    sheet1_modified = modified[["fid", "split_part", "text_data_new"]].rename(
        columns={"text_data_new": "text_data"}
    )
    sheet1_modified["translation"] = pd.NA
    sheet1_modified["status"] = "修改"

    sheet1 = pd.concat(
        [sort_block(sheet1_same), sort_block(sheet1_added), sort_block(sheet1_modified)],
        ignore_index=True,
    )
    sheet1 = sheet1[["fid", "split_part", "text_data", "translation", "status"]]

    # ----------------------------
    # Sheet2：修改(含旧/新/translation) -> 删除(含旧/translation)，加 status
    # ----------------------------
    progress(75, "生成Sheet2...")

    s2_mod = modified[["fid", "split_part", "text_data_old", "text_data_new", "translation"]].copy()
    s2_mod = s2_mod.rename(columns={"text_data_old": "old_text_data", "text_data_new": "new_text_data"})
    s2_mod["status"] = "修改"

    s2_del = deleted[["fid", "split_part", "text_data_old", "translation"]].copy()
    s2_del = s2_del.rename(columns={"text_data_old": "old_text_data"})
    s2_del["new_text_data"] = pd.NA
    s2_del["status"] = "删除"

    sheet2 = pd.concat([sort_block(s2_mod), sort_block(s2_del)], ignore_index=True)
    sheet2 = sheet2[["fid", "split_part", "old_text_data", "new_text_data", "translation", "status"]]

    log(f"一致(忽略大小写): {len(sheet1_same)}")
    log(f"新增: {len(sheet1_added)}")
    log(f"修改: {len(sheet1_modified)}")
    log(f"删除: {len(s2_del)}")

    progress(90, "写出结果Excel...")
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet1.to_excel(writer, index=False, sheet_name="sheet1")
        sheet2.to_excel(writer, index=False, sheet_name="sheet2")

    # ----------------------------
    # 样式修补：把“一致部分”的 text_data 单元格样式从旧表复制到输出 sheet1
    # ----------------------------
    progress(96, "复制一致部分 text_data 样式...")
    styled_count, miss_count = apply_text_data_style_for_same_rows(
        old_path=old_path,
        out_path=out_path,
        old_sheet=old_sheet,   # None=旧文件 active sheet；如果你指定了旧sheet名，这里会跟着用
        out_sheet="sheet1",
        fid_col="fid",
        split_part_col="split_part",
        text_col="translation",
        status_col="status",
        same_status_value="一致",
    )
    log(f"样式复制完成：复制 {styled_count} 条；未匹配到旧表key {miss_count} 条")

    progress(100, "完成")
    log("完成：已生成两张sheet的输出Excel。")


# ----------------------------
# GUI
# ----------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel合并输出：一致/新增/修改 + 修改/删除明细")
        self.geometry("900x600")

        self.old_path = tk.StringVar()
        self.new_path = tk.StringVar()
        self.out_path = tk.StringVar()

        self.old_sheet = tk.StringVar()
        self.new_sheet = tk.StringVar()

        self._build_ui()
        self.worker = None

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        frm_paths = ttk.LabelFrame(self, text="文件路径")
        frm_paths.pack(fill="x", **pad)

        ttk.Label(frm_paths, text="旧Excel：").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_paths, textvariable=self.old_path).grid(row=0, column=1, sticky="ew", **pad)
        ttk.Button(frm_paths, text="选择...", command=self.pick_old).grid(row=0, column=2, **pad)

        ttk.Label(frm_paths, text="新Excel：").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_paths, textvariable=self.new_path).grid(row=1, column=1, sticky="ew", **pad)
        ttk.Button(frm_paths, text="选择...", command=self.pick_new).grid(row=1, column=2, **pad)

        ttk.Label(frm_paths, text="输出Excel：").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm_paths, textvariable=self.out_path).grid(row=2, column=1, sticky="ew", **pad)
        ttk.Button(frm_paths, text="选择...", command=self.pick_out).grid(row=2, column=2, **pad)

        frm_paths.columnconfigure(1, weight=1)

        frm_opts = ttk.LabelFrame(self, text="可选参数")
        frm_opts.pack(fill="x", **pad)

        ttk.Label(frm_opts, text="旧sheet(可选)：").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_opts, textvariable=self.old_sheet).grid(row=0, column=1, sticky="ew", **pad)

        ttk.Label(frm_opts, text="新sheet(可选)：").grid(row=0, column=2, sticky="w", **pad)
        ttk.Entry(frm_opts, textvariable=self.new_sheet).grid(row=0, column=3, sticky="ew", **pad)

        frm_opts.columnconfigure(1, weight=1)
        frm_opts.columnconfigure(3, weight=1)

        frm_run = ttk.Frame(self)
        frm_run.pack(fill="x", **pad)

        self.btn_run = ttk.Button(frm_run, text="生成输出Excel", command=self.on_run)
        self.btn_run.pack(side="left")

        self.btn_clear = ttk.Button(frm_run, text="清空日志", command=self.clear_log)
        self.btn_clear.pack(side="left", padx=10)

        self.prog_msg = tk.StringVar(value="等待开始")
        ttk.Label(frm_run, textvariable=self.prog_msg).pack(side="left", padx=10)

        self.prog = ttk.Progressbar(self, mode="determinate", maximum=100)
        self.prog.pack(fill="x", padx=10, pady=4)

        frm_log = ttk.LabelFrame(self, text="日志")
        frm_log.pack(fill="both", expand=True, **pad)

        self.txt = tk.Text(frm_log, wrap="word")
        self.txt.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frm_log, command=self.txt.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt.configure(yscrollcommand=scrollbar.set)

    def log(self, msg: str):
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")

    def clear_log(self):
        self.txt.delete("1.0", "end")

    def set_progress(self, pct: int, msg: str):
        self.prog["value"] = max(0, min(100, pct))
        self.prog_msg.set(msg)

    def pick_old(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.old_path.set(p)
            if not self.out_path.get():
                base = os.path.splitext(os.path.basename(p))[0]
                self.out_path.set(os.path.join(os.path.dirname(p), f"{base}_merged.xlsx"))

    def pick_new(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.new_path.set(p)

    def pick_out(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if p:
            self.out_path.set(p)

    def on_run(self):
        old_path = self.old_path.get().strip()
        new_path = self.new_path.get().strip()
        out_path = self.out_path.get().strip()

        if not old_path or not os.path.exists(old_path):
            messagebox.showerror("错误", "请选择有效的旧Excel路径")
            return
        if not new_path or not os.path.exists(new_path):
            messagebox.showerror("错误", "请选择有效的新Excel路径")
            return
        if not out_path:
            messagebox.showerror("错误", "请选择输出Excel路径")
            return

        old_sheet = self.old_sheet.get().strip() or None
        new_sheet = self.new_sheet.get().strip() or None

        self.set_progress(0, "开始...")
        self.log("========================================")
        self.log("启动任务...")

        self.btn_run.configure(state="disabled")

        def ui_progress(pct, msg):
            self.after(0, lambda: self.set_progress(pct, msg))

        def ui_log(msg):
            self.after(0, lambda: self.log(msg))

        def worker():
            try:
                build_two_sheet_output(
                    old_path=old_path,
                    new_path=new_path,
                    out_path=out_path,
                    old_sheet=old_sheet,
                    new_sheet=new_sheet,
                    progress_cb=ui_progress,
                    log_cb=ui_log,
                )
                self.after(0, lambda: messagebox.showinfo("完成", f"已输出：\n{out_path}"))
            except Exception as e:
                tb = traceback.format_exc()
                self.after(0, lambda: self.log(tb))
                self.after(0, lambda: messagebox.showerror("失败", f"{e}\n\n详情见日志"))
            finally:
                self.after(0, lambda: self.btn_run.configure(state="normal"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
