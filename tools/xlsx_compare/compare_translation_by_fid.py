"""按 fid 聚合 xlsx 中的 translation 并进行严格比对。"""

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple

import yaml
from openpyxl import load_workbook


class ConfigError(Exception):
    pass


@dataclass(frozen=True)
class OrderConfig:
    mode: str
    column: str | None
    value_type: str | None
    require_unique: bool | None


@dataclass(frozen=True)
class FileConfig:
    label: str
    path: Path
    sheet: str
    header_row: int
    data_start_row: int
    key_column: str
    compare_column: str
    order: OrderConfig


@dataclass(frozen=True)
class RowRecord:
    row_index: int
    key_value: str
    compare_value: str
    order_value: Any


@dataclass(frozen=True)
class AggregatedRecord:
    fid: str
    row_count: int
    merged_translation: str


def require_key(data: Dict[str, Any], key: str, path: str) -> Any:
    if key not in data:
        raise ConfigError(f"缺少配置项: {path}{key}")
    return data[key]


def require_type(value: Any, expected_type: type, path: str) -> Any:
    if not isinstance(value, expected_type):
        raise ConfigError(f"配置项类型错误: {path} 期望 {expected_type.__name__}")
    return value


def _find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "config" / "lotro.yaml").exists() or (candidate / ".git").exists():
            return candidate
    raise ConfigError("无法自动定位项目根目录，请在 base_dir 中填写绝对路径")


def _resolve_base_dir(base_dir: str, config_path: Path) -> Path:
    if base_dir == "__PROJECT_ROOT__":
        return _find_project_root(config_path.parent)
    return Path(base_dir).expanduser().resolve()


def _load_yaml_config(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"配置文件不存在: {path}")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ConfigError("配置文件根对象必须为 YAML 对象")
    return data


def _normalize_key_cell(value: Any) -> str:
    if value is None:
        raise ConfigError("fid 不能为空")
    if isinstance(value, bool):
        raise ConfigError("fid 不能为布尔值")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        raise ConfigError(f"fid 不能为非整数浮点数: {value}")
    text = str(value).strip()
    if text == "":
        raise ConfigError("fid 不能为空字符串")
    return text


def _normalize_compare_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _parse_order_value(value: Any, value_type: str, row_index: int, column_name: str) -> Any:
    if value_type == "int":
        if value is None:
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 为空")
        if isinstance(value, bool):
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 不能为布尔值")
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 不是整数: {value}")
        text = str(value).strip()
        if text == "":
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 为空字符串")
        if text.startswith("-"):
            number = text[1:]
            if number.isdigit():
                return int(text)
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 不是整数: {text}")
        if not text.isdigit():
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 不是整数: {text}")
        return int(text)
    if value_type == "string":
        if value is None:
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 为空")
        text = str(value)
        if text == "":
            raise ConfigError(f"第 {row_index} 行排序列 {column_name} 为空字符串")
        return text
    raise ConfigError(f"不支持的排序值类型: {value_type}")


def _validate_order_config(data: Dict[str, Any], path: str) -> OrderConfig:
    mode = require_type(require_key(data, "mode", path), str, f"{path}mode")
    if mode == "document":
        return OrderConfig(mode=mode, column=None, value_type=None, require_unique=None)

    if mode != "column":
        raise ConfigError(f"{path}mode 只能是 document 或 column")

    column = require_type(require_key(data, "column", path), str, f"{path}column")
    value_type = require_type(require_key(data, "value_type", path), str, f"{path}value_type")
    if value_type not in ("int", "string"):
        raise ConfigError(f"{path}value_type 只能是 int 或 string")
    require_unique = require_type(
        require_key(data, "require_unique", path),
        bool,
        f"{path}require_unique",
    )
    return OrderConfig(mode=mode, column=column, value_type=value_type, require_unique=require_unique)


def _validate_file_config(
    label: str,
    data: Dict[str, Any],
    base_dir: Path,
    path: str,
) -> FileConfig:
    relative_path = require_type(require_key(data, "path", path), str, f"{path}path")
    sheet = require_type(require_key(data, "sheet", path), str, f"{path}sheet")
    header_row = require_type(require_key(data, "header_row", path), int, f"{path}header_row")
    data_start_row = require_type(require_key(data, "data_start_row", path), int, f"{path}data_start_row")
    key_column = require_type(require_key(data, "key_column", path), str, f"{path}key_column")
    compare_column = require_type(require_key(data, "compare_column", path), str, f"{path}compare_column")
    order = _validate_order_config(
        require_type(require_key(data, "order", path), dict, f"{path}order"),
        f"{path}order.",
    )

    if header_row <= 0:
        raise ConfigError(f"{path}header_row 必须大于 0")
    if data_start_row <= header_row:
        raise ConfigError(f"{path}data_start_row 必须大于 header_row")

    file_path = (base_dir / relative_path).resolve()
    return FileConfig(
        label=label,
        path=file_path,
        sheet=sheet,
        header_row=header_row,
        data_start_row=data_start_row,
        key_column=key_column,
        compare_column=compare_column,
        order=order,
    )


def load_config(path: Path) -> Tuple[FileConfig, FileConfig, Path, Path]:
    data = _load_yaml_config(path)
    base_dir = _resolve_base_dir(
        require_type(require_key(data, "base_dir", ""), str, "base_dir"),
        path,
    )
    files_cfg = require_type(require_key(data, "files", ""), dict, "files")
    output_cfg = require_type(require_key(data, "output", ""), dict, "output")

    left_cfg = _validate_file_config(
        "left",
        require_type(require_key(files_cfg, "left", "files."), dict, "files.left"),
        base_dir,
        "files.left.",
    )
    right_cfg = _validate_file_config(
        "right",
        require_type(require_key(files_cfg, "right", "files."), dict, "files.right"),
        base_dir,
        "files.right.",
    )

    report_path = (base_dir / require_type(require_key(output_cfg, "report_path", "output."), str, "output.report_path")).resolve()
    summary_path = (base_dir / require_type(require_key(output_cfg, "summary_path", "output."), str, "output.summary_path")).resolve()
    return left_cfg, right_cfg, report_path, summary_path


class XlsxTranslationComparer:
    def __init__(self, left: FileConfig, right: FileConfig, report_path: Path, summary_path: Path):
        self.left = left
        self.right = right
        self.report_path = report_path
        self.summary_path = summary_path

    def run(self) -> int:
        left_records = self._aggregate_file(self.left)
        right_records = self._aggregate_file(self.right)
        only_left, only_right, mismatches = self._compare(left_records, right_records)
        self._write_report(left_records, right_records, only_left, only_right, mismatches)
        self._write_summary(left_records, right_records, only_left, only_right, mismatches)
        if not only_left and not only_right and not mismatches:
            return 0
        return 1

    def _aggregate_file(self, config: FileConfig) -> Dict[str, AggregatedRecord]:
        if not config.path.exists():
            raise FileNotFoundError(f"{config.label} 文件不存在: {config.path}")

        workbook = load_workbook(config.path, read_only=True, data_only=True)
        if config.sheet not in workbook.sheetnames:
            raise ConfigError(f"{config.label} sheet 不存在: {config.sheet}")
        worksheet = workbook[config.sheet]

        header_values = next(
            worksheet.iter_rows(
                min_row=config.header_row,
                max_row=config.header_row,
                values_only=True,
            )
        )
        header_map = self._build_header_map(header_values, config)
        key_index = header_map[config.key_column]
        compare_index = header_map[config.compare_column]
        order_index = None
        if config.order.mode == "column":
            order_index = header_map[config.order.column]  # type: ignore[index]

        grouped: Dict[str, List[RowRecord]] = {}
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=config.data_start_row, values_only=True),
            start=config.data_start_row,
        ):
            key_value = _normalize_key_cell(row[key_index])
            compare_value = _normalize_compare_cell(row[compare_index])
            if config.order.mode == "document":
                order_value = row_index
            else:
                order_value = _parse_order_value(
                    row[order_index],  # type: ignore[index]
                    config.order.value_type,  # type: ignore[arg-type]
                    row_index,
                    config.order.column,  # type: ignore[arg-type]
                )
            grouped.setdefault(key_value, []).append(
                RowRecord(
                    row_index=row_index,
                    key_value=key_value,
                    compare_value=compare_value,
                    order_value=order_value,
                )
            )

        return self._merge_grouped_records(grouped, config)

    def _build_header_map(self, header_values: Tuple[Any, ...], config: FileConfig) -> Dict[str, int]:
        header_map: Dict[str, int] = {}
        for idx, value in enumerate(header_values):
            if value is None:
                continue
            header_name = str(value).strip()
            if header_name == "":
                continue
            if header_name in header_map:
                raise ConfigError(f"{config.label} 表头重复: {header_name}")
            header_map[header_name] = idx

        required_columns = [config.key_column, config.compare_column]
        if config.order.mode == "column":
            required_columns.append(config.order.column)  # type: ignore[arg-type]

        for column in required_columns:
            if column not in header_map:
                raise ConfigError(f"{config.label} 缺少表头列: {column}")
        return header_map

    def _merge_grouped_records(
        self,
        grouped: Dict[str, List[RowRecord]],
        config: FileConfig,
    ) -> Dict[str, AggregatedRecord]:
        aggregated: Dict[str, AggregatedRecord] = {}
        for fid, records in grouped.items():
            if config.order.mode == "document":
                ordered_records = sorted(records, key=lambda item: item.row_index)
            else:
                ordered_records = sorted(records, key=lambda item: (item.order_value, item.row_index))
                if config.order.require_unique:
                    seen_values = set()
                    for item in ordered_records:
                        if item.order_value in seen_values:
                            raise ConfigError(
                                f"{config.label} 中 fid={fid} 的排序列 {config.order.column} 存在重复值: {item.order_value}"
                            )
                        seen_values.add(item.order_value)

            merged_translation = "".join(item.compare_value for item in ordered_records)
            aggregated[fid] = AggregatedRecord(
                fid=fid,
                row_count=len(ordered_records),
                merged_translation=merged_translation,
            )
        return aggregated

    def _compare(
        self,
        left_records: Dict[str, AggregatedRecord],
        right_records: Dict[str, AggregatedRecord],
    ) -> Tuple[List[str], List[str], List[str]]:
        left_keys = set(left_records)
        right_keys = set(right_records)
        only_left = sorted(left_keys - right_keys)
        only_right = sorted(right_keys - left_keys)
        mismatches: List[str] = []
        for fid in sorted(left_keys & right_keys):
            if left_records[fid].merged_translation != right_records[fid].merged_translation:
                mismatches.append(fid)
        return only_left, only_right, mismatches

    def _write_report(
        self,
        left_records: Dict[str, AggregatedRecord],
        right_records: Dict[str, AggregatedRecord],
        only_left: List[str],
        only_right: List[str],
        mismatches: List[str],
    ) -> None:
        self.report_path.parent.mkdir(parents=True, exist_ok=True)
        with self.report_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "status",
                    "fid",
                    "left_row_count",
                    "right_row_count",
                    "left_translation",
                    "right_translation",
                ]
            )

            for fid in only_left:
                left = left_records[fid]
                writer.writerow(["only_left", fid, left.row_count, "", left.merged_translation, ""])
            for fid in only_right:
                right = right_records[fid]
                writer.writerow(["only_right", fid, "", right.row_count, "", right.merged_translation])
            for fid in mismatches:
                left = left_records[fid]
                right = right_records[fid]
                writer.writerow(
                    [
                        "translation_mismatch",
                        fid,
                        left.row_count,
                        right.row_count,
                        left.merged_translation,
                        right.merged_translation,
                    ]
                )

    def _write_summary(
        self,
        left_records: Dict[str, AggregatedRecord],
        right_records: Dict[str, AggregatedRecord],
        only_left: List[str],
        only_right: List[str],
        mismatches: List[str],
    ) -> None:
        total_matched = len(left_records) - len(only_left) - len(mismatches)
        result_text = "完全匹配" if not only_left and not only_right and not mismatches else "存在差异"
        self.summary_path.parent.mkdir(parents=True, exist_ok=True)
        lines = [
            f"result={result_text}",
            f"left_unique_fid={len(left_records)}",
            f"right_unique_fid={len(right_records)}",
            f"matched_fid={total_matched}",
            f"only_left_fid={len(only_left)}",
            f"only_right_fid={len(only_right)}",
            f"translation_mismatch_fid={len(mismatches)}",
            f"report_path={self.report_path}",
        ]
        self.summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按 fid 聚合 translation 并比对两个 xlsx")
    parser.add_argument("--config", required=True, help="YAML 配置文件路径")
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    config_path = Path(args.config).expanduser().resolve()
    left_cfg, right_cfg, report_path, summary_path = load_config(config_path)
    comparer = XlsxTranslationComparer(left_cfg, right_cfg, report_path, summary_path)
    exit_code = comparer.run()
    summary_text = summary_path.read_text(encoding="utf-8").strip()
    print(summary_text)
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())

    # python tools/xlsx_compare/compare_translation_by_fid.py --config tools/xlsx_compare/config.yaml