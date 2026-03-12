"""从 xlsx 原始数据推导 textId 错误→正确映射，生成 UPDATE SQL。

原理：
  与 generate_fix_sql_from_sqlite.py 相同，区别在于数据来源是 xlsx 文件。
  对每条分段，同时用"旧错误正则"和"新正确正则"各提取一次 textId，
  找出不一致的条目，建立 (fid, old_textId) → new_textId 映射，
  无需连接数据库，直接生成可执行的 UPDATE SQL。

旧正则（错误方式，复现当时导入时的行为）：
  格式1 {num}::::::[text]        → textId = {num}         ← 无需修复
  格式2 {num}:::{n}:::[text]     → textId = {num}         ← 错误（丢失 :::{n}）
  格式3 {num}:::{m-n}:::[text]   → textId = {num}         ← 错误（丢失 :::{m-n}）

新正则（正确方式）：
  格式1 {num}::::::[text]        → textId = {num}         ← 同上，无需修复
  格式2 {num}:::{n}:::[text]     → textId = {num}:::{n}   ← 正确
  格式3 {num}:::{m-n}:::[text]   → textId = {num}:::{m-n} ← 正确

xlsx 读取逻辑：
  同一 fid 可能分布在多行（通过 splitPart 列区分）。
  先按 fid 聚合、按 splitPart 排序后合并 sourceText，再按 splitDelimiter 拆分为段。
  只解析 sourceText 列，无需 translatedText（textId 在两列中相同）。

用法:
  python tools/fix_textid/generate_fix_sql_from_xlsx.py --config tools/fix_textid/generate_fix_sql_from_xlsx.yaml
"""

import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

_TOOLS_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_TOOLS_ROOT / "version_iteration_tool"))

from common import ConfigError, load_yaml_config, require_key, require_type  # noqa: E402


# ---------------------------------------------------------------------------
# 配置加载
# ---------------------------------------------------------------------------

def _find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "config" / "lotro.yaml").exists() or (candidate / ".git").exists():
            return candidate
    raise ConfigError("Cannot locate project root from config path")


def _resolve_base_dir(base_dir: str, config_path: Path) -> Path:
    if base_dir == "__PROJECT_ROOT__":
        return _find_project_root(config_path.parent)
    return Path(base_dir).expanduser().resolve()


def _require_column_letter(value: Any, path: str) -> str:
    text = require_type(value, str, path).strip().upper()
    if text == "":
        raise ConfigError(f"Column letter cannot be empty: {path}")
    try:
        column_index_from_string(text)
    except ValueError as exc:
        raise ConfigError(f"Invalid column letter: {path}={text}") from exc
    return text


def _parse_row_end(value: Any, max_row: int) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().lower() == "max":
        return max_row
    raise ConfigError("input.row_end must be integer or 'max'")


def _validate_config(config: Dict[str, Any], config_path: Path) -> Dict[str, Any]:
    base_dir_raw = require_type(require_key(config, "base_dir", ""), str, "base_dir")
    input_cfg = require_type(require_key(config, "input", ""), dict, "input")
    parsing_cfg = require_type(require_key(config, "parsing", ""), dict, "parsing")
    output_cfg = require_type(require_key(config, "output", ""), dict, "output")
    db_cfg = require_type(require_key(config, "db", ""), dict, "db")

    xlsx_path = require_type(require_key(input_cfg, "path", "input."), str, "input.path")
    sheet = require_type(require_key(input_cfg, "sheet", "input."), str, "input.sheet")
    row_start = require_type(require_key(input_cfg, "row_start", "input."), int, "input.row_start")
    row_end_raw = require_key(input_cfg, "row_end", "input.")

    columns_cfg = require_type(require_key(input_cfg, "columns", "input."), dict, "input.columns")
    fid_col = _require_column_letter(
        require_key(columns_cfg, "fid", "input.columns."), "input.columns.fid"
    )
    split_part_col = _require_column_letter(
        require_key(columns_cfg, "splitPart", "input.columns."), "input.columns.splitPart"
    )
    source_col = _require_column_letter(
        require_key(columns_cfg, "sourceText", "input.columns."), "input.columns.sourceText"
    )

    split_delimiter = require_type(
        require_key(parsing_cfg, "splitDelimiter", "parsing."), str, "parsing.splitDelimiter"
    )
    id_pattern = require_type(require_key(parsing_cfg, "idPattern", "parsing."), str, "parsing.idPattern")
    try:
        re.compile(id_pattern)
    except re.error as exc:
        raise ConfigError(f"parsing.idPattern 无效: {exc}") from exc

    csv_path = require_type(require_key(output_cfg, "csvPath", "output."), str, "output.csvPath")
    sql_path = require_type(require_key(output_cfg, "sqlPath", "output."), str, "output.sqlPath")
    overwrite = require_type(require_key(output_cfg, "overwrite", "output."), bool, "output.overwrite")

    db_table = require_type(require_key(db_cfg, "table", "db."), str, "db.table")

    if row_start <= 0:
        raise ConfigError("input.row_start must be > 0")
    if split_delimiter == "":
        raise ConfigError("parsing.splitDelimiter 不能为空")

    base_dir = _resolve_base_dir(base_dir_raw, config_path)

    return {
        "baseDir": base_dir,
        "xlsxPath": xlsx_path,
        "sheet": sheet,
        "rowStart": row_start,
        "rowEndRaw": row_end_raw,
        "fidCol": fid_col,
        "splitPartCol": split_part_col,
        "sourceCol": source_col,
        "splitDelimiter": split_delimiter,
        "idPattern": id_pattern,
        "csvPath": csv_path,
        "sqlPath": sql_path,
        "overwrite": overwrite,
        "dbTable": db_table,
    }


# ---------------------------------------------------------------------------
# 正则构建（与 sqlite 版本完全一致）
# ---------------------------------------------------------------------------

def _build_wrong_patterns(
    id_pattern: str,
) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    """构建旧的（错误的）正则——复现导入时的行为。
    格式2/3 的 textId 命名组只捕获数字前缀，不含 :::{n} 或 :::{m-n} 部分。
    """
    p1 = re.compile(rf"^(?P<textId>{id_pattern})::::::\[(?P<text>.*)\]$", re.DOTALL)
    p2 = re.compile(rf"^(?P<textId>{id_pattern}):::\d+:::\[(?P<text>.*)\]$", re.DOTALL)
    p3 = re.compile(rf"^(?P<textId>{id_pattern}):::\d+(?:-\d+)+:::\[(?P<text>.*)\]$", re.DOTALL)
    return p1, p2, p3


def _build_correct_patterns(
    id_pattern: str,
) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    """构建新的（正确的）正则。
    格式2/3 的 textId 命名组包含完整业务标识。
    """
    p1 = re.compile(rf"^(?P<textId>{id_pattern})::::::\[(?P<text>.*)\]$", re.DOTALL)
    p2 = re.compile(rf"^(?P<textId>{id_pattern}:::\d+):::\[(?P<text>.*)\]$", re.DOTALL)
    p3 = re.compile(
        rf"^(?P<textId>{id_pattern}:::\d+(?:-\d+)+):::\[(?P<text>.*)\]$", re.DOTALL
    )
    return p1, p2, p3


def _extract_textid(
    segment: str,
    patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
) -> Optional[str]:
    """从一条分段文本中提取 textId，返回 None 表示格式无法识别。"""
    for p in patterns:
        m = p.fullmatch(segment)
        if m:
            return m.group("textId")
    return None


# ---------------------------------------------------------------------------
# xlsx 读取：按 fid 聚合 + 按 splitPart 排序合并 sourceText
# ---------------------------------------------------------------------------

def _normalize_fid(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value)


def _normalize_split_part(value: Any, row_index: int) -> int:
    if value is None:
        raise ValueError(f"Row {row_index} splitPart is empty")
    if isinstance(value, bool):
        raise ValueError(f"Row {row_index} splitPart cannot be bool")
    if isinstance(value, int):
        part_value = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"Row {row_index} splitPart must be integer, got {value}")
        part_value = int(value)
    else:
        text = str(value).strip()
        if not re.fullmatch(r"-?\d+", text):
            raise ValueError(f"Row {row_index} splitPart must be integer, got {text}")
        part_value = int(text)
    if part_value < 0:
        raise ValueError(f"Row {row_index} splitPart must be >= 0, got {part_value}")
    return part_value


def _collect_merged_source(
    worksheet,
    row_start: int,
    row_end: int,
    fid_idx: int,
    split_part_idx: int,
    source_idx: int,
) -> List[Tuple[str, str]]:
    """从 xlsx 读取并按 fid 聚合，返回 [(fid, merged_sourceText), ...] 有序列表。
    同一 fid 的多行按 splitPart 排序后 sourceText 直接拼接（无分隔符，段协议本身携带边界）。
    """
    max_idx = max(fid_idx, split_part_idx, source_idx)
    fid_order: List[str] = []
    grouped: Dict[str, Dict[int, Tuple[int, str]]] = {}

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True),
        start=row_start,
    ):
        padded = list(row)
        if len(padded) <= max_idx:
            padded.extend([None] * (max_idx + 1 - len(padded)))

        fid = _normalize_fid(padded[fid_idx])
        source_raw = "" if padded[source_idx] is None else str(padded[source_idx])

        if fid.strip() == "" and source_raw.strip() == "":
            continue  # 空行跳过
        if fid.strip() == "":
            raise ValueError(f"Row {row_index} fid is empty")

        split_part = _normalize_split_part(padded[split_part_idx], row_index)

        if fid not in grouped:
            grouped[fid] = {}
            fid_order.append(fid)

        part_map = grouped[fid]
        if split_part in part_map:
            existing_row = part_map[split_part][0]
            raise ValueError(
                f"Row {row_index} duplicate splitPart for fid={fid}: {split_part} "
                f"(already seen at row {existing_row})"
            )
        part_map[split_part] = (row_index, source_raw)

    result: List[Tuple[str, str]] = []
    for fid in fid_order:
        part_map = grouped[fid]
        ordered_parts = sorted(part_map.keys())
        merged = "".join(part_map[p][1] for p in ordered_parts)
        result.append((fid, merged))
    return result


# ---------------------------------------------------------------------------
# 核心扫描：双正则提取 + 建立映射
# ---------------------------------------------------------------------------

def _scan_xlsx(
    config: Dict[str, Any],
    wrong_patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
    correct_patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
) -> Dict[Tuple[str, str], str]:
    """扫描 xlsx 数据，建立 (fid, old_textId) → new_textId 映射。
    只包含需要修复的条目（old_textId != new_textId）。
    """
    xlsx_path = Path(config["xlsxPath"]).expanduser()
    if not xlsx_path.is_absolute():
        xlsx_path = config["baseDir"] / xlsx_path
    xlsx_path = xlsx_path.resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx 文件不存在: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if config["sheet"] not in workbook.sheetnames:
        raise ConfigError(f"Worksheet not found: {config['sheet']}")
    worksheet = workbook[config["sheet"]]

    row_end = _parse_row_end(config["rowEndRaw"], worksheet.max_row)
    if row_end > worksheet.max_row:
        row_end = worksheet.max_row
    if row_end < config["rowStart"]:
        raise ConfigError("input.row_end must be >= input.row_start")

    fid_idx = column_index_from_string(config["fidCol"]) - 1
    split_part_idx = column_index_from_string(config["splitPartCol"]) - 1
    source_idx = column_index_from_string(config["sourceCol"]) - 1

    merged_rows = _collect_merged_source(
        worksheet=worksheet,
        row_start=config["rowStart"],
        row_end=row_end,
        fid_idx=fid_idx,
        split_part_idx=split_part_idx,
        source_idx=source_idx,
    )

    fix_map: Dict[Tuple[str, str], str] = {}
    total_fid_rows = len(merged_rows)
    total_segments = 0
    skipped_segments = 0

    for fid, merged_source in merged_rows:
        if merged_source.strip() == "":
            continue
        for raw_segment in merged_source.split(config["splitDelimiter"]):
            segment = raw_segment.strip()
            if segment == "":
                continue
            total_segments += 1

            old_tid = _extract_textid(segment, wrong_patterns)
            new_tid = _extract_textid(segment, correct_patterns)

            if old_tid is None or new_tid is None:
                skipped_segments += 1
                continue

            # 格式1：旧/新 textId 相同，无需修复
            if old_tid == new_tid:
                continue

            key = (fid, old_tid)
            if key in fix_map:
                if fix_map[key] != new_tid:
                    raise RuntimeError(
                        f"映射冲突: fid={fid}, old_textId={old_tid} "
                        f"映射到了不同的 new_textId: {fix_map[key]!r} vs {new_tid!r}"
                    )
            else:
                fix_map[key] = new_tid

    print(
        f"[INFO] xlsx 扫描完成: fid_rows={total_fid_rows}, "
        f"segments={total_segments}, skipped(格式未识别)={skipped_segments}"
    )
    print(f"[INFO] 需要修复的 (fid, textId) 条目: {len(fix_map)} 条")
    return fix_map


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------

def _sql_escape(value: str) -> str:
    """MySQL 字符串字面量转义（单引号包裹）。"""
    value = value.replace("\\", "\\\\")
    value = value.replace("'", "''")
    return "'" + value + "'"


_INSERT_CHUNK_SIZE = 1000  # 每个 INSERT 语句最多包含的行数


def _write_outputs(fix_map: Dict[Tuple[str, str], str], config: Dict[str, Any], source_label: str) -> None:
    """将修复映射写出为 CSV 和 UPDATE SQL。
    SQL 格式：临时表 + JOIN UPDATE，一次性批量执行，避免逐条 fsync 开销。
    """
    csv_path = Path(config["csvPath"]).expanduser().resolve()
    sql_path = Path(config["sqlPath"]).expanduser().resolve()
    db_table = config["dbTable"]
    now_str = datetime.now(timezone.utc).isoformat()

    sorted_entries = sorted(fix_map.items(), key=lambda x: (x[0][0], x[0][1]))

    with csv_path.open("w", encoding="utf-8") as f:
        f.write("fid,old_textId,new_textId\n")
        for (fid, old_tid), new_tid in sorted_entries:
            fid_esc = fid.replace('"', '""')
            old_esc = old_tid.replace('"', '""')
            new_esc = new_tid.replace('"', '""')
            f.write(f'"{fid_esc}","{old_esc}","{new_esc}"\n')
    print(f"[FILE] CSV: {csv_path}")

    with sql_path.open("w", encoding="utf-8") as f:
        f.write("-- Auto-generated by generate_fix_sql_from_xlsx.py\n")
        f.write(f"-- generated_at_utc: {now_str}\n")
        f.write(f"-- source: {source_label}\n")
        f.write(f"-- total_fix_entries: {len(fix_map)}\n\n")

        # 1. 创建临时表
        f.write("CREATE TEMPORARY TABLE `_textid_fix` (\n")
        f.write("  `fid` VARCHAR(64) NOT NULL,\n")
        f.write("  `old_textId` VARCHAR(200) NOT NULL,\n")
        f.write("  `new_textId` VARCHAR(200) NOT NULL,\n")
        f.write("  PRIMARY KEY (`fid`, `old_textId`)\n")
        f.write(") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_general_ci;\n\n")

        # 2. 分批 INSERT（每批 _INSERT_CHUNK_SIZE 行，避免单条 SQL 过长）
        for i in range(0, len(sorted_entries), _INSERT_CHUNK_SIZE):
            chunk = sorted_entries[i : i + _INSERT_CHUNK_SIZE]
            f.write("INSERT INTO `_textid_fix` (`fid`, `old_textId`, `new_textId`) VALUES\n")
            for j, ((fid, old_tid), new_tid) in enumerate(chunk):
                suffix = "," if j < len(chunk) - 1 else ";"
                f.write(
                    f"  ({_sql_escape(fid)}, {_sql_escape(old_tid)}, {_sql_escape(new_tid)}){suffix}\n"
                )
            f.write("\n")

        # 3. 一次 JOIN UPDATE 批量修复（走 (fid, old_textId) 联合主键，极快）
        f.write(f"UPDATE `{db_table}` t\n")
        f.write("  JOIN `_textid_fix` m ON t.`fid` = m.`fid` AND t.`textId` = m.`old_textId`\n")
        f.write("  SET t.`textId` = m.`new_textId`;\n\n")

        # 4. 清理临时表
        f.write("DROP TEMPORARY TABLE `_textid_fix`;\n")
    print(f"[FILE] SQL: {sql_path}")


# ---------------------------------------------------------------------------
# 入口
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="从 xlsx 原始数据推导 textId 修复映射，生成 UPDATE SQL（无需连接数据库）"
    )
    parser.add_argument("--config", required=True, help="配置文件路径")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser().resolve()
    config = _validate_config(load_yaml_config(config_path), config_path)

    csv_path = Path(config["csvPath"]).expanduser().resolve()
    sql_path = Path(config["sqlPath"]).expanduser().resolve()
    for out_path in (csv_path, sql_path):
        if out_path.exists() and not config["overwrite"]:
            raise RuntimeError(f"输出文件已存在且 overwrite=false: {out_path}")
        if not out_path.parent.exists():
            raise FileNotFoundError(f"输出目录不存在: {out_path.parent}")

    wrong_patterns = _build_wrong_patterns(config["idPattern"])
    correct_patterns = _build_correct_patterns(config["idPattern"])

    fix_map = _scan_xlsx(config, wrong_patterns, correct_patterns)
    _write_outputs(fix_map, config, source_label=config["xlsxPath"])

    print("[DONE] 修复映射生成完成")
    print(f"[STAT] total_fix_entries={len(fix_map)}")


if __name__ == "__main__":
    main()
    # python tools/fix_textid/generate_fix_sql_from_xlsx.py --config tools/fix_textid/generate_fix_sql_from_xlsx.yaml
