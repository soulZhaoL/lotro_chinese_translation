import argparse
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


_EXCEL_MAX_CELL_TEXT_LENGTH = 32767


class ConfigError(Exception):
    pass


def _require_key(obj: Dict[str, Any], key: str, path: str) -> Any:
    if key not in obj:
        raise ConfigError(f"Missing config key: {path}{key}")
    return obj[key]


def _require_type(value: Any, expected_type: type, path: str) -> Any:
    if not isinstance(value, expected_type):
        raise ConfigError(f"Invalid config type: {path}, expected {expected_type.__name__}")
    return value


def _require_positive_int(value: Any, path: str) -> int:
    number = _require_type(value, int, path)
    if number <= 0:
        raise ConfigError(f"{path} must be > 0")
    return number


def _require_non_negative_int(value: Any, path: str) -> int:
    number = _require_type(value, int, path)
    if number < 0:
        raise ConfigError(f"{path} must be >= 0")
    return number


def _find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "config" / "lotro.yaml").exists() or (candidate / ".git").exists():
            return candidate
    raise ConfigError("Cannot locate project root from config path")


def _resolve_base_dir(base_dir: str, config_path: Path) -> Path:
    if base_dir == "__PROJECT_ROOT__":
        return _find_project_root(config_path.parent)
    return Path(base_dir).expanduser().resolve()


def _require_column_letter(value: Any, path: str) -> str:
    text = _require_type(value, str, path).strip().upper()
    if text == "":
        raise ConfigError(f"Column letter cannot be empty: {path}")
    try:
        column_index_from_string(text)
    except ValueError as exc:
        raise ConfigError(f"Invalid column letter: {path}={text}") from exc
    return text


def _parse_row_end(value: Any, max_row: int) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().lower() == "max":
        return max_row
    raise ConfigError("input.row_end must be integer or 'max'")


def _load_config(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ConfigError("Config root must be a YAML object")
    return data


def _validate_config(data: Dict[str, Any], config_path: Path) -> Dict[str, Any]:
    base_dir_value = _require_type(_require_key(data, "base_dir", ""), str, "base_dir")
    input_cfg = _require_type(_require_key(data, "input", ""), dict, "input")
    parsing_cfg = _require_type(_require_key(data, "parsing", ""), dict, "parsing")
    repair_cfg = _require_type(_require_key(data, "repair", ""), dict, "repair")
    output_cfg = _require_type(_require_key(data, "output", ""), dict, "output")
    behavior_cfg = _require_type(_require_key(data, "behavior", ""), dict, "behavior")

    row_start = _require_positive_int(_require_key(input_cfg, "row_start", "input."), "input.row_start")
    sheet = _require_type(_require_key(input_cfg, "sheet", "input."), str, "input.sheet")
    column = _require_column_letter(_require_key(input_cfg, "column", "input."), "input.column")
    row_end_raw = _require_key(input_cfg, "row_end", "input.")
    input_path = _require_type(_require_key(input_cfg, "path", "input."), str, "input.path")

    split_delimiter = _require_type(
        _require_key(parsing_cfg, "splitDelimiter", "parsing."),
        str,
        "parsing.splitDelimiter",
    )
    if split_delimiter == "":
        raise ConfigError("parsing.splitDelimiter cannot be empty")
    id_pattern = _require_type(_require_key(parsing_cfg, "idPattern", "parsing."), str, "parsing.idPattern")
    if id_pattern == "":
        raise ConfigError("parsing.idPattern cannot be empty")

    allow_missing_opening = _require_type(
        _require_key(repair_cfg, "allowMissingOpeningBracket", "repair."),
        bool,
        "repair.allowMissingOpeningBracket",
    )
    allow_missing_closing = _require_type(
        _require_key(repair_cfg, "allowMissingClosingBracket", "repair."),
        bool,
        "repair.allowMissingClosingBracket",
    )
    allow_missing_colon = _require_type(
        _require_key(repair_cfg, "allowMissingColonInColon6Format", "repair."),
        bool,
        "repair.allowMissingColonInColon6Format",
    )

    output_path = _require_type(_require_key(output_cfg, "path", "output."), str, "output.path")
    overwrite = _require_type(_require_key(output_cfg, "overwrite", "output."), bool, "output.overwrite")

    skip_blank_cells = _require_type(
        _require_key(behavior_cfg, "skipBlankCells", "behavior."),
        bool,
        "behavior.skipBlankCells",
    )
    max_samples_per_kind = _require_non_negative_int(
        _require_key(behavior_cfg, "maxSamplesPerKind", "behavior."),
        "behavior.maxSamplesPerKind",
    )

    base_dir = _resolve_base_dir(base_dir_value, config_path)
    return {
        "base_dir": base_dir,
        "input": {
            "path": base_dir / input_path,
            "sheet": sheet,
            "row_start": row_start,
            "row_end_raw": row_end_raw,
            "column": column,
        },
        "parsing": {
            "splitDelimiter": split_delimiter,
            "idPattern": id_pattern,
        },
        "repair": {
            "allowMissingOpeningBracket": allow_missing_opening,
            "allowMissingClosingBracket": allow_missing_closing,
            "allowMissingColonInColon6Format": allow_missing_colon,
        },
        "output": {
            "path": base_dir / output_path,
            "overwrite": overwrite,
        },
        "behavior": {
            "skipBlankCells": skip_blank_cells,
            "maxSamplesPerKind": max_samples_per_kind,
        },
    }


def _build_valid_patterns(id_pattern: str) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    return (
        re.compile(rf"^(?P<textId>{id_pattern})::::::\[(?P<text>.*)\]$", re.DOTALL),
        re.compile(rf"^(?P<textId>{id_pattern}:::\d+):::\[(?P<text>.*)\]$", re.DOTALL),
        re.compile(rf"^(?P<textId>{id_pattern}:::\d+(?:-\d+)+):::\[(?P<text>.*)\]$", re.DOTALL),
    )


def _build_missing_open_patterns(id_pattern: str) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    return (
        re.compile(rf"^(?P<head>{id_pattern}::::::)(?P<text>(?!\[).*)\]$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::\d+:::)(?P<text>(?!\[).*)\]$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::\d+(?:-\d+)+:::)(?P<text>(?!\[).*)\]$", re.DOTALL),
    )


def _build_missing_close_patterns(id_pattern: str) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    return (
        re.compile(rf"^(?P<head>{id_pattern}::::::)\[(?P<text>.*)$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::\d+:::)\[(?P<text>.*)$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::\d+(?:-\d+)+:::)\[(?P<text>.*)$", re.DOTALL),
    )


def _build_missing_colon_patterns(id_pattern: str) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    return (
        re.compile(rf"^(?P<head>{id_pattern}:::::)(?!:)\[(?P<text>.*)\]$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::::)(?!:)(?P<text>(?!\[).*)\]$", re.DOTALL),
        re.compile(rf"^(?P<head>{id_pattern}:::::)(?!:)\[(?P<text>.*)$", re.DOTALL),
    )


def _first_fullmatch(patterns: Tuple[re.Pattern[str], ...], text: str) -> Optional[re.Match[str]]:
    for pattern in patterns:
        matched = pattern.fullmatch(text)
        if matched is not None:
            return matched
    return None


def _classify_segment(
    segment: str,
    valid_patterns: Tuple[re.Pattern[str], ...],
    missing_colon_patterns: Tuple[re.Pattern[str], ...],
    missing_open_patterns: Tuple[re.Pattern[str], ...],
    missing_close_patterns: Tuple[re.Pattern[str], ...],
) -> str:
    if _first_fullmatch(valid_patterns, segment) is not None:
        return "valid"
    if _first_fullmatch(missing_colon_patterns, segment) is not None:
        return "missing_colon"
    if _first_fullmatch(missing_open_patterns, segment) is not None:
        return "missing_open"
    if _first_fullmatch(missing_close_patterns, segment) is not None:
        return "missing_close"
    return "other_invalid"


def _repair_segment(
    segment: str,
    valid_patterns: Tuple[re.Pattern[str], ...],
    missing_colon_patterns: Tuple[re.Pattern[str], ...],
    missing_open_patterns: Tuple[re.Pattern[str], ...],
    missing_close_patterns: Tuple[re.Pattern[str], ...],
    allow_missing_colon: bool,
    allow_missing_opening: bool,
    allow_missing_closing: bool,
) -> Tuple[str, List[str]]:
    if _first_fullmatch(valid_patterns, segment) is not None:
        return segment, []

    if allow_missing_colon:
        matched = _first_fullmatch(missing_colon_patterns, segment)
        if matched is not None:
            repaired = f"{matched.group('head')}:[{matched.group('text')}]"
            if _first_fullmatch(valid_patterns, repaired) is None:
                raise ValueError(f"repair produced invalid segment: {repaired}")
            repair_kinds = ["missing_colon"]
            if "[" not in segment:
                repair_kinds.append("missing_open")
            if segment.endswith("]") is False:
                repair_kinds.append("missing_close")
            return repaired, repair_kinds

    if allow_missing_opening:
        matched = _first_fullmatch(missing_open_patterns, segment)
        if matched is not None:
            repaired = f"{matched.group('head')}[{matched.group('text')}]"
            if _first_fullmatch(valid_patterns, repaired) is None:
                raise ValueError(f"repair produced invalid segment: {repaired}")
            return repaired, ["missing_open"]

    if allow_missing_closing:
        matched = _first_fullmatch(missing_close_patterns, segment)
        if matched is not None:
            repaired = f"{matched.group('head')}[{matched.group('text')}]"
            if _first_fullmatch(valid_patterns, repaired) is None:
                raise ValueError(f"repair produced invalid segment: {repaired}")
            return repaired, ["missing_close"]

    return segment, []


def _split_piece_whitespace(piece: str) -> Tuple[str, str, str]:
    leading_length = len(piece) - len(piece.lstrip())
    trailing_length = len(piece) - len(piece.rstrip())
    leading = piece[:leading_length]
    if trailing_length == 0:
        trailing = ""
        core_end = len(piece)
    else:
        trailing = piece[-trailing_length:]
        core_end = len(piece) - trailing_length
    core = piece[leading_length:core_end]
    return leading, core, trailing


def _summarize_segment(text: str, max_len: int) -> str:
    snippet = text.replace("\n", "\\n")
    if len(snippet) <= max_len:
        return snippet
    return snippet[: max_len - 3] + "..."


def _repair_cell_text(
    raw_text: str,
    split_delimiter: str,
    valid_patterns: Tuple[re.Pattern[str], ...],
    missing_colon_patterns: Tuple[re.Pattern[str], ...],
    missing_open_patterns: Tuple[re.Pattern[str], ...],
    missing_close_patterns: Tuple[re.Pattern[str], ...],
    allow_missing_colon: bool,
    allow_missing_opening: bool,
    allow_missing_closing: bool,
) -> Dict[str, Any]:
    raw_pieces = raw_text.split(split_delimiter)
    repaired_pieces: List[str] = []
    repaired_missing_colon = 0
    repaired_missing_open = 0
    repaired_missing_close = 0
    invalid_segments: List[Tuple[int, str, str]] = []

    for segment_index, raw_piece in enumerate(raw_pieces, start=1):
        leading, core, trailing = _split_piece_whitespace(raw_piece)
        stripped_core = core.strip()
        if stripped_core == "":
            repaired_pieces.append(raw_piece)
            invalid_segments.append((segment_index, "empty_segment", ""))
            continue

        repaired_core, repair_kinds = _repair_segment(
            stripped_core,
            valid_patterns,
            missing_colon_patterns,
            missing_open_patterns,
            missing_close_patterns,
            allow_missing_colon,
            allow_missing_opening,
            allow_missing_closing,
        )

        repaired_missing_colon += repair_kinds.count("missing_colon")
        repaired_missing_open += repair_kinds.count("missing_open")
        repaired_missing_close += repair_kinds.count("missing_close")

        if len(repair_kinds) == 0:
            final_piece = raw_piece
            final_segment = stripped_core
        else:
            final_piece = f"{leading}{repaired_core}{trailing}"
            final_segment = repaired_core
        repaired_pieces.append(final_piece)

        final_kind = _classify_segment(
            final_segment,
            valid_patterns,
            missing_colon_patterns,
            missing_open_patterns,
            missing_close_patterns,
        )
        if final_kind != "valid":
            invalid_segments.append((segment_index, final_kind, final_segment))

    return {
        "text": split_delimiter.join(repaired_pieces),
        "repairedMissingColon": repaired_missing_colon,
        "repairedMissingOpen": repaired_missing_open,
        "repairedMissingClose": repaired_missing_close,
        "invalidSegments": invalid_segments,
    }


def _collect_invalid_segments(
    raw_text: str,
    split_delimiter: str,
    valid_patterns: Tuple[re.Pattern[str], ...],
    missing_colon_patterns: Tuple[re.Pattern[str], ...],
    missing_open_patterns: Tuple[re.Pattern[str], ...],
    missing_close_patterns: Tuple[re.Pattern[str], ...],
) -> List[Tuple[int, str, str]]:
    invalid_segments: List[Tuple[int, str, str]] = []
    for segment_index, raw_piece in enumerate(raw_text.split(split_delimiter), start=1):
        _, core, _ = _split_piece_whitespace(raw_piece)
        stripped_core = core.strip()
        if stripped_core == "":
            invalid_segments.append((segment_index, "empty_segment", ""))
            continue
        kind = _classify_segment(
            stripped_core,
            valid_patterns,
            missing_colon_patterns,
            missing_open_patterns,
            missing_close_patterns,
        )
        if kind != "valid":
            invalid_segments.append((segment_index, kind, stripped_core))
    return invalid_segments


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def run_from_config(config_path: Path) -> Dict[str, Any]:
    config_data = _load_config(config_path)
    config = _validate_config(config_data, config_path)

    input_path = config["input"]["path"]
    output_path = config["output"]["path"]
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if output_path.exists() and config["output"]["overwrite"] is False:
        raise FileExistsError(f"Output file already exists: {output_path}")

    workbook = load_workbook(input_path)
    sheet_name = config["input"]["sheet"]
    if sheet_name not in workbook.sheetnames:
        raise ConfigError(f"Sheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]

    row_end = _parse_row_end(config["input"]["row_end_raw"], worksheet.max_row)
    row_start = config["input"]["row_start"]
    if row_end < row_start:
        raise ConfigError("input.row_end must be >= input.row_start")

    column_index = column_index_from_string(config["input"]["column"])
    split_delimiter = config["parsing"]["splitDelimiter"]
    id_pattern = config["parsing"]["idPattern"]
    allow_missing_opening = config["repair"]["allowMissingOpeningBracket"]
    allow_missing_closing = config["repair"]["allowMissingClosingBracket"]
    allow_missing_colon = config["repair"]["allowMissingColonInColon6Format"]
    max_samples_per_kind = config["behavior"]["maxSamplesPerKind"]
    skip_blank_cells = config["behavior"]["skipBlankCells"]

    valid_patterns = _build_valid_patterns(id_pattern)
    missing_colon_patterns = _build_missing_colon_patterns(id_pattern)
    missing_open_patterns = _build_missing_open_patterns(id_pattern)
    missing_close_patterns = _build_missing_close_patterns(id_pattern)

    invalid_samples: Dict[str, List[Tuple[int, int, str]]] = {
        "empty_segment": [],
        "missing_colon": [],
        "missing_open": [],
        "missing_close": [],
        "other_invalid": [],
    }
    remaining_invalid_counts: Dict[str, int] = {
        "empty_segment": 0,
        "missing_colon": 0,
        "missing_open": 0,
        "missing_close": 0,
        "other_invalid": 0,
    }

    changed_cells = 0
    scanned_cells = 0
    repaired_missing_colon = 0
    repaired_missing_open = 0
    repaired_missing_close = 0
    overflow_blocked_cells = 0
    overflow_blocked_segments = 0
    overflow_samples: List[Tuple[int, int, int, str]] = []

    for row_index in range(row_start, row_end + 1):
        cell = worksheet.cell(row=row_index, column=column_index)
        raw_text = _normalize_cell(cell.value)
        if skip_blank_cells and raw_text.strip() == "":
            continue

        scanned_cells += 1
        result = _repair_cell_text(
            raw_text,
            split_delimiter,
            valid_patterns,
            missing_colon_patterns,
            missing_open_patterns,
            missing_close_patterns,
            allow_missing_colon,
            allow_missing_opening,
            allow_missing_closing,
        )

        repaired_text = result["text"]
        final_invalid_segments = result["invalidSegments"]
        if repaired_text != raw_text and len(repaired_text) > _EXCEL_MAX_CELL_TEXT_LENGTH:
            overflow_blocked_cells += 1
            overflow_blocked_segments += (
                result["repairedMissingColon"]
                + result["repairedMissingOpen"]
                + result["repairedMissingClose"]
            )
            if len(overflow_samples) < max_samples_per_kind:
                overflow_samples.append(
                    (
                        row_index,
                        len(raw_text),
                        len(repaired_text),
                        _summarize_segment(repaired_text[-120:], 120),
                    )
                )
            final_invalid_segments = _collect_invalid_segments(
                raw_text,
                split_delimiter,
                valid_patterns,
                missing_colon_patterns,
                missing_open_patterns,
                missing_close_patterns,
            )
        else:
            if repaired_text != raw_text:
                cell.value = repaired_text
                changed_cells += 1
            repaired_missing_colon += result["repairedMissingColon"]
            repaired_missing_open += result["repairedMissingOpen"]
            repaired_missing_close += result["repairedMissingClose"]

        for segment_index, invalid_kind, segment in final_invalid_segments:
            remaining_invalid_counts[invalid_kind] += 1
            samples = invalid_samples[invalid_kind]
            if len(samples) < max_samples_per_kind:
                samples.append((row_index, segment_index, _summarize_segment(segment, 120)))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "inputPath": str(input_path),
        "outputPath": str(output_path),
        "sheet": sheet_name,
        "column": config["input"]["column"],
        "rowStart": row_start,
        "rowEnd": row_end,
        "scannedCells": scanned_cells,
        "changedCells": changed_cells,
        "repairedMissingColon": repaired_missing_colon,
        "repairedMissingOpen": repaired_missing_open,
        "repairedMissingClose": repaired_missing_close,
        "overflowBlockedCells": overflow_blocked_cells,
        "overflowBlockedSegments": overflow_blocked_segments,
        "overflowBlockedSamples": overflow_samples,
        "remainingInvalidCounts": remaining_invalid_counts,
        "remainingInvalidSamples": invalid_samples,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="修复 xlsx 分段协议中明显缺失的外层方括号")
    parser.add_argument("--config", required=True, help="YAML 配置文件路径")
    return parser.parse_args()


def _print_summary(summary: Dict[str, Any]) -> None:
    repaired_total = (
        summary["repairedMissingColon"]
        + summary["repairedMissingOpen"]
        + summary["repairedMissingClose"]
    )
    print(f"input_path={summary['inputPath']}")
    print(f"output_path={summary['outputPath']}")
    print(f"sheet={summary['sheet']} column={summary['column']} rows={summary['rowStart']}-{summary['rowEnd']}")
    print(f"scanned_cells={summary['scannedCells']}")
    print(f"changed_cells={summary['changedCells']}")
    print(f"repaired_total={repaired_total}")
    print(f"repaired_missing_colon={summary['repairedMissingColon']}")
    print(f"repaired_missing_open={summary['repairedMissingOpen']}")
    print(f"repaired_missing_close={summary['repairedMissingClose']}")
    print(f"overflow_blocked_cells={summary['overflowBlockedCells']}")
    print(f"overflow_blocked_segments={summary['overflowBlockedSegments']}")
    if len(summary["overflowBlockedSamples"]) > 0:
        print("overflow_blocked_samples:")
        for row_index, original_len, repaired_len, snippet in summary["overflowBlockedSamples"]:
            print(
                f"  row={row_index} original_len={original_len} "
                f"repaired_len={repaired_len} tail={snippet}"
            )
    print("remaining_invalid_counts:")
    for kind, count in summary["remainingInvalidCounts"].items():
        print(f"  {kind}={count}")
    print("remaining_invalid_samples:")
    for kind, samples in summary["remainingInvalidSamples"].items():
        if len(samples) == 0:
            continue
        print(f"  {kind}:")
        for row_index, segment_index, snippet in samples:
            print(f"    row={row_index} segment={segment_index} snippet={snippet}")


def main() -> None:
    args = _parse_args()
    summary = run_from_config(Path(args.config).expanduser().resolve())
    _print_summary(summary)


if __name__ == "__main__":
    main()
