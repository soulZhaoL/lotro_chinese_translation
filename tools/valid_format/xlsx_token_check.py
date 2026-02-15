# xlsx 标记计数脚本，用于检查文本中的特定标记数量。
# 支持按列/行范围读取并对比两列数量，输出差异并生成 xlsx 报告。

import argparse
import os
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _count_token(text: str, token: str) -> int:
    if not token:
        return 0
    return text.count(token)


def _summarize_segment(text: str, max_len: int) -> str:
    if text == "":
        return "<EMPTY>"
    snippet = text.replace("\n", "\\n")
    if len(snippet) <= max_len:
        return snippet
    return snippet[: max_len - 3] + "..."


def _split_segments(text: str) -> List[str]:
    if not text:
        return []
    return text.split("|||")


def _extract_segment_info(segment: str) -> Tuple[Optional[str], str]:
    segment = segment.strip()
    if not segment:
        return None, segment
    # 4-10 位数字 + :::::: + [text]
    match = re.match(r"^(\d{4,10})::::::\[(.*)\]$", segment, re.DOTALL)
    if match:
        return match.group(1), match.group(2)
    # 4-10 位数字 + ::: + 数字/范围 + ::: + [text]
    match = re.match(r"^(\d{4,10}):::[0-9-]+:::\[(.*)\]$", segment, re.DOTALL)
    if match:
        return match.group(1), match.group(2)
    return None, segment


def _segment_token_diffs(
    left_text: str, right_text: str, token: str, max_items: int
) -> List[Tuple[str, int, int, str, str]]:
    left_segments = _split_segments(left_text)
    right_segments = _split_segments(right_text)

    left_map: List[Tuple[str, str]] = []
    right_map: List[Tuple[str, str]] = []
    left_invalid: List[Tuple[int, str]] = []
    right_invalid: List[Tuple[int, str]] = []

    for idx, seg in enumerate(left_segments, start=1):
        seg_id, seg_text = _extract_segment_info(seg)
        if seg_id is None:
            if seg_text.strip():
                left_invalid.append((idx, seg_text))
            continue
        left_map.append((seg_id, seg_text))

    for idx, seg in enumerate(right_segments, start=1):
        seg_id, seg_text = _extract_segment_info(seg)
        if seg_id is None:
            if seg_text.strip():
                right_invalid.append((idx, seg_text))
            continue
        right_map.append((seg_id, seg_text))

    # build id -> list of texts, preserving order
    def _group(items: List[Tuple[str, str]]) -> dict:
        grouped: dict[str, List[str]] = {}
        for key, text in items:
            grouped.setdefault(key, []).append(text)
        return grouped

    left_grouped = _group(left_map)
    right_grouped = _group(right_map)

    diffs: List[Tuple[str, int, int, str, str]] = []

    for idx, seg_text in right_invalid:
        diffs.append(
            (
                f"译文格式错误#{idx}",
                0,
                _count_token(seg_text, token),
                "",
                seg_text,
            )
        )
        if len(diffs) >= max_items:
            return diffs

    for idx, seg_text in left_invalid:
        diffs.append(
            (
                f"原文格式错误#{idx}",
                _count_token(seg_text, token),
                0,
                seg_text,
                "",
            )
        )
        if len(diffs) >= max_items:
            return diffs

    # iterate in left order to keep stable output
    seen_keys: set[str] = set()
    for key, _ in left_map:
        if key in seen_keys:
            continue
        seen_keys.add(key)
        left_list = left_grouped.get(key, [])
        right_list = right_grouped.get(key, [])
        max_len = max(len(left_list), len(right_list))
        for idx in range(max_len):
            left_seg = left_list[idx] if idx < len(left_list) else ""
            right_seg = right_list[idx] if idx < len(right_list) else ""
            left_count = _count_token(left_seg, token)
            right_count = _count_token(right_seg, token)
            if left_count != right_count:
                label = key
                diffs.append((label, left_count, right_count, left_seg, right_seg))
                if len(diffs) >= max_items:
                    return diffs

    # include keys only in right (extra tokens)
    for key, _ in right_map:
        if key in seen_keys:
            continue
        seen_keys.add(key)
        left_list = left_grouped.get(key, [])
        right_list = right_grouped.get(key, [])
        max_len = max(len(left_list), len(right_list))
        for idx in range(max_len):
            left_seg = left_list[idx] if idx < len(left_list) else ""
            right_seg = right_list[idx] if idx < len(right_list) else ""
            left_count = _count_token(left_seg, token)
            right_count = _count_token(right_seg, token)
            if left_count != right_count:
                label = key
                diffs.append((label, left_count, right_count, left_seg, right_seg))
                if len(diffs) >= max_items:
                    return diffs
    return diffs


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="xlsx 标记计数脚本")
    parser.add_argument(
        "--path", default="work_text/text_work.xlsx", help="xlsx 文件路径"
    )
    parser.add_argument("--sheet", default="Sheet1", help="工作表名称，默认使用活动表")
    parser.add_argument("--column", default="C", help="待解析列（如 C）")
    parser.add_argument(
        "--row-start", type=int, default=2, help="起始行号（从 2 开始,不考虑标题）"
    )
    parser.add_argument("--row-end", type=int, default=100, help="结束行号（包含）")
    parser.add_argument("--compare-column", default="D", help="对比列（如 D）")
    parser.add_argument(
        "--only-mismatch", action="store_true", help="仅输出数量不一致的行"
    )
    parser.add_argument(
        "--show-mismatch", action="store_true", help="输出不一致行的附近文本片段"
    )
    parser.add_argument(
        "--token",
        default="<--DO_NOT_TOUCH!-->",
        help="需要统计的标记文本",
    )
    parser.add_argument(
        "--snippet-len", type=int, default=120, help="片段输出最大长度"
    )
    parser.add_argument(
        "--max-items", type=int, default=20, help="最多输出的差异片段数"
    )
    parser.add_argument(
        "--output-mismatch-xlsx",
        default=None,
        help="输出不一致行的 xlsx 文件路径",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=max(1, min(32, (os.cpu_count() or 4) + 4)),
        help="并发线程数（默认 CPU 核心数 + 4，上限 32）",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=10000,
        help="进度输出间隔行数（0 表示不输出）",
    )
    return parser.parse_args()


def _resolve_sheet(workbook, sheet_name: Optional[str]):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        return workbook[sheet_name]
    return workbook.active


def _progress_print(current: int, total: int) -> None:
    percent = 0.0
    if total > 0:
        percent = current * 100.0 / total
    print(f"Progress: {current}/{total} ({percent:.2f}%)", file=sys.stderr)


def _format_mismatch_detail(
    diffs: List[Tuple[str, int, int, str, str]],
    snippet_len: int,
) -> str:
    lines: List[str] = []
    for seg_label, left_count, right_count, left_seg, right_seg in diffs:
        if seg_label.startswith("译文格式错误#"):
            lines.append(f"{seg_label}: {_summarize_segment(right_seg, snippet_len)}")
            continue
        if seg_label.startswith("原文格式错误#"):
            lines.append(f"{seg_label}: {_summarize_segment(left_seg, snippet_len)}")
            continue
        status = "缺失" if right_count < left_count else "多出"
        lines.append(
            f"{seg_label} {status} L{left_count}->R{right_count}: "
            f"{_summarize_segment(left_seg, snippet_len)}"
        )
        lines.append(f"{seg_label} R: {_summarize_segment(right_seg, snippet_len)}")
    return "\n".join(lines)


def _write_mismatch_xlsx(
    path: Path,
    rows: List[Tuple[int, str, str, int, int, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "mismatch"
    worksheet.append(
        [
            "row",
            "fid",
            "part",
            "left_count",
            "right_count",
            "mismatch_detail",
        ]
    )
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)


def _analyze_compare_row(
    row_index: int,
    a_value: str,
    b_value: str,
    left_text: str,
    right_text: str,
    token: str,
    max_items: int,
) -> Tuple[int, str, str, int, int, List[Tuple[str, int, int, str, str]]]:
    left_count = _count_token(left_text, token)
    right_count = _count_token(right_text, token)
    diffs = _segment_token_diffs(left_text, right_text, token, max_items)
    return (
        row_index,
        a_value,
        b_value,
        left_count,
        right_count,
        diffs,
    )


def main() -> None:
    args = _parse_args()
    path = Path(args.path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")
    if args.row_start <= 0:
        raise ValueError("row-start 必须大于 0")
    if args.row_end < args.row_start:
        raise ValueError("row-end 必须大于等于 row-start")
    if not args.compare_column:
        raise ValueError("compare-column 不能为空")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = _resolve_sheet(workbook, args.sheet)

    row_end = min(args.row_end, worksheet.max_row or args.row_end)
    if row_end < args.row_start:
        raise ValueError("row-start 超出工作表范围")
    total_rows = row_end - args.row_start + 1

    column_index = column_index_from_string(args.column)
    compare_index = column_index_from_string(args.compare_column)
    max_col = max(column_index, compare_index, 2)

    rows_iter = worksheet.iter_rows(
        min_row=args.row_start,
        max_row=row_end,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )

    print("row\tleft\tright\tmatch")
    mismatch_rows: List[Tuple[int, str, str, int, int, str]] = []

    def _compare_task_iter():
        for offset, row_values in enumerate(rows_iter, start=0):
            row_index = args.row_start + offset
            left_value = row_values[column_index - 1]
            right_value = row_values[compare_index - 1]
            left_text = _normalize_cell(left_value)
            right_text = _normalize_cell(right_value)
            a_value = _normalize_cell(row_values[0])
            b_value = _normalize_cell(row_values[1] if len(row_values) > 1 else "")
            yield row_index, a_value, b_value, left_text, right_text

    completed = 0
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        for (
            row_index,
            a_value,
            b_value,
            left_count,
            right_count,
            diffs,
        ) in executor.map(
            lambda args_tuple: _analyze_compare_row(*args_tuple),
            (
                (
                    row_index,
                    a_value,
                    b_value,
                    left_text,
                    right_text,
                    args.token,
                    args.max_items,
                )
                for row_index, a_value, b_value, left_text, right_text in _compare_task_iter()
            ),
            chunksize=200,
        ):
            match = left_count == right_count
            if not (args.only_mismatch and match):
                flag = "OK" if match else "DIFF"
                print(f"{row_index}\t{left_count}\t{right_count}\t{flag}")
            if not match and args.output_mismatch_xlsx:
                detail = _format_mismatch_detail(diffs, args.snippet_len)
                mismatch_rows.append(
                    (
                        row_index,
                        a_value,
                        b_value,
                        left_count,
                        right_count,
                        detail,
                    )
                )
            if not match and args.show_mismatch:
                print(f"--- row {row_index} ---")
                for seg_label, left_c, right_c, left_seg, right_seg in diffs[: args.max_items]:
                    print(
                        f"{seg_label} L{left_c}->R{right_c}: "
                        f"{_summarize_segment(left_seg, args.snippet_len)}"
                    )
                    print(
                        f"{seg_label} R: {_summarize_segment(right_seg, args.snippet_len)}"
                    )
            completed += 1
            if args.progress_every and completed % args.progress_every == 0:
                _progress_print(completed, total_rows)

    if args.output_mismatch_xlsx:
        output_path = Path(args.output_mismatch_xlsx)
        _write_mismatch_xlsx(output_path, mismatch_rows)


if __name__ == "__main__":
    main()
# python ./tools/xlsx_token_check.py --row-start 2 --row-end 1000000 --compare-column D --output-mismatch-xlsx mismatch.xlsx 