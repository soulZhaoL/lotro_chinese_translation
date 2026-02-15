import argparse
import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class ConfigError(Exception):
    pass


class RowParseError(Exception):
    pass


def _require_key(obj: Dict[str, Any], key: str, path: str) -> Any:
    if key not in obj:
        raise ConfigError(f"Missing config key: {path}{key}")
    return obj[key]


def _require_type(value: Any, expected_type: type, path: str) -> Any:
    if not isinstance(value, expected_type):
        raise ConfigError(f"Invalid config type: {path}, expected {expected_type.__name__}")
    return value


def _require_identifier(value: str, path: str) -> str:
    if value == "":
        raise ConfigError(f"Identifier cannot be empty: {path}")
    first = value[0]
    if not (first.isalpha() or first == "_"):
        raise ConfigError(f"Invalid identifier: {path}={value}")
    for char in value[1:]:
        if not (char.isalnum() or char == "_"):
            raise ConfigError(f"Invalid identifier: {path}={value}")
    return value


def _require_column_letter(value: Any, path: str) -> str:
    text = _require_type(value, str, path).strip().upper()
    if text == "":
        raise ConfigError(f"Column letter cannot be empty: {path}")
    try:
        column_index_from_string(text)
    except ValueError as exc:
        raise ConfigError(f"Invalid column letter: {path}={text}") from exc
    return text


def _parse_row_end(value: Any, max_row: int) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().lower() == "max":
        return max_row
    raise ConfigError("input.row_end must be integer or 'max'")


def _parse_row_range(value: str) -> Tuple[int, int]:
    parts = value.split("-", 1)
    if len(parts) != 2:
        raise ConfigError("row-range must be m-n")
    start_text, end_text = parts[0].strip(), parts[1].strip()
    if not start_text.isdigit() or not end_text.isdigit():
        raise ConfigError("row-range values must be digits")
    row_start = int(start_text)
    row_end = int(end_text)
    if row_start <= 0:
        raise ConfigError("row-range start must be > 0")
    if row_end < row_start:
        raise ConfigError("row-range end must be >= start")
    return row_start, row_end


def _find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "config" / "lotro.yaml").exists() or (candidate / ".git").exists():
            return candidate
    raise ConfigError("Cannot locate project root from config path")


def _resolve_base_dir(base_dir: str, config_path: Path) -> Path:
    if base_dir == "__PROJECT_ROOT__":
        return _find_project_root(config_path.parent)
    return Path(base_dir).expanduser().resolve()


def _normalize_text_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_fid(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _normalize_split_part(value: Any, row_index: int) -> int:
    if value is None:
        raise RowParseError(f"Row {row_index} splitPart is empty")
    if isinstance(value, bool):
        raise RowParseError(f"Row {row_index} splitPart cannot be bool")
    if isinstance(value, int):
        part_value = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise RowParseError(f"Row {row_index} splitPart must be integer, got {value}")
        part_value = int(value)
    else:
        text = str(value).strip()
        if text == "":
            raise RowParseError(f"Row {row_index} splitPart is empty")
        if not re.fullmatch(r"-?\d+", text):
            raise RowParseError(f"Row {row_index} splitPart must be integer, got {text}")
        part_value = int(text)
    if part_value < 0:
        raise RowParseError(f"Row {row_index} splitPart must be >= 0, got {part_value}")
    return part_value


def _sql_identifier(name: str) -> str:
    if name == "":
        raise ConfigError("SQL identifier cannot be empty")
    return '"' + name.replace('"', '""') + '"'


def _sql_qualified_identifier(name: str) -> str:
    parts = [item.strip() for item in name.split(".")]
    if not parts or any(not item for item in parts):
        raise ConfigError(f"Invalid SQL identifier: {name}")
    return ".".join(_sql_identifier(item) for item in parts)


def _sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("'", "''")
    return "'" + text + "'"


def _load_config(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {path}")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ConfigError("Config root must be a YAML object")
    return data


def _validate_policy(value: str, path: str, choices: Tuple[str, ...]) -> str:
    if value not in choices:
        joined = "/".join(choices)
        raise ConfigError(f"{path} must be one of {joined}")
    return value


def _validate_output_columns(columns_cfg: Dict[str, Any]) -> Dict[str, str]:
    required_keys = (
        "fid",
        "part",
        "textId",
        "sourceText",
        "sourceTextHash",
        "translatedText",
        "status",
        "isClaimed",
    )
    normalized: Dict[str, str] = {}
    for key in required_keys:
        value = _require_type(_require_key(columns_cfg, key, "output.columns."), str, f"output.columns.{key}")
        normalized[key] = _require_identifier(value, f"output.columns.{key}")
    return normalized


def _validate_config(data: Dict[str, Any]) -> Dict[str, Any]:
    base_dir = _require_type(_require_key(data, "base_dir", ""), str, "base_dir")
    input_cfg = _require_type(_require_key(data, "input", ""), dict, "input")
    parsing_cfg = _require_type(_require_key(data, "parsing", ""), dict, "parsing")
    output_cfg = _require_type(_require_key(data, "output", ""), dict, "output")
    fixed_values = _require_type(_require_key(data, "fixedValues", ""), dict, "fixedValues")
    behavior_cfg = _require_type(_require_key(data, "behavior", ""), dict, "behavior")

    input_path = _require_type(_require_key(input_cfg, "path", "input."), str, "input.path")
    sheet = _require_type(_require_key(input_cfg, "sheet", "input."), str, "input.sheet")
    row_start = _require_type(_require_key(input_cfg, "row_start", "input."), int, "input.row_start")
    row_end = _require_key(input_cfg, "row_end", "input.")
    columns_cfg = _require_type(_require_key(input_cfg, "columns", "input."), dict, "input.columns")
    fid_col = _require_column_letter(_require_key(columns_cfg, "fid", "input.columns."), "input.columns.fid")
    split_part_col = _require_column_letter(
        _require_key(columns_cfg, "splitPart", "input.columns."),
        "input.columns.splitPart",
    )
    source_col = _require_column_letter(
        _require_key(columns_cfg, "sourceText", "input.columns."),
        "input.columns.sourceText",
    )
    translated_col = _require_column_letter(
        _require_key(columns_cfg, "translatedText", "input.columns."),
        "input.columns.translatedText",
    )

    split_delimiter = _require_type(
        _require_key(parsing_cfg, "splitDelimiter", "parsing."),
        str,
        "parsing.splitDelimiter",
    )
    id_pattern = _require_type(_require_key(parsing_cfg, "idPattern", "parsing."), str, "parsing.idPattern")

    output_path = _require_type(_require_key(output_cfg, "path", "output."), str, "output.path")
    table = _require_type(_require_key(output_cfg, "table", "output."), str, "output.table")
    chunk_size = _require_type(_require_key(output_cfg, "chunkSize", "output."), int, "output.chunkSize")
    overwrite = _require_type(_require_key(output_cfg, "overwrite", "output."), bool, "output.overwrite")
    output_columns = _validate_output_columns(
        _require_type(_require_key(output_cfg, "columns", "output."), dict, "output.columns")
    )

    status_value = _require_type(_require_key(fixed_values, "status", "fixedValues."), int, "fixedValues.status")
    is_claimed_value = _require_type(
        _require_key(fixed_values, "isClaimed", "fixedValues."),
        bool,
        "fixedValues.isClaimed",
    )

    skip_blank_rows = _require_type(
        _require_key(behavior_cfg, "skipBlankRows", "behavior."),
        bool,
        "behavior.skipBlankRows",
    )
    row_error_policy = _validate_policy(
        _require_type(_require_key(behavior_cfg, "rowErrorPolicy", "behavior."), str, "behavior.rowErrorPolicy"),
        "behavior.rowErrorPolicy",
        ("error", "skip"),
    )

    if row_start <= 0:
        raise ConfigError("input.row_start must be > 0")
    if chunk_size <= 0:
        raise ConfigError("output.chunkSize must be > 0")
    if split_delimiter == "":
        raise ConfigError("parsing.splitDelimiter cannot be empty")
    if status_value not in (1, 2, 3):
        raise ConfigError("fixedValues.status must be 1/2/3")
    if is_claimed_value:
        raise ConfigError("fixedValues.isClaimed must be false for unclaimed rows")
    try:
        re.compile(id_pattern)
    except re.error as exc:
        raise ConfigError(f"Invalid parsing.idPattern: {exc}") from exc

    return {
        "base_dir": base_dir,
        "input_path": input_path,
        "sheet": sheet,
        "row_start": row_start,
        "row_end": row_end,
        "fid_col": fid_col,
        "split_part_col": split_part_col,
        "source_col": source_col,
        "translated_col": translated_col,
        "split_delimiter": split_delimiter,
        "id_pattern": id_pattern,
        "output_path": output_path,
        "table": table,
        "chunk_size": chunk_size,
        "overwrite": overwrite,
        "output_columns": output_columns,
        "status_value": status_value,
        "is_claimed_value": is_claimed_value,
        "skip_blank_rows": skip_blank_rows,
        "row_error_policy": row_error_policy,
    }


def _build_patterns(id_pattern: str) -> Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]]:
    pattern_colon6 = re.compile(rf"^(?P<textId>{id_pattern})::::::\[(?P<text>.*)\]$", re.DOTALL)
    pattern_triple_colon_num = re.compile(rf"^(?P<textId>{id_pattern}):::\d+:::\[(?P<text>.*)\]$", re.DOTALL)
    pattern_triple_colon_range = re.compile(
        rf"^(?P<textId>{id_pattern}):::\d+(?:-\d+)+:::\[(?P<text>.*)\]$",
        re.DOTALL,
    )
    return pattern_colon6, pattern_triple_colon_num, pattern_triple_colon_range


def _parse_segment(
    segment: str,
    patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
) -> Optional[Tuple[int, str]]:
    pattern_colon6, pattern_triple_colon_num, pattern_triple_colon_range = patterns
    matched = (
        pattern_colon6.fullmatch(segment)
        or pattern_triple_colon_num.fullmatch(segment)
        or pattern_triple_colon_range.fullmatch(segment)
    )
    if matched is None:
        return None
    return int(matched.group("textId")), matched.group("text")


def _is_blank_row(fid: str, source_text: str, translated_text: str) -> bool:
    return fid.strip() == "" and source_text.strip() == "" and translated_text.strip() == ""


def _parse_cell_segments(
    raw_text: str,
    split_delimiter: str,
    patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
    row_index: int,
    column_name: str,
) -> List[Tuple[int, str]]:
    if raw_text.strip() == "":
        raise RowParseError(f"Row {row_index} {column_name} is empty")
    pieces = raw_text.split(split_delimiter)
    if len(pieces) == 0:
        raise RowParseError(f"Row {row_index} {column_name} has no segment")

    parsed: List[Tuple[int, str]] = []
    for idx, piece in enumerate(pieces, start=1):
        segment = piece.strip()
        if segment == "":
            raise RowParseError(f"Row {row_index} {column_name} segment #{idx} is empty")
        extracted = _parse_segment(segment, patterns)
        if extracted is None:
            snippet = segment.replace("\n", "\\n")
            if len(snippet) > 120:
                snippet = snippet[:117] + "..."
            raise RowParseError(
                f"Row {row_index} {column_name} segment #{idx} format invalid: {snippet}"
            )
        parsed.append(extracted)
    return parsed


def _collect_merged_fid_rows(
    worksheet,
    row_start: int,
    row_end: int,
    fid_idx: int,
    split_part_idx: int,
    source_idx: int,
    translated_idx: int,
    skip_blank_rows: bool,
    row_error_policy: str,
) -> Tuple[List[Tuple[str, str, str, int]], int, int]:
    max_idx = max(fid_idx, split_part_idx, source_idx, translated_idx)
    fid_order: List[str] = []
    grouped: Dict[str, Dict[int, Tuple[int, str, str]]] = {}
    skipped_blank_rows = 0
    skipped_error_rows = 0

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True),
        start=row_start,
    ):
        padded = list(row)
        if len(padded) <= max_idx:
            padded.extend([None] * (max_idx + 1 - len(padded)))

        fid_text = _normalize_fid(padded[fid_idx])
        source_raw = _normalize_text_cell(padded[source_idx])
        translated_raw = _normalize_text_cell(padded[translated_idx])

        if skip_blank_rows and _is_blank_row(fid_text, source_raw, translated_raw):
            skipped_blank_rows += 1
            continue

        try:
            if fid_text.strip() == "":
                raise RowParseError(f"Row {row_index} fid is empty")
            split_part_value = _normalize_split_part(padded[split_part_idx], row_index)
            if fid_text not in grouped:
                grouped[fid_text] = {}
                fid_order.append(fid_text)
            part_map = grouped[fid_text]
            if split_part_value in part_map:
                existing_row = part_map[split_part_value][0]
                raise RowParseError(
                    f"Row {row_index} duplicate splitPart for fid={fid_text}: {split_part_value} "
                    f"(already seen at row {existing_row})"
                )
            part_map[split_part_value] = (row_index, source_raw, translated_raw)
        except RowParseError as exc:
            if row_error_policy == "skip":
                skipped_error_rows += 1
                print(f"[SKIP] {exc}")
                continue
            raise

    merged_rows: List[Tuple[str, str, str, int]] = []
    for fid in fid_order:
        part_map = grouped[fid]
        ordered_parts = sorted(part_map.keys())
        source_chunks: List[str] = []
        translated_chunks: List[str] = []
        first_row = part_map[ordered_parts[0]][0]
        for part in ordered_parts:
            _, source_piece, translated_piece = part_map[part]
            source_chunks.append(source_piece)
            translated_chunks.append(translated_piece)
        merged_rows.append((fid, "".join(source_chunks), "".join(translated_chunks), first_row))

    return merged_rows, skipped_blank_rows, skipped_error_rows


def _build_output_rows_for_excel_row(
    fid: str,
    source_raw: str,
    translated_raw: str,
    split_delimiter: str,
    patterns: Tuple[re.Pattern[str], re.Pattern[str], re.Pattern[str]],
    row_index: int,
    status_value: int,
    is_claimed_value: bool,
) -> List[List[str]]:
    source_segments = _parse_cell_segments(source_raw, split_delimiter, patterns, row_index, "sourceText")
    translated_segments = _parse_cell_segments(translated_raw, split_delimiter, patterns, row_index, "translatedText")

    if len(source_segments) != len(translated_segments):
        raise RowParseError(
            f"Row {row_index} segment count mismatch: source={len(source_segments)}, translated={len(translated_segments)}"
        )

    rows: List[List[str]] = []
    for idx, (source_item, translated_item) in enumerate(zip(source_segments, translated_segments), start=1):
        source_text_id, source_text = source_item
        translated_text_id, translated_text = translated_item
        if source_text_id != translated_text_id:
            raise RowParseError(
                f"Row {row_index} segment #{idx} textId mismatch: source={source_text_id}, translated={translated_text_id}"
            )
        source_hash = hashlib.sha256(source_text.encode("utf-8")).hexdigest()
        row_values = {
            "fid": fid,
            "part": idx,
            "textId": source_text_id,
            "sourceText": source_text,
            "sourceTextHash": source_hash,
            "translatedText": translated_text,
            "status": status_value,
            "isClaimed": is_claimed_value,
        }
        rows.append(
            [
                _sql_literal(row_values["fid"]),
                _sql_literal(row_values["part"]),
                _sql_literal(row_values["textId"]),
                _sql_literal(row_values["sourceText"]),
                _sql_literal(row_values["sourceTextHash"]),
                _sql_literal(row_values["translatedText"]),
                _sql_literal(row_values["status"]),
                _sql_literal(row_values["isClaimed"]),
            ]
        )
    return rows


def _write_insert(handle, table: str, output_columns: Dict[str, str], rows: List[List[str]]) -> None:
    if not rows:
        return
    ordered_column_names = [
        output_columns["fid"],
        output_columns["part"],
        output_columns["textId"],
        output_columns["sourceText"],
        output_columns["sourceTextHash"],
        output_columns["translatedText"],
        output_columns["status"],
        output_columns["isClaimed"],
    ]
    table_sql = _sql_qualified_identifier(table)
    columns_sql = ", ".join(_sql_identifier(name) for name in ordered_column_names)
    handle.write(f"INSERT INTO {table_sql} ({columns_sql}) VALUES\n")
    for idx, row in enumerate(rows):
        suffix = ",\n" if idx < len(rows) - 1 else ";\n"
        handle.write("(" + ", ".join(row) + ")" + suffix)


def main() -> None:
    parser = argparse.ArgumentParser(description="Split fixed-format xlsx text and generate INSERT SQL")
    parser.add_argument("--config", required=True, help="YAML config path")
    parser.add_argument("--row-range", help="Override rows with m-n")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser().resolve()
    config = _validate_config(_load_config(config_path))
    base_dir = _resolve_base_dir(config["base_dir"], config_path)

    input_path = Path(config["input_path"]).expanduser()
    if not input_path.is_absolute():
        input_path = base_dir / input_path
    input_path = input_path.resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_path = Path(config["output_path"]).expanduser()
    if not output_path.is_absolute():
        output_path = base_dir / output_path
    output_path = output_path.resolve()
    if output_path.exists() and not config["overwrite"]:
        raise RuntimeError(f"Output file exists and overwrite=false: {output_path}")
    if not output_path.parent.exists():
        raise FileNotFoundError(f"Output directory not found: {output_path.parent}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if config["sheet"] not in workbook.sheetnames:
        raise ConfigError(f"Worksheet not found: {config['sheet']}")
    worksheet = workbook[config["sheet"]]

    row_start = config["row_start"]
    row_end_cfg = config["row_end"]
    if args.row_range:
        row_start, row_end_cfg = _parse_row_range(args.row_range)
    row_end = _parse_row_end(row_end_cfg, worksheet.max_row)
    if row_end > worksheet.max_row:
        row_end = worksheet.max_row
    if row_end < row_start:
        raise ConfigError("input.row_end must be >= input.row_start")

    fid_idx = column_index_from_string(config["fid_col"]) - 1
    split_part_idx = column_index_from_string(config["split_part_col"]) - 1
    source_idx = column_index_from_string(config["source_col"]) - 1
    translated_idx = column_index_from_string(config["translated_col"]) - 1
    patterns = _build_patterns(config["id_pattern"])

    merged_fid_rows, skipped_blank_rows, skipped_error_rows = _collect_merged_fid_rows(
        worksheet=worksheet,
        row_start=row_start,
        row_end=row_end,
        fid_idx=fid_idx,
        split_part_idx=split_part_idx,
        source_idx=source_idx,
        translated_idx=translated_idx,
        skip_blank_rows=config["skip_blank_rows"],
        row_error_policy=config["row_error_policy"],
    )

    row_buffer: List[List[str]] = []
    generated_rows = 0
    skipped_fid_rows = 0

    with output_path.open("w", encoding="utf-8") as handle:
        handle.write("-- Auto-generated by xlsx_to_insert_segmented.py\n")
        for fid_text, source_raw, translated_raw, row_index in merged_fid_rows:
            try:
                output_rows = _build_output_rows_for_excel_row(
                    fid=fid_text,
                    source_raw=source_raw,
                    translated_raw=translated_raw,
                    split_delimiter=config["split_delimiter"],
                    patterns=patterns,
                    row_index=row_index,
                    status_value=config["status_value"],
                    is_claimed_value=config["is_claimed_value"],
                )
            except RowParseError as exc:
                if config["row_error_policy"] == "skip":
                    skipped_fid_rows += 1
                    print(f"[SKIP] {exc}")
                    continue
                raise

            for generated in output_rows:
                row_buffer.append(generated)
                generated_rows += 1
                if len(row_buffer) >= config["chunk_size"]:
                    _write_insert(handle, config["table"], config["output_columns"], row_buffer)
                    row_buffer = []

        if row_buffer:
            _write_insert(handle, config["table"], config["output_columns"], row_buffer)

    print(f"[DONE] Merged fid rows: {len(merged_fid_rows)}")
    print(f"[DONE] Generated rows: {generated_rows}")
    print(f"[SKIP] Blank rows: {skipped_blank_rows}")
    print(f"[SKIP] Error rows: {skipped_error_rows}")
    print(f"[SKIP] Error fid rows: {skipped_fid_rows}")
    print(f"[OUTPUT] {output_path}")


if __name__ == "__main__":
    main()
