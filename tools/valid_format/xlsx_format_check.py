# xlsx 固定格式计数脚本，用于检查文本中的固定格式数量。
# 支持按列/行范围读取并输出计数结果，可选对比两列数量是否一致。

import argparse
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string


_ID_PATTERN = r"\d{2,10}"
_PATTERN_COLON6 = re.compile(rf"^{_ID_PATTERN}::::::\[.*\]$", re.DOTALL)
_PATTERN_TRIPLE_COLON_NUM = re.compile(rf"^{_ID_PATTERN}:::\d+:::\[.*\]$", re.DOTALL)
_PATTERN_TRIPLE_COLON_RANGE = re.compile(
    rf"^{_ID_PATTERN}:::\d+(?:-\d+)+:::\[.*\]$", re.DOTALL
)


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _split_segments(text: str) -> List[str]:
    if not text:
        return []
    return text.split("|||")


def _match_fixed_format(segment: str) -> bool:
    return bool(
        _PATTERN_COLON6.fullmatch(segment)
        or _PATTERN_TRIPLE_COLON_RANGE.fullmatch(segment)
        or _PATTERN_TRIPLE_COLON_NUM.fullmatch(segment)
    )


def count_fixed_formats(text: str) -> Tuple[int, List[str]]:
    segments = _split_segments(text)
    count = 0
    invalid_segments: List[str] = []
    for raw in segments:
        segment = raw.strip()
        if not segment:
            invalid_segments.append(segment)
            continue
        if _match_fixed_format(segment):
            count += 1
        else:
            invalid_segments.append(segment)
    return count, invalid_segments


def _analyze_text(text: str) -> Tuple[int, List[Tuple[int, str]]]:
    segments = _split_segments(text)
    count = 0
    invalid_segments: List[Tuple[int, str]] = []
    for idx, raw in enumerate(segments, start=1):
        segment = raw.strip()
        if not segment:
            invalid_segments.append((idx, segment))
            continue
        if _match_fixed_format(segment):
            count += 1
        else:
            invalid_segments.append((idx, segment))
    return count, invalid_segments


def _summarize_segment(text: str, max_len: int) -> str:
    if text == "":
        return "<EMPTY>"
    snippet = text.replace("\n", "\\n")
    if len(snippet) <= max_len:
        return snippet
    return snippet[: max_len - 3] + "..."


def _iter_row_counts(
    worksheet,
    column_index: int,
    row_start: int,
    row_end: int,
) -> Iterable[Tuple[int, int, int]]:
    for row_index in range(row_start, row_end + 1):
        cell_value = worksheet.cell(row=row_index, column=column_index).value
        text = _normalize_cell(cell_value)
        # print(text)
        count, invalid = count_fixed_formats(text)
        yield row_index, count, len(invalid)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="xlsx 固定格式计数脚本")
    parser.add_argument(
        "--path", default="work_text/text_work.xlsx", help="xlsx 文件路径"
    )
    parser.add_argument("--sheet", default=None, help="工作表名称，默认使用活动表")
    parser.add_argument("--column", default="C", help="待解析列（如 C）")
    parser.add_argument(
        "--row-start", type=int, default=2, help="起始行号（从 2 开始,不考虑标题）"
    )
    parser.add_argument("--row-end", type=int, default=100, help="结束行号（包含）")
    parser.add_argument("--compare-column", default=None, help="对比列（如 D）")
    parser.add_argument(
        "--only-mismatch", action="store_true", help="仅输出数量不一致的行"
    )
    parser.add_argument(
        "--show-mismatch", action="store_true", help="输出不一致行的附近文本片段"
    )
    parser.add_argument(
        "--snippet-len", type=int, default=120, help="片段输出最大长度"
    )
    parser.add_argument(
        "--max-items", type=int, default=20, help="每列最多输出的可疑片段数"
    )
    parser.add_argument(
        "--output-mismatch-xlsx",
        default=None,
        help="输出不一致行的 xlsx 文件路径（需搭配 --compare-column）",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=8,
        help="并发线程数（默认 8）",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=10000,
        help="进度输出间隔行数（0 表示不输出）",
    )
    return parser.parse_args()


def _resolve_sheet(workbook, sheet_name: Optional[str]):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        return workbook[sheet_name]
    return workbook.active


def _print_single_column(rows: Iterable[Tuple[int, int, int]]) -> None:
    print("row\tcount\tinvalid")
    for row_index, count, invalid_count in rows:
        print(f"{row_index}\t{count}\t{invalid_count}")


def _print_compare_columns(
    left_rows: List[Tuple[int, int, int]],
    right_rows: List[Tuple[int, int, int]],
    only_mismatch: bool,
) -> None:
    print("row\tleft\tleft_invalid\tright\tright_invalid\tmatch")
    for (row_index, left_count, left_invalid), (_, right_count, right_invalid) in zip(
        left_rows, right_rows
    ):
        match = left_count == right_count
        if only_mismatch and match:
            continue
        flag = "OK" if match else "DIFF"
        print(
            f"{row_index}\t{left_count}\t{left_invalid}\t{right_count}\t{right_invalid}\t{flag}"
        )


def _print_mismatch_detail(
    row_index: int,
    left_invalid: List[Tuple[int, str]],
    right_invalid: List[Tuple[int, str]],
    snippet_len: int,
    max_items: int,
) -> None:
    print(f"--- row {row_index} ---")
    if left_invalid:
        print("L#\tsegment")
        for idx, segment in left_invalid[:max_items]:
            print(f"{idx}\t{_summarize_segment(segment, snippet_len)}")
    if right_invalid:
        print("R#\tsegment")
        for idx, segment in right_invalid[:max_items]:
            print(f"{idx}\t{_summarize_segment(segment, snippet_len)}")


def _format_mismatch_detail(
    left_invalid: List[Tuple[int, str]],
    right_invalid: List[Tuple[int, str]],
    snippet_len: int,
    max_items: int,
) -> str:
    lines: List[str] = []
    if left_invalid:
        for idx, segment in left_invalid[:max_items]:
            lines.append(f"L#{idx}: {_summarize_segment(segment, snippet_len)}")
    if right_invalid:
        for idx, segment in right_invalid[:max_items]:
            lines.append(f"R#{idx}: {_summarize_segment(segment, snippet_len)}")
    return "\n".join(lines)


def _write_mismatch_xlsx(
    path: Path,
    rows: List[Tuple[int, str, str, int, int, int, int, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "mismatch"
    worksheet.append(
        [
            "row",
            "fid",
            "part",
            "left_count",
            "right_count",
            "left_invalid",
            "right_invalid",
            "mismatch_detail",
        ]
    )
    for row in rows:
        worksheet.append(list(row))
    workbook.save(path)


def _progress_print(current: int, total: int) -> None:
    percent = 0.0
    if total > 0:
        percent = current * 100.0 / total
    print(f"Progress: {current}/{total} ({percent:.2f}%)", file=sys.stderr)


def _analyze_single_row(
    row_index: int,
    left_text: str,
    max_items: int,
) -> Tuple[int, int, int]:
    left_count, left_invalid = _analyze_text(left_text)
    return row_index, left_count, len(left_invalid)


def _analyze_compare_row(
    row_index: int,
    a_value: str,
    b_value: str,
    left_text: str,
    right_text: str,
    max_items: int,
) -> Tuple[
    int,
    str,
    str,
    int,
    int,
    List[Tuple[int, str]],
    List[Tuple[int, str]],
]:
    left_count, left_invalid = _analyze_text(left_text)
    right_count, right_invalid = _analyze_text(right_text)
    return (
        row_index,
        a_value,
        b_value,
        left_count,
        right_count,
        left_invalid,
        right_invalid,
    )


def main() -> None:
    args = _parse_args()
    path = Path(args.path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")
    if args.row_start <= 0:
        raise ValueError("row-start 必须大于 0")
    if args.row_end < args.row_start:
        raise ValueError("row-end 必须大于等于 row-start")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = _resolve_sheet(workbook, args.sheet)

    row_end = min(args.row_end, worksheet.max_row or args.row_end)
    if row_end < args.row_start:
        raise ValueError("row-start 超出工作表范围")
    total_rows = row_end - args.row_start + 1

    column_index = column_index_from_string(args.column)
    compare_index = (
        column_index_from_string(args.compare_column) if args.compare_column else None
    )
    max_col = max(column_index, compare_index or 1, 2)

    rows_iter = worksheet.iter_rows(
        min_row=args.row_start,
        max_row=row_end,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )

    if not args.compare_column:
        print("row\tcount\tinvalid")
        def _single_task_iter():
            for offset, row_values in enumerate(rows_iter, start=0):
                row_index = args.row_start + offset
                left_value = row_values[column_index - 1]
                left_text = _normalize_cell(left_value)
                yield row_index, left_text

        completed = 0
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
            for row_index, count, invalid_count in executor.map(
                lambda args_tuple: _analyze_single_row(*args_tuple),
                (
                    (row_index, left_text, args.max_items)
                    for row_index, left_text in _single_task_iter()
                ),
                chunksize=200,
            ):
                print(f"{row_index}\t{count}\t{invalid_count}")
                completed += 1
                if args.progress_every and completed % args.progress_every == 0:
                    _progress_print(completed, total_rows)
        return

    print("row\tleft\tleft_invalid\tright\tright_invalid\tmatch")
    mismatch_rows: List[Tuple[int, str, str, int, int, int, int, str]] = []

    def _compare_task_iter():
        for offset, row_values in enumerate(rows_iter, start=0):
            row_index = args.row_start + offset
            left_value = row_values[column_index - 1]
            right_value = row_values[compare_index - 1] if compare_index else None
            left_text = _normalize_cell(left_value)
            right_text = _normalize_cell(right_value)
            a_value = _normalize_cell(row_values[0])
            b_value = _normalize_cell(row_values[1] if len(row_values) > 1 else "")
            yield row_index, a_value, b_value, left_text, right_text

    completed = 0
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        for (
            row_index,
            a_value,
            b_value,
            left_count,
            right_count,
            left_invalid,
            right_invalid,
            left_contexts,
            right_contexts,
        ) in executor.map(
            lambda args_tuple: _analyze_compare_row(*args_tuple),
            (
                (
                    row_index,
                    a_value,
                    b_value,
                    left_text,
                    right_text,
                    args.max_items,
                )
                for row_index, a_value, b_value, left_text, right_text in _compare_task_iter()
            ),
            chunksize=200,
        ):
            match = left_count == right_count
            if not (args.only_mismatch and match):
                flag = "OK" if match else "DIFF"
                print(
                    f"{row_index}\t{left_count}\t{len(left_invalid)}\t"
                    f"{right_count}\t{len(right_invalid)}\t{flag}"
                )
            if not match:
                if args.show_mismatch:
                    _print_mismatch_detail(
                        row_index,
                        left_invalid,
                        right_invalid,
                        args.snippet_len,
                        args.max_items,
                    )
                if args.output_mismatch_xlsx:
                    detail = _format_mismatch_detail(
                        left_invalid,
                        right_invalid,
                        args.snippet_len,
                        args.max_items,
                    )
                    mismatch_rows.append(
                        (
                            row_index,
                            a_value,
                            b_value,
                            left_count,
                            right_count,
                            len(left_invalid),
                            len(right_invalid),
                            detail,
                        )
                    )
            completed += 1
            if args.progress_every and completed % args.progress_every == 0:
                _progress_print(completed, total_rows)

    if args.output_mismatch_xlsx:
        output_path = Path(args.output_mismatch_xlsx)
        _write_mismatch_xlsx(output_path, mismatch_rows)


if __name__ == "__main__":
    main()
    # 数字::::::[]
    #  /Users/zhaolei/miniconda3/envs/lotro/bin/python /Users/zhaolei/My/my-python/lotro_chinese_translation/tools/xlsx_format_check.py --compare-column D --row-start 2 --row-end 1000000  --show-mismatch --output-mismatch-xlsx mismatch.xlsx --workers 8 --progress-every 10000 
    
    
