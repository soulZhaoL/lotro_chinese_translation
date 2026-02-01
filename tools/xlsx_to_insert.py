import argparse
from pathlib import Path
from typing import Any, Dict, List, Tuple

import yaml
from openpyxl import load_workbook


class ConfigError(Exception):
    pass


def _require_key(obj: Dict[str, Any], key: str, path: str) -> Any:
    if key not in obj:
        raise ConfigError(f"缺少配置项: {path}{key}")
    return obj[key]


def _require_type(value: Any, expected_type: type, path: str) -> Any:
    if not isinstance(value, expected_type):
        raise ConfigError(f"配置项类型错误: {path} 期望 {expected_type.__name__}")
    return value


def _require_list_of_str(value: Any, path: str) -> List[str]:
    value = _require_type(value, list, path)
    for item in value:
        if not isinstance(item, str):
            raise ConfigError(f"配置项类型错误: {path} 期望 str 列表")
    return value


def _require_list_of_int(value: Any, path: str) -> List[int]:
    value = _require_type(value, list, path)
    for item in value:
        if not isinstance(item, int):
            raise ConfigError(f"配置项类型错误: {path} 期望 int 列表")
    return value


def _parse_row_end(value: Any, max_row: int) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().lower() == "max":
        return max_row
    raise ConfigError("row_end 必须为整数或 'max'")


def _parse_row_range(value: str) -> Tuple[int, int]:
    parts = value.split("-", 1)
    if len(parts) != 2:
        raise ConfigError("row_range 格式必须为 m-n")
    start_text, end_text = parts[0].strip(), parts[1].strip()
    if not start_text.isdigit() or not end_text.isdigit():
        raise ConfigError("row_range 必须为数字范围")
    row_start = int(start_text)
    row_end = int(end_text)
    if row_start <= 0:
        raise ConfigError("row_range 起始行必须大于 0")
    if row_end < row_start:
        raise ConfigError("row_range 结束行必须大于等于起始行")
    return row_start, row_end


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    return str(value)


def _sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("'", "''")
    return "'" + text + "'"


def _build_output_row(
    row_values: Tuple[Any, ...],
    output_columns: List[str],
    source_map: Dict[str, int],
    fixed_values: Dict[str, Any],
    nullable_columns: List[str],
    on_missing: str,
    row_index: int,
) -> Tuple[bool, List[str]]:
    values: List[str] = []
    for col in output_columns:
        if col in source_map:
            source_index = source_map[col]
            if source_index < 0 or source_index >= len(row_values):
                raise ConfigError(f"源列索引越界: {col} -> {source_index + 1}")
            raw_value = _normalize_cell(row_values[source_index])
            if isinstance(raw_value, str) and raw_value.strip() == "":
                if col in nullable_columns:
                    raw_value = None
            if raw_value is None and col not in nullable_columns:
                if on_missing == "skip":
                    return False, []
                raise ConfigError(f"缺少必填列数据: xlsx 行{row_index} 列{col}")
            values.append(_sql_literal(raw_value))
            continue
        if col in fixed_values:
            values.append(_sql_literal(fixed_values[col]))
            continue
        raise ConfigError(f"输出列未映射: {col}")
    return True, values


def _load_config(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"配置文件不存在: {path}")
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ConfigError("配置文件根对象必须为 YAML 对象")
    return data


def _find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (candidate / "config" / "lotro.yaml").exists() or (candidate / ".git").exists():
            return candidate
    raise ConfigError("无法自动定位项目根目录，请在 base_dir 中填写绝对路径")


def _resolve_base_dir(base_dir: str, config_path: Path) -> Path:
    if base_dir == "__PROJECT_ROOT__":
        return _find_project_root(config_path.parent)
    return Path(base_dir).expanduser().resolve()


def _validate_config(data: Dict[str, Any]) -> Dict[str, Any]:
    base_dir = _require_type(_require_key(data, "base_dir", ""), str, "base_dir")
    input_cfg = _require_type(_require_key(data, "input", ""), dict, "input")
    output_cfg = _require_type(_require_key(data, "output", ""), dict, "output")
    behavior_cfg = _require_type(_require_key(data, "behavior", ""), dict, "behavior")

    input_path = _require_type(_require_key(input_cfg, "path", "input."), str, "input.path")
    sheet = _require_type(_require_key(input_cfg, "sheet", "input."), str, "input.sheet")
    row_start = _require_type(_require_key(input_cfg, "row_start", "input."), int, "input.row_start")
    row_end = _require_key(input_cfg, "row_end", "input.")

    source_columns = _require_type(_require_key(input_cfg, "source_columns", "input."), dict, "input.source_columns")
    source_indexes = {}
    for key, value in source_columns.items():
        if not isinstance(key, str):
            raise ConfigError("input.source_columns 的键必须为字符串")
        if not isinstance(value, int):
            raise ConfigError("input.source_columns 的值必须为整数")
        source_indexes[key] = value

    output_path = _require_type(_require_key(output_cfg, "path", "output."), str, "output.path")
    table = _require_type(_require_key(output_cfg, "table", "output."), str, "output.table")
    columns = _require_list_of_str(_require_key(output_cfg, "columns", "output."), "output.columns")
    chunk_size = _require_type(_require_key(output_cfg, "chunk_size", "output."), int, "output.chunk_size")
    fixed_values = _require_type(_require_key(output_cfg, "fixed_values", "output."), dict, "output.fixed_values")
    if "status" in fixed_values:
        status_value = fixed_values["status"]
        if not isinstance(status_value, int):
            raise ConfigError("output.fixed_values.status 必须为整数 1/2/3")
        if status_value not in (1, 2, 3):
            raise ConfigError("output.fixed_values.status 仅支持 1/2/3")

    on_missing = _require_type(_require_key(behavior_cfg, "on_missing", "behavior."), str, "behavior.on_missing")
    if on_missing not in ("error", "skip"):
        raise ConfigError("behavior.on_missing 只能是 error 或 skip")
    nullable_columns = _require_list_of_str(
        _require_key(behavior_cfg, "nullable_columns", "behavior."),
        "behavior.nullable_columns",
    )
    skip_blank_rows = _require_type(
        _require_key(behavior_cfg, "skip_blank_rows", "behavior."),
        bool,
        "behavior.skip_blank_rows",
    )

    if row_start <= 0:
        raise ConfigError("input.row_start 必须大于 0")
    if chunk_size <= 0:
        raise ConfigError("output.chunk_size 必须大于 0")

    for column in columns:
        if column in source_indexes and column in fixed_values:
            raise ConfigError(f"输出列重复映射: {column}")
        if column not in source_indexes and column not in fixed_values:
            raise ConfigError(f"输出列未映射: {column}")

    return {
        "base_dir": base_dir,
        "input_path": input_path,
        "sheet": sheet,
        "row_start": row_start,
        "row_end": row_end,
        "source_indexes": source_indexes,
        "output_path": output_path,
        "table": table,
        "columns": columns,
        "chunk_size": chunk_size,
        "fixed_values": fixed_values,
        "on_missing": on_missing,
        "nullable_columns": nullable_columns,
        "skip_blank_rows": skip_blank_rows,
    }


def _is_blank_row(row_values: Tuple[Any, ...]) -> bool:
    for value in row_values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def _write_insert(
    handle,
    table: str,
    columns: List[str],
    rows: List[List[str]],
) -> None:
    if not rows:
        return
    handle.write(f"INSERT INTO {table} ({', '.join(columns)}) VALUES\n")
    for index, row in enumerate(rows):
        suffix = ",\n" if index < len(rows) - 1 else ";\n"
        handle.write("(" + ", ".join(row) + ")" + suffix)


def main() -> None:
    parser = argparse.ArgumentParser(description="xlsx 转 INSERT 语句生成器")
    parser.add_argument("--config", required=True, help="配置文件路径")
    parser.add_argument("--row-range", help="仅处理 m-n 行数据（覆盖配置中的 row_start/row_end）")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser().resolve()
    config = _validate_config(_load_config(config_path))

    base_dir = _resolve_base_dir(config["base_dir"], config_path)
    input_path = Path(config["input_path"]).expanduser()
    if not input_path.is_absolute():
        input_path = base_dir / input_path
    input_path = input_path.resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    output_path = Path(config["output_path"]).expanduser()
    if not output_path.is_absolute():
        output_path = base_dir / output_path
    output_path = output_path.resolve()
    output_dir = output_path.parent
    if not output_dir.exists():
        raise FileNotFoundError(f"输出目录不存在: {output_dir}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    if config["sheet"] not in wb.sheetnames:
        raise ConfigError(f"工作表不存在: {config['sheet']}")
    ws = wb[config["sheet"]]

    row_start = config["row_start"]
    row_end_value = config["row_end"]
    if args.row_range:
        row_start, row_end_value = _parse_row_range(args.row_range)
    row_end = _parse_row_end(row_end_value, ws.max_row)
    if row_end < row_start:
        raise ConfigError("input.row_end 必须大于等于 input.row_start")

    source_indexes = {key: value - 1 for key, value in config["source_indexes"].items()}

    chunk_size = config["chunk_size"]
    rows_buffer: List[List[str]] = []
    written_rows = 0

    with output_path.open("w", encoding="utf-8") as handle:
        handle.write("-- Auto-generated inserts from xlsx\n")
        for row_index, row in enumerate(
            ws.iter_rows(min_row=row_start, max_row=row_end, values_only=True),
            start=row_start,
        ):
            if config["skip_blank_rows"] and _is_blank_row(row):
                continue
            ok, values = _build_output_row(
                row,
                config["columns"],
                source_indexes,
                config["fixed_values"],
                config["nullable_columns"],
                config["on_missing"],
                row_index,
            )
            if not ok:
                continue
            rows_buffer.append(values)
            written_rows += 1
            if len(rows_buffer) >= chunk_size:
                _write_insert(handle, config["table"], config["columns"], rows_buffer)
                rows_buffer = []
        if rows_buffer:
            _write_insert(handle, config["table"], config["columns"], rows_buffer)

    print(f"已生成 INSERT 语句, 共 {written_rows} 行: {output_path}")


if __name__ == "__main__":
    main()

# python tools/xlsx_to_insert.py --config tools/xlsx_to_insert.yaml --row-range 2-20000